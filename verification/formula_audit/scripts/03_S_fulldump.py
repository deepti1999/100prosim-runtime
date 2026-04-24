"""Full cell dump of _S.xlsx scenario sheets.

For each sheet, dump every non-empty cell as `<sheet>!<ref>=<value>`.
Writes one .md per sheet under 01_value_parity/S_dump/.
"""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
SRC = ROOT / "docs" / "100prosim_d_250517_250517.1817m" / "_S.xlsx"
OUT_DIR = ROOT / "verification" / "formula_audit" / "01_value_parity" / "S_dump"
OUT_DIR.mkdir(parents=True, exist_ok=True)

SHEETS = [
    "1. Flächen",
    "2. Erneuerbare",
    "3. Bedarfsniveau",
    "4. Verbrauch",
    "5. Bilanz",
    "6. Fossile",
    "7. Verbrauch Status",
    "8. Kennzahlen",
]

def dump_sheet(wb_v, wb_f, sname, outpath):
    ws_v = wb_v[sname]
    ws_f = wb_f[sname]
    with open(outpath, "w", encoding="utf-8") as f:
        f.write(f"# `{sname}` full cell dump\n\n")
        f.write(f"Dim: {ws_v.max_row} × {ws_v.max_column}\n\n")
        f.write("Row-major. `v=` is cached computed value, `f=` is formula text.\n\n")
        f.write("```\n")
        for r in range(1, ws_v.max_row + 1):
            row_has = False
            for c in range(1, ws_v.max_column + 1):
                v = ws_v.cell(r, c).value
                ftxt = ws_f.cell(r, c).value
                if v is None and ftxt is None:
                    continue
                if not row_has:
                    f.write(f"\n  row {r}:\n")
                    row_has = True
                ref = f"{get_column_letter(c)}{r}"
                vs = repr(v) if v is not None else ""
                if isinstance(ftxt, str) and ftxt.startswith("="):
                    f.write(f"    {ref}: v={vs}  f={ftxt!r}\n")
                else:
                    f.write(f"    {ref}: v={vs}\n")
        f.write("```\n")

def main():
    wb_v = load_workbook(SRC, data_only=True)
    wb_f = load_workbook(SRC, data_only=False)
    for sname in SHEETS:
        safe = sname.replace(" ", "_").replace(".", "_")
        dump_sheet(wb_v, wb_f, sname, OUT_DIR / f"{safe}.md")
        print(f"  wrote {safe}.md")
    print("done")

if __name__ == "__main__":
    main()
