"""§4 Bilanz parity — compare `calculate_bilanz_data()` output to _S.xlsx!5. Bilanz.

Writes:
  03_bilanz_parity/row_by_row.csv
  03_bilanz_parity/discrepancies.md
  03_bilanz_parity/summary.md
"""
from __future__ import annotations
import os, sys, csv, math
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "landuse_project.settings")
import django
django.setup()

from openpyxl import load_workbook
from calculation_engine.bilanz_engine import calculate_bilanz_data

SRC = ROOT / "docs" / "100prosim_d_250517_250517.1817m" / "_S.xlsx"
OUT = ROOT / "verification" / "formula_audit" / "03_bilanz_parity"
OUT.mkdir(parents=True, exist_ok=True)

# Map of Bilanz engine keys → Excel row number on 5. Bilanz (KLIK, GW, PW, MA, total cells)
# _S.xlsx!5. Bilanz columns per row:
#   H = KLIK, K = GW, N = PW, Q = MA, T = total   (for rows 9, 12, 15, 18 — the "Verbrauch" rows)
#   I = KLIK, L = GW, O = PW, R = MA, U = total   (for rows 10, 13, 16, 19 — the "erneuerbar" rows)
#   J = KLIK, M = GW, P = PW, S = MA, V = total   (for rows 11, 14, 17, 20 — the "fossil" rows)
MAP = [
    # (engine_key, label, row_verbrauch, row_erneuerbar, row_fossil)
    ("verbrauch_strom", "Strom", 9, 10, 11),
    ("verbrauch_fuels", "Brennstoff_gas", 12, 13, 14),
    # Note: Excel has separate rows for liquid (15-17) and solid (18-20) fuels;
    # our engine has one "fuels" aggregate. We compare to SUM.
    # (Skip detailed liquid/solid for this pass.)
    ("verbrauch_heat", "Wärme", 27, 28, 29),   # Need to verify row numbers
]

def safe_get(sheet_data, row, col):
    return sheet_data.cell(row=row, column=col).value

def main():
    wb = load_workbook(SRC, data_only=True)
    ws = wb["5. Bilanz"]

    bilanz = calculate_bilanz_data()

    rows = []
    def rel(a, b):
        if a is None or b is None:
            return math.inf
        try:
            fa, fb = float(a), float(b)
        except:
            return math.inf
        if fa == 0 and fb == 0:
            return 0.0
        if max(abs(fa), abs(fb)) == 0:
            return 0.0
        return abs(fa - fb) / max(abs(fa), abs(fb))

    # Column letter → Excel row col index
    # KLIK (status 'H'=8, erneuerbar 'I'=9, fossil 'J'=10)
    # GW   (status 'K'=11, erneuerbar 'L'=12, fossil 'M'=13)
    # PW   (status 'N'=14, erneuerbar 'O'=15, fossil 'P'=16)
    # MA   (status 'Q'=17, erneuerbar 'R'=18, fossil 'S'=19)
    # total(T=20, U=21, V=22)

    # verbrauch_strom rows at 9/10/11
    COLS_STATUS = dict(KLIK="H", GW="K", PW="N", MA="Q", total="T")
    COLS_EN = dict(KLIK="I", GW="L", PW="O", MA="R", total="U")
    COLS_FS = dict(KLIK="J", GW="M", PW="P", MA="S", total="V")

    # For each engine key + excel rows, compare
    def bilanz_vals(key):
        d = bilanz.get(key, {})
        status = d.get("status", {}) or {}
        return dict(
            KLIK=status.get("kraft_licht", 0),
            GW=status.get("gebaeudewaerme", 0),
            PW=status.get("prozesswaerme", 0),
            MA=status.get("mobile", 0),
            total=status.get("gesamt", 0),
        )

    def excel_row(row, cols):
        r = {}
        for sector, col in cols.items():
            r[sector] = ws[f"{col}{row}"].value
        return r

    sections = [
        # (db_engine_key, excel_row, cols, label)
        ("verbrauch_strom", 9, COLS_STATUS, "Verbrauch Strom (total)"),
        ("verbrauch_strom_renewable", 10, COLS_EN, "Verbrauch Strom (renewable)"),
        ("verbrauch_strom_fossil", 11, COLS_FS, "Verbrauch Strom (fossil)"),
    ]
    for (key, row, cols, label) in sections:
        db_vals = bilanz_vals(key)
        xl_vals = excel_row(row, cols)
        for sector in ["KLIK", "GW", "PW", "MA", "total"]:
            rows.append({
                "section": label,
                "engine_key": key,
                "sector": sector,
                "excel_cell": f"{cols[sector]}{row}",
                "db": db_vals[sector],
                "xl": xl_vals[sector],
                "drift": rel(db_vals[sector], xl_vals[sector]),
            })

    # Save CSV
    out_csv = OUT / "row_by_row.csv"
    with open(out_csv, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["section", "engine_key", "sector", "excel_cell", "db", "xl", "drift"])
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print(f"wrote {out_csv} ({len(rows)} rows)")

    # Discrepancies summary
    with open(OUT / "discrepancies.md", "w", encoding="utf-8") as f:
        f.write("# §4 Bilanz Parity — discrepancies\n\n")
        f.write("| section | sector | excel | db | drift | verdict |\n")
        f.write("|---------|--------|-------|-----|-------|---------|\n")
        for r in rows:
            d = r["drift"]
            verdict = "PASS" if d < 0.001 else ("PASS_COSMETIC" if d < 0.01 else "DRIFT")
            f.write(f"| {r['section']} | {r['sector']} | {r['xl']} | {r['db']} | {d:.4f} | {verdict} |\n")
    # Summary counts
    from collections import Counter
    def verdict(d):
        if d < 0.001: return "PASS"
        if d < 0.01: return "PASS_COSMETIC"
        return "DRIFT"
    c = Counter(verdict(r["drift"]) for r in rows)
    with open(OUT / "summary.md", "w", encoding="utf-8") as f:
        f.write("# §4 Bilanz Parity — summary\n\n")
        f.write(f"Total cells compared: {len(rows)}\n\n")
        f.write("Verdict distribution:\n")
        for k, v in sorted(c.items()):
            f.write(f"- {k}: {v}\n")

if __name__ == "__main__":
    main()
