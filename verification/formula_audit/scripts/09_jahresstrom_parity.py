"""§6 Jahresstrom parity.

Compare every numeric value on /annual-electricity/ (via the
`compute_ws_diagram_reference()` function) to the corresponding cell
on WS.xlsm!1.Jahresbilanz_Strom.

Outputs:
  05_jahresstrom_parity/every_diagram_node.csv
  05_jahresstrom_parity/discrepancies.md
  05_jahresstrom_parity/summary.md
"""
from __future__ import annotations
import os, sys, csv, math
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "landuse_project.settings")
import django
django.setup()

from openpyxl import load_workbook
from simulator.signals import compute_ws_diagram_reference

SRC = ROOT / "docs" / "100prosim_d_250517_250517.1817m" / "WS.xlsm"
OUT = ROOT / "verification" / "formula_audit" / "05_jahresstrom_parity"
OUT.mkdir(parents=True, exist_ok=True)

# Mapping: diagram key → excel cell on 1.Jahresbilanz_Strom
MAP = [
    # Source circles (top row)
    ("pv_value",       "E19",  "PV value (Solarstrom)"),
    ("pv_tages",       "E20",  "PV Tagesladungen"),
    ("pv_pct",         "E21",  "PV % share (× 100)"),

    ("wind_value",     "E25",  "Wind value"),
    ("wind_tages",     "E26",  "Wind Tagesladungen"),
    ("wind_pct",       "E27",  "Wind % share (× 100)"),

    ("hydro_value",    "E31",  "Hydro/Laufwasser value"),
    ("hydro_tages",    "E32",  "Hydro Tagesladungen"),
    ("hydro_pct",      "E33",  "Hydro % share (× 100)"),

    ("bio_value",      "E13",  "Biomass / Fossile Brennstoffe"),
    ("bio_tages",      "E14",  "Bio Tagesladungen"),
    ("bio_pct",        "E15",  "Bio % share"),

    # M-circle (top middle)
    ("m_total",        "H25",  "Wind+Solar+konstant (M)"),
    ("ely_branch_value", "H28", "Ely branch value (overshoot O)"),

    # J-circle (splitter)
    ("n_value",        "J25",  "Stromverbr. J-branch"),
    ("flow_n_value_tages", "J26", "J-branch Tagesladungen"),

    # Q-circle (middle)
    ("n_output_branch","L28",  "Einspeich P (= n_output_branch)"),
    ("n_input_branch", "L23",  "Abregelung input (AbregCopy)"),
    ("flow_q_abregelung_tages", "L24", "Abregelung Tagesladungen"),

    # S-circle (demand net)
    ("n_to_right",     "N25",  "Direktverbrauch (N)"),
    ("flow_n_to_right_tages", "N26", "Direktverbr. Tagesladungen"),

    ("final_stromnetz","S25",  "Final Stromnetz (VerbrauchStrom)"),
    ("flow_final_tages","S26", "Final Tagesladungen (365)"),

    # Bottom row — Speicher
    ("h2_offer",       "H36",  "H2 offer (=Einspeich sum capacity)"),
    ("gas_storage",    "L36",  "Gas storage (P sum)"),
    ("flow_gas_storage_tages", "L37", "Gas storage Tagesladungen"),
    ("t_value",        "Q36",  "T value (Ausspeich Rückverstr)"),
    ("flow_t_value_tages", "Q37", "T Tagesladungen"),

    # Storage capacity
    ("storage_capacity","M44", "Speicherkapazität"),
    ("flow_storage_capacity_tages", "M45", "Speicherkapazität Tages"),

    # Abgleich
    ("abgleichdifferenz", "Q44", "Abgleichdifferenz"),
]

def rel(a, b):
    if a is None or b is None:
        return math.inf
    try:
        fa, fb = float(a), float(b)
    except:
        return math.inf
    if fa == 0 and fb == 0:
        return 0.0
    m = max(abs(fa), abs(fb))
    if m < 1e-9:
        return 0.0
    return abs(fa - fb) / m

def verdict(d):
    if d is None: return "NO_DATA"
    if d == math.inf: return "NO_MATCH"
    if d == 0.0: return "EXACT"
    if d < 0.0001: return "PASS_COSMETIC"
    if d < 0.001: return "PASS"
    if d < 0.01: return "PASS_LOOSE (<1%)"
    return "DRIFT"

def main():
    wb = load_workbook(SRC, data_only=True)
    ws = wb["1.Jahresbilanz_Strom"]
    d = compute_ws_diagram_reference()

    rows = []
    # Special-case pct values (Excel stores as decimal 0.62, we use 62)
    pct_keys = {"pv_pct", "wind_pct", "hydro_pct", "bio_pct"}
    for (key, cell, desc) in MAP:
        our = d.get(key)
        xl = ws[cell].value
        if key in pct_keys and xl is not None:
            try:
                xl_scaled = float(xl) * 100.0
            except:
                xl_scaled = xl
        else:
            xl_scaled = xl
        drift = rel(our, xl_scaled)
        rows.append({
            "key": key,
            "excel_cell": cell,
            "description": desc,
            "our_value": our,
            "excel_raw": xl,
            "excel_scaled": xl_scaled,
            "drift": drift,
            "verdict": verdict(drift),
        })

    out_csv = OUT / "every_diagram_node.csv"
    with open(out_csv, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["key", "excel_cell", "description", "our_value", "excel_raw", "excel_scaled", "drift", "verdict"])
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print(f"wrote {out_csv} ({len(rows)} rows)")

    # Summary
    from collections import Counter
    cnts = Counter(r["verdict"] for r in rows)
    with open(OUT / "summary.md", "w", encoding="utf-8") as fo:
        fo.write("# §6 Jahresstrom Parity — summary\n\n")
        fo.write(f"Total nodes: {len(rows)}\n\n")
        fo.write("Verdict distribution:\n")
        for k, v in sorted(cnts.items()):
            fo.write(f"- {k}: {v}\n")

    # Discrepancies
    with open(OUT / "discrepancies.md", "w", encoding="utf-8") as fo:
        fo.write("# §6 Jahresstrom Parity — node-by-node\n\n")
        fo.write("| verdict | our_value | excel_scaled | drift | key | cell | description |\n")
        fo.write("|---------|-----------|--------------|-------|-----|------|-------------|\n")
        for r in rows:
            drift_s = f"{r['drift']:.4f}" if r['drift'] != math.inf else "∞"
            fo.write(f"| {r['verdict']} | {r['our_value']} | {r['excel_scaled']} | {drift_s} | {r['key']} | {r['excel_cell']} | {r['description']} |\n")

if __name__ == "__main__":
    main()
