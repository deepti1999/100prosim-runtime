"""§2 Value parity — compare every DB parameter row to _S.xlsx.

Approach:
  1. Build Excel row index per sheet (sheet → list of rows with
     full_name, L_value, M_value, D_ref, source tag).
  2. Load DB rows (LandUse, VerbrauchData, RenewableData, GebaeudewaermeData).
  3. Match DB row → Excel row by normalized name. For LandUse, also
     cross-check via D_ref column when name is ambiguous.
  4. For each match, compare at tolerance 0.1% (and 0.01% for the
     cosmetic layer). Try scale factors {1, 1000, 1e-3, 1e4, 1e-4}.
  5. Emit per-row CSV + discrepancies.md + summary.md.

OUTPUTS
  verification/formula_audit/01_value_parity/per_row_comparison.csv
  verification/formula_audit/01_value_parity/discrepancies.md
  verification/formula_audit/01_value_parity/summary.md
"""
from __future__ import annotations
import os, sys, csv, json, math, re, unicodedata
from pathlib import Path
from collections import defaultdict

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "landuse_project.settings")

import django
django.setup()

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SRC = ROOT / "docs" / "100prosim_d_250517_250517.1817m" / "_S.xlsx"
OUT_DIR = ROOT / "verification" / "formula_audit" / "01_value_parity"
OUT_CSV = OUT_DIR / "per_row_comparison.csv"
OUT_DISC = OUT_DIR / "discrepancies.md"
OUT_SUMM = OUT_DIR / "summary.md"

# --- Normalization ----------------------------------------------------

def norm_name(s: str) -> str:
    """Normalize a German name for fuzzy matching."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = s.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    s = s.replace("Ä", "ae").replace("Ö", "oe").replace("Ü", "ue")
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

TOL_REL = 0.001          # 0.1 %
TOL_REL_COSMETIC = 0.0001  # 0.01 %
TOL_ABS_SMALL = 0.01     # for |value| < 1

def drift(a, b):
    """Return relative drift (|a-b|/max(|a|,|b|)) or abs if near zero. None if both None."""
    if a is None and b is None:
        return None
    if a is None or b is None:
        return math.inf
    try:
        fa, fb = float(a), float(b)
    except Exception:
        return math.inf
    if math.isclose(fa, fb, rel_tol=0, abs_tol=1e-12):
        return 0.0
    m = max(abs(fa), abs(fb))
    if m < 1.0:
        return abs(fa - fb)
    return abs(fa - fb) / m

def try_scales(db_val, xl_val):
    """Try multiple scale factors — return (best_scale, best_drift) or (1, drift)."""
    if db_val is None or xl_val is None:
        return (1.0, drift(db_val, xl_val))
    try:
        fa = float(db_val)
    except Exception:
        return (1.0, math.inf)
    best = (1.0, drift(fa, xl_val))
    for sc in (1000.0, 0.001, 10000.0, 0.0001, 100.0, 0.01):
        d = drift(fa * sc, xl_val)
        if d < best[1]:
            best = (sc, d)
    return best

# --- Build Excel row index --------------------------------------------

SHEET_LAYOUTS = {
    # sheet_name: {
    #   "row_range": (start, end),
    #   "status_col": "I" or "L",
    #   "ziel_col": "L" or "M",
    #   "name_cols": list of cols for hierarchical name,
    #   "full_name_col": concatenated name column (AZ/AQ/AL),
    #   "dref_col": D reference column (e.g. AG for Flächen),
    # }
    "1. Flächen": {
        "row_range": (5, 35),
        "status_col": "I",
        "ziel_col": "L",
        "unit_col": "I",   # unit header "ha" at row 6
        "name_cols": ["E", "F", "G", "H"],
        "full_name_col": "AL",
        "dref_col": "AG",
    },
    "2. Erneuerbare": {
        "row_range": (5, 230),
        "status_col": "L",
        "ziel_col": "M",
        "unit_col": "K",
        "name_cols": ["E", "F", "G", "H", "I", "J"],
        "full_name_col": "AZ",
        "dref_col": "AV",
    },
    "3. Bedarfsniveau": {
        "row_range": (5, 53),
        "status_col": "L",
        "ziel_col": "M",
        "unit_col": "K",
        "name_cols": ["E", "F", "G", "H"],
        "full_name_col": None,  # need to build
        "dref_col": "BA",
    },
    "4. Verbrauch": {
        "row_range": (5, 213),
        "status_col": "L",
        "ziel_col": "M",
        "unit_col": "K",
        "name_cols": ["E", "F", "G", "H", "I", "J"],
        "full_name_col": "AQ",
        "dref_col": "AM",
    },
    "6. Fossile": {
        "row_range": (5, 93),
        "status_col": "L",
        "ziel_col": "M",
        "unit_col": "K",
        "name_cols": ["E", "F", "G", "H", "I", "J"],
        "full_name_col": None,
        "dref_col": None,
    },
}

def build_excel_index():
    """Return dict sheet → list of dicts {row, full_name, norm_name, status, ziel, dref, unit}."""
    wb = load_workbook(SRC, data_only=True)
    idx = {}
    for sname, layout in SHEET_LAYOUTS.items():
        ws = wb[sname]
        rows = []
        r0, r1 = layout["row_range"]
        status_col = layout["status_col"]
        ziel_col = layout["ziel_col"]
        for r in range(r0, min(r1, ws.max_row) + 1):
            status = ws[f"{status_col}{r}"].value
            ziel = ws[f"{ziel_col}{r}"].value
            # Build full name from hierarchy columns
            if layout["full_name_col"]:
                full = ws[f"{layout['full_name_col']}{r}"].value
            else:
                parts = []
                for col in layout["name_cols"]:
                    v = ws[f"{col}{r}"].value
                    if v is not None and str(v).strip() != "":
                        parts.append(str(v).strip())
                full = " ".join(parts) if parts else None
            unit = ws[f"{layout['unit_col']}{r}"].value if layout.get("unit_col") else None
            dref = None
            if layout["dref_col"]:
                dref_val = ws[f"{layout['dref_col']}{r}"].value
                if dref_val is not None:
                    dref = str(dref_val).strip()
            if full is None and status is None and ziel is None:
                continue
            rows.append({
                "sheet": sname,
                "row": r,
                "full_name": full,
                "norm_name": norm_name(full),
                "status": status,
                "ziel": ziel,
                "unit": unit,
                "dref": dref,
            })
        idx[sname] = rows
    return idx

# --- DB extraction -----------------------------------------------------

def load_db():
    from simulator.models import (
        LandUse, VerbrauchData, RenewableData, GebaeudewaermeData
    )
    out = []
    # LandUse: only owner=testsim (avoid duplicates; also global None copy exists)
    from django.contrib.auth import get_user_model
    User = get_user_model()
    user = User.objects.get(username="testsim")
    # Use owner=None (global canonical seed rows) for audit to avoid workspace drift.
    for lu in LandUse.all_objects.filter(owner=None):
        out.append({
            "db_table": "LandUse",
            "db_id": lu.id,
            "db_code": lu.code,
            "db_name": lu.name,
            "db_status": float(lu.status_ha) if lu.status_ha is not None else None,
            "db_ziel": float(lu.target_ha) if lu.target_ha is not None else None,
            "db_unit": "ha",
            "hint_sheet": "1. Flächen",
        })
    # VerbrauchData: code-based, needs dispatch
    for v in VerbrauchData.all_objects.filter(owner=None):
        # Verbrauch sheet dispatch by leading code char
        code = (v.code or "").strip()
        if code.startswith("1.") or code == "1":
            hint = "3. Bedarfsniveau"  # KLIK-related
            # but detailed bedarfsniveau items might be in 4. Verbrauch; rely on name match
            hint2 = "4. Verbrauch"
        elif code.startswith("6.") or code == "6":
            hint = "6. Fossile"
            hint2 = None
        elif code.startswith("10") or code == "10":
            hint = "6. Fossile"
            hint2 = None
        else:
            hint = "4. Verbrauch"
            hint2 = None
        out.append({
            "db_table": "VerbrauchData",
            "db_id": v.id,
            "db_code": code,
            "db_name": v.category or v.code,
            "db_status": float(v.status) if v.status is not None else None,
            "db_ziel": float(v.ziel) if v.ziel is not None else None,
            "db_unit": v.unit,
            "hint_sheet": hint,
            "hint_sheet2": hint2,
        })
    for r in RenewableData.all_objects.filter(owner=None):
        out.append({
            "db_table": "RenewableData",
            "db_id": r.id,
            "db_code": r.code,
            "db_name": r.name,
            "db_status": float(r.status_value) if r.status_value is not None else None,
            "db_ziel": float(r.target_value) if r.target_value is not None else None,
            "db_unit": r.unit,
            "hint_sheet": "2. Erneuerbare",
        })
    for g in GebaeudewaermeData.all_objects.all():  # no owner scoping
        out.append({
            "db_table": "GebaeudewaermeData",
            "db_id": g.id,
            "db_code": g.code,
            "db_name": g.category or g.code,
            "db_status": float(g.status) if g.status is not None else None,
            "db_ziel": float(g.ziel) if g.ziel is not None else None,
            "db_unit": g.unit,
            "hint_sheet": "4. Verbrauch",  # GW rows are in 4. Verbrauch section 2.x
        })
    return out

# --- Match + compare ---------------------------------------------------

def best_match(db_name, candidates):
    """Return (candidate, score) where score = 1 for exact norm match, 0.9 for substring, 0 otherwise."""
    dn = norm_name(db_name)
    if not dn:
        return (None, 0)
    # Exact
    for c in candidates:
        if c["norm_name"] == dn:
            return (c, 1.0)
    # Exact with hyphen/space variants
    dn_strip = dn.replace(" ", "")
    for c in candidates:
        if c["norm_name"].replace(" ", "") == dn_strip:
            return (c, 0.99)
    # Substring match
    best = (None, 0.0)
    for c in candidates:
        cn = c["norm_name"]
        if not cn:
            continue
        if dn in cn or cn in dn:
            # Prefer shorter match (more specific)
            overlap = min(len(dn), len(cn)) / max(len(dn), len(cn))
            if overlap > best[1]:
                best = (c, 0.5 + 0.5 * overlap)
    return best

# --- Main --------------------------------------------------------------

def main():
    print("Building Excel index...")
    xl_idx = build_excel_index()
    for sname, rows in xl_idx.items():
        print(f"  {sname}: {len(rows)} rows indexed")

    print("Loading DB rows...")
    db_rows = load_db()
    print(f"  {len(db_rows)} DB rows")

    results = []
    for db in db_rows:
        hint = db.get("hint_sheet")
        hint2 = db.get("hint_sheet2")
        candidates = list(xl_idx.get(hint, []))
        if hint2:
            candidates += list(xl_idx.get(hint2, []))
        match, score = best_match(db["db_name"], candidates)

        if not match:
            # Try all sheets as a fallback
            all_cands = [r for rows in xl_idx.values() for r in rows]
            match, score = best_match(db["db_name"], all_cands)
            if match:
                score = max(0.0, score - 0.1)  # penalty for cross-sheet match

        row = {
            "db_table": db["db_table"],
            "db_id": db["db_id"],
            "db_code": db["db_code"],
            "db_name": db["db_name"],
            "db_unit": db.get("db_unit"),
            "db_status": db["db_status"],
            "db_ziel": db["db_ziel"],
            "match_score": round(score, 3),
            "match_sheet": match["sheet"] if match else "",
            "match_row": match["row"] if match else "",
            "match_name": match["full_name"] if match else "",
            "match_unit": match["unit"] if match else "",
            "match_dref": match["dref"] if match else "",
            "xl_status": match["status"] if match else None,
            "xl_ziel": match["ziel"] if match else None,
        }

        # Compute drift at scale=1
        row["drift_status"] = drift(db["db_status"], row["xl_status"]) if row["xl_status"] is not None else None
        row["drift_ziel"] = drift(db["db_ziel"], row["xl_ziel"]) if row["xl_ziel"] is not None else None

        # Try scales
        if row["xl_status"] is not None:
            sc, d = try_scales(db["db_status"], row["xl_status"])
            row["best_scale_status"] = sc
            row["best_drift_status"] = d
        else:
            row["best_scale_status"] = None
            row["best_drift_status"] = None

        if row["xl_ziel"] is not None:
            sc, d = try_scales(db["db_ziel"], row["xl_ziel"])
            row["best_scale_ziel"] = sc
            row["best_drift_ziel"] = d
        else:
            row["best_scale_ziel"] = None
            row["best_drift_ziel"] = None

        # Verdict
        def verdict(d, sc):
            if d is None:
                return "NO_DATA"
            if d == math.inf:
                return "NO_MATCH"
            if d == 0.0:
                return "EXACT"
            if d < TOL_REL_COSMETIC:
                return "PASS_COSMETIC"
            if d < TOL_REL:
                return "PASS"
            if sc and sc != 1.0:
                return f"DRIFT_SCALE_{sc}"
            return "DRIFT"

        row["verdict_status"] = verdict(row["best_drift_status"], row["best_scale_status"])
        row["verdict_ziel"] = verdict(row["best_drift_ziel"], row["best_scale_ziel"])
        results.append(row)

    # --- Emit CSV ---
    cols = [
        "db_table", "db_id", "db_code", "db_name", "db_unit", "db_status", "db_ziel",
        "match_score", "match_sheet", "match_row", "match_name", "match_unit", "match_dref",
        "xl_status", "xl_ziel",
        "drift_status", "drift_ziel",
        "best_scale_status", "best_drift_status",
        "best_scale_ziel", "best_drift_ziel",
        "verdict_status", "verdict_ziel",
    ]
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUT_CSV, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in results:
            w.writerow(r)
    print(f"wrote {OUT_CSV}: {len(results)} rows")

    # --- Summary ---
    from collections import Counter
    verdicts_s = Counter(r["verdict_status"] for r in results)
    verdicts_z = Counter(r["verdict_ziel"] for r in results)
    no_match = [r for r in results if r["match_score"] < 0.5]
    drift_rows = [r for r in results if r["verdict_status"].startswith("DRIFT") or r["verdict_ziel"].startswith("DRIFT")]

    with open(OUT_SUMM, "w", encoding="utf-8") as f:
        f.write("# §2 Value Parity — summary\n\n")
        f.write(f"**Total DB rows compared**: {len(results)}\n\n")
        f.write("## Verdict distribution — status column\n\n")
        for k, v in sorted(verdicts_s.items()):
            f.write(f"- {k}: {v}\n")
        f.write("\n## Verdict distribution — ziel column\n\n")
        for k, v in sorted(verdicts_z.items()):
            f.write(f"- {k}: {v}\n")
        f.write(f"\n## Matching quality\n\n")
        f.write(f"- Rows with match_score < 0.5 (no good Excel match): {len(no_match)}\n")
        f.write(f"- Rows with DRIFT in status or ziel: {len(drift_rows)}\n")
        f.write(f"\n## Self-skepticism log\n\n")
        f.write(f"- Tolerances tried: 0.001 (PASS) and 0.0001 (PASS_COSMETIC).\n")
        f.write(f"- Scale factors tried: 1, 1000, 1/1000, 10000, 1/10000, 100, 1/100.\n")
        f.write(f"- NO_MATCH rows demand manual cross-reference — see discrepancies.md.\n")

    # --- Discrepancies ---
    with open(OUT_DISC, "w", encoding="utf-8") as f:
        f.write("# §2 Value Parity — discrepancies\n\n")
        f.write("Rows flagged DRIFT, DRIFT_SCALE_*, or NO_MATCH.\n\n")
        # NO_MATCH first
        nmr = [r for r in results if r["match_score"] < 0.5]
        f.write(f"## NO_MATCH — {len(nmr)} rows\n\n")
        f.write("DB rows whose name couldn't be matched in any Excel sheet.\n\n")
        for r in nmr:
            f.write(f"- `{r['db_table']}[{r['db_code']}]` **{r['db_name']}** "
                    f"(status={r['db_status']}, ziel={r['db_ziel']}, unit={r['db_unit']}) "
                    f"— best sheet hint: `{r['match_sheet']}` row {r['match_row']} name={r['match_name']!r} score={r['match_score']}\n")

        # DRIFT rows
        f.write(f"\n## DRIFT — rows with significant status or ziel divergence\n\n")
        f.write("| table | code | name | col | db | xl | drift | best_scale | verdict |\n")
        f.write("|-------|------|------|-----|----|----|-------|------------|---------|\n")
        for r in results:
            for col in ("status", "ziel"):
                v = r[f"verdict_{col}"]
                if v.startswith("DRIFT") or v == "NO_MATCH":
                    d = r[f"best_drift_{col}"]
                    sc = r[f"best_scale_{col}"]
                    db_v = r[f"db_{col}"]
                    xl_v = r[f"xl_{col}"]
                    name = str(r["db_name"])[:60].replace("|", "\\|")
                    f.write(f"| {r['db_table']} | {r['db_code']} | {name} | {col} | {db_v} | {xl_v} | {d} | {sc} | {v} |\n")

    print(f"wrote {OUT_DISC}")
    print(f"wrote {OUT_SUMM}")

if __name__ == "__main__":
    main()
