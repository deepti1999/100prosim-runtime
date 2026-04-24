"""Probe _S.xlsx sheet layout.

For each scenario sheet (1.Flächen through 8.Kennzahlen), print the
first column B (row code / label) down to the last non-empty row.
This gives us a grep-able index of code → row mapping.
"""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
SRC = ROOT / "docs" / "100prosim_d_250517_250517.1817m" / "_S.xlsx"
OUT = ROOT / "verification" / "formula_audit" / "01_value_parity" / "_S_layout.md"

SHEETS = [
    "1. Flächen",
    "2. Erneuerbare",
    "3. Bedarfsniveau",
    "4. Verbrauch",
    "5. Bilanz",
    "6. Fossile",
    "7. Verbrauch Status",
    "8. Kennzahlen",
]

def main():
    wb_vals = load_workbook(SRC, data_only=True)
    wb_form = load_workbook(SRC, data_only=False)
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("# _S.xlsx layout probe\n\n")
        for sname in SHEETS:
            ws_v = wb_vals[sname]
            ws_f = wb_form[sname]
            f.write(f"\n## `{sname}` — {ws_v.max_row}×{ws_v.max_column}\n\n")
            # Dump first row (header) - cols A..Z
            hdr = [(ws_v.cell(1, c).value, ws_v.cell(2, c).value, ws_v.cell(3, c).value) for c in range(1, min(26, ws_v.max_column+1))]
            f.write("**First-3-rows by column (A..Z)**:\n\n")
            f.write("| col | row1 | row2 | row3 |\n|-----|------|------|------|\n")
            from openpyxl.utils import get_column_letter
            for i, (r1, r2, r3) in enumerate(hdr, start=1):
                def safe(x):
                    s = str(x) if x is not None else ""
                    return s.replace("|", "\\|").replace("\n", " ")[:40]
                f.write(f"| {get_column_letter(i)} | {safe(r1)} | {safe(r2)} | {safe(r3)} |\n")
            # Dump column B (common "code/label" column) + C (often label)
            f.write("\n**All rows (A, B, C, then status/ziel columns if obvious)**:\n\n")
            f.write("```\n")
            for r in range(1, ws_v.max_row + 1):
                a = ws_v.cell(r, 1).value
                b = ws_v.cell(r, 2).value
                c = ws_v.cell(r, 3).value
                if a is None and b is None and c is None:
                    continue
                f.write(f"  r{r}: A={a!r:<40} B={b!r:<40} C={c!r}\n")
            f.write("```\n")
    print(f"wrote {OUT}")

if __name__ == "__main__":
    main()
