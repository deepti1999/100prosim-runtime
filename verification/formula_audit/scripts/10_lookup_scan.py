"""§9 — Scan for VLOOKUP/INDEX/MATCH/SUMIF formulas across all workbooks.

Enumerates every cell in every scenario-bearing sheet whose formula
contains VLOOKUP, INDEX, MATCH, SUMIF, SUMIFS, INDIRECT or COUNTIF.

Outputs: verification/formula_audit/08_cross_references/lookup_tables.md
"""
from __future__ import annotations
import os, sys, re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
OUT_DIR = ROOT / "verification" / "formula_audit" / "08_cross_references"
OUT_DIR.mkdir(parents=True, exist_ok=True)

BOOKS = [
    ("_S.xlsx", ["1. Flächen", "2. Erneuerbare", "3. Bedarfsniveau", "4. Verbrauch",
                 "5. Bilanz", "6. Fossile", "7. Verbrauch Status", "8. Kennzahlen"]),
    ("D.xlsx", ["1.", "7.VerbrauchStatus", "8.Kennzahlen", "9.Quellen",
                "O_", "I_BS.2", "I_BS.3", "I_Region", "I_Basisdaten", "WS_"]),
    ("WS.xlsm", ["1.Jahresbilanz_Strom", "Zeitreihen Kalkulation", "WS_",
                 "2. Jahresgang Strom", "2a. Jahresgang segmentiert"]),
]

LOOKUP_PATTERNS = ["VLOOKUP", "INDEX", "MATCH", "SUMIF", "SUMIFS",
                   "INDIRECT", "COUNTIF", "COUNTIFS", "HLOOKUP",
                   "OFFSET", "CHOOSE"]

PAT = re.compile(r"(" + "|".join(LOOKUP_PATTERNS) + r")\s*\(", re.IGNORECASE)

def scan_book(book, sheets):
    src = ROOT / "docs" / "100prosim_d_250517_250517.1817m" / book
    if not src.exists():
        return []
    wb = load_workbook(src, data_only=False)
    hits = []
    for sname in sheets:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if not isinstance(v, str):
                    continue
                if not v.startswith("="):
                    continue
                if PAT.search(v):
                    ref = f"{get_column_letter(c)}{r}"
                    hits.append({
                        "book": book,
                        "sheet": sname,
                        "ref": ref,
                        "formula": v,
                    })
    return hits

def main():
    all_hits = []
    for (book, sheets) in BOOKS:
        print(f"scanning {book}...")
        hits = scan_book(book, sheets)
        print(f"  {len(hits)} hits")
        all_hits.extend(hits)

    out = OUT_DIR / "lookup_tables.md"
    with open(out, "w", encoding="utf-8") as f:
        f.write("# §9 — VLOOKUP / INDEX / MATCH / SUMIF / INDIRECT scan\n\n")
        f.write(f"Total hits: {len(all_hits)}\n\n")

        # Group by function
        from collections import Counter, defaultdict
        by_func = defaultdict(list)
        for h in all_hits:
            # Extract first matching function
            m = PAT.search(h["formula"])
            func = m.group(1).upper() if m else "?"
            by_func[func].append(h)

        f.write("## Hits by function\n\n")
        f.write("| function | count |\n|----------|------:|\n")
        for func in sorted(by_func.keys()):
            f.write(f"| {func} | {len(by_func[func])} |\n")
        f.write("\n")

        for func in sorted(by_func.keys()):
            f.write(f"\n## {func} — {len(by_func[func])} hits\n\n")
            # Sample first 20
            sample = by_func[func][:20]
            f.write("| book | sheet | ref | formula |\n|---|---|---|---|\n")
            for h in sample:
                fo = h["formula"].replace("|", "\\|").replace("\n", " ")[:150]
                f.write(f"| {h['book']} | {h['sheet']} | {h['ref']} | `{fo}` |\n")
            if len(by_func[func]) > 20:
                f.write(f"\n(... {len(by_func[func]) - 20} more hits not shown — see CSV if needed)\n")

    print(f"wrote {out} with {len(all_hits)} hits")

if __name__ == "__main__":
    main()
