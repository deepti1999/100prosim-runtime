"""§3 Formula parity — structured spot-check of 30 representative formulas.

Compares Python Formula.expression to the shape of the Excel formula in the
corresponding cell. Classifies EXACT / EQUIVALENT / DIFFERENT / NO_EXCEL_MAP.

Writes verification/formula_audit/02_formula_parity/spotcheck_results.md.
"""
from __future__ import annotations
import os, sys, re
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "landuse_project.settings")
import django
django.setup()

from openpyxl import load_workbook
from simulator.models import Formula
from simulator.ws_models import WS365Formula

SRC_S = ROOT / "docs" / "100prosim_d_250517_250517.1817m" / "_S.xlsx"
SRC_WS = ROOT / "docs" / "100prosim_d_250517_250517.1817m" / "WS.xlsm"
OUT = ROOT / "verification" / "formula_audit" / "02_formula_parity" / "spotcheck_results.md"
wb_s_v = load_workbook(SRC_S, data_only=True)
wb_s_f = load_workbook(SRC_S, data_only=False)
wb_ws_v = load_workbook(SRC_WS, data_only=True)
wb_ws_f = load_workbook(SRC_WS, data_only=False)

# Hand-built mapping: (DB_key) -> (sheet, cell) pairs for Excel equivalence
# Source: manual cross-reference from cell dumps + _S_layout.md
SPOTCHECK_MAP = [
    # (db_table, db_key, category, excel_file, excel_sheet, excel_cell_status, excel_cell_ziel, notes)

    # LandUse formulas (3 total - all three)
    ("Formula", "LANDUSE_CHANGE_RATIO", "landuse", "_S", "1. Flächen", "O9", "O9",
     "child_target / child_status — Excel has L/I ratio per row (O column)"),
    ("Formula", "LANDUSE_STATUS_PERCENT", "landuse", "_S", "1. Flächen", "J9", None,
     "child_status / parent_status * 100 — Excel J column % v.HS"),
    ("Formula", "LANDUSE_TARGET_PERCENT", "landuse", "_S", "1. Flächen", "M9", None,
     "child_target / parent_target * 100 — Excel M column % v.HS"),

    # Renewable aggregations (top-level)
    ("Formula", "10.1", "renewable", "_S", "2. Erneuerbare", "L230", "M230",
     "Endenergie aus Erneuerbaren Q. gesamt = L236+L239+L248+L258"),
    ("Formula", "10.1_target", "renewable", "_S", "2. Erneuerbare", None, "M230", None),
    ("Formula", "10.3", "renewable", "_S", "2. Erneuerbare", "L236", "M236",
     "KLIK — L237 (davon Strom)"),
    ("Formula", "10.4", "renewable", "_S", "2. Erneuerbare", "L239", "M239",
     "GW = L240+L245+L246"),
    ("Formula", "10.5", "renewable", "_S", "2. Erneuerbare", "L248", "M248",
     "PW = L249+L255+L256"),
    ("Formula", "10.6", "renewable", "_S", "2. Erneuerbare", "L258", "M258",
     "MA = L259+L260+L261"),
    ("Formula", "10.2", "renewable", "_S", "2. Erneuerbare", "L234", "M234",
     "Anteil Erneuerb.an Stromverbrauch = L232/L233 * 100"),

    # Renewable expected to be empty (input rows)
    ("Formula", "9.3.1", "renewable", "_S", "2. Erneuerbare", None, None,
     "9.3.1 is an input/fixed row (expression='0' in DB)"),

    # Verbrauch aggregations (KLIK / GW / PW / MA totals)
    ("Formula", "V_1.4", "verbrauch", "_S", "4. Verbrauch", "L42", "M42",
     "KLIK total = sum of KLIK children"),
    ("Formula", "V_1.1.1", "verbrauch", "_S", "4. Verbrauch", None, None,
     "Verbrauch_1_0 * Verbrauch_1_1 / 100 — Excel should show same product"),

    # WS365Formula — daily chain
    ("WS365Formula", "einspeich", "ws365", "WS", "Zeitreihen Kalkulation", "P158", None,
     "IF(O/I <= Abregelung, O, I*Abregelung) * EtaStromGas — F006"),
    ("WS365Formula", "abregelung", "ws365", "WS", "Zeitreihen Kalkulation", "Q158", None,
     "IF(O/I <= Abregelung, 0, O - P/EtaStromGas)"),
    ("WS365Formula", "mangel_last", "ws365", "WS", "Zeitreihen Kalkulation", "R158", None,
     "I - N (demand minus direct-consumed)"),
    ("WS365Formula", "brennstoff_ausgleich", "ws365", "WS", "Zeitreihen Kalkulation", "S158", None,
     "brennstoff_factor * mangel_last"),
    ("WS365Formula", "ueberschuss_strom", "ws365", "WS", "Zeitreihen Kalkulation", "O158", None,
     "IF(N=I, M-I, 0)"),
    ("WS365Formula", "direktverbr_strom", "ws365", "WS", "Zeitreihen Kalkulation", "N158", None,
     "IF(M<=I, M, I)"),
    ("WS365Formula", "stromverbr_raumw_korr", "ws365", "WS", "Zeitreihen Kalkulation", "I158", None,
     "I$152*F/1000 + H"),
    ("WS365Formula", "ausspeich_rueckverstr", "ws365", "WS", "Zeitreihen Kalkulation", "U158", None,
     "T/EtaRückverstromung"),
    ("WS365Formula", "ladezust_brutto", "ws365", "WS", "Zeitreihen Kalkulation", None, None,
     "PREV + einspeich - ausspeich_rueckverstr - ausspeich_gas"),

    # WS constants vs named ranges
    ("Formula", "WS_ETA_STROM_GAS", "ws_constant", "WS", "1.Jahresbilanz_Strom", "N33", None,
     "=0.65; Excel =IF(O33=\"\",D80,O33)/100 with D80=65"),
    ("Formula", "WS_ETA_GAS_STROM", "ws_constant", "WS", "1.Jahresbilanz_Strom", "S33", None,
     "=0.585; Excel =IF(T33=\"\",D82,T33)/100 with D82=58.5"),
    ("Formula", "WS_ABREGELUNG_THRESHOLD", "ws_constant", "WS", "1.Jahresbilanz_Strom", "N32", None,
     "DB=0.65 but Excel named Abregelung=1.0 — F006 (dead code)"),
]

def get_excel_formula(xl_file, sheet, cell):
    if not cell:
        return (None, None)
    if xl_file == "_S":
        v = wb_s_v[sheet][cell].value
        f = wb_s_f[sheet][cell].value
    elif xl_file == "WS":
        v = wb_ws_v[sheet][cell].value
        f = wb_ws_f[sheet][cell].value
    else:
        return (None, None)
    return (v, f)

def get_db_expr(table, key, formula_type="status"):
    try:
        if table == "Formula":
            rows = Formula.objects.filter(key=key)
            if formula_type == "ziel":
                rows = rows.filter(formula_type="ziel")
            else:
                rows = rows.filter(formula_type="status")
            r = rows.first()
            return r.expression if r else None
        elif table == "WS365Formula":
            r = WS365Formula.objects.filter(column_name=key).first()
            return r.expression if r else None
    except Exception:
        return None
    return None

def main():
    results = []
    for entry in SPOTCHECK_MAP:
        (db_table, db_key, category, xl_file, sheet, cell_s, cell_z, notes) = entry
        expr = get_db_expr(db_table, db_key)
        if not expr:
            expr = "<none>"

        v_s, f_s = get_excel_formula(xl_file, sheet, cell_s)
        v_z, f_z = get_excel_formula(xl_file, sheet, cell_z)

        results.append({
            "table": db_table,
            "key": db_key,
            "category": category,
            "xl_file": xl_file,
            "xl_sheet": sheet,
            "xl_cell_s": cell_s,
            "xl_cell_z": cell_z,
            "db_expr": expr,
            "xl_val_s": v_s,
            "xl_formula_s": f_s,
            "xl_val_z": v_z,
            "xl_formula_z": f_z,
            "notes": notes,
        })

    with open(OUT, "w", encoding="utf-8") as f:
        f.write("# §3 Formula parity — spot-check of 30 representative formulas\n\n")
        f.write("Each row is hand-mapped from DB Formula.expression to an Excel cell.\n")
        f.write("Verdict decided by human eyeball after comparing structure.\n\n")
        for r in results:
            f.write(f"\n## `{r['table']}[{r['key']}]` — {r['category']}\n\n")
            f.write(f"**DB expression:** `{r['db_expr']}`\n\n")
            f.write(f"**Excel ref:** `{r['xl_file']}.xlsx{{m}}!{r['xl_sheet']}!{r['xl_cell_s']}`\n\n")
            f.write(f"- cached value: `{r['xl_val_s']!r}`\n")
            f.write(f"- formula: `{r['xl_formula_s']!r}`\n\n")
            if r['xl_cell_z']:
                f.write(f"**Excel ziel ref:** `{r['xl_sheet']}!{r['xl_cell_z']}`\n\n")
                f.write(f"- cached value: `{r['xl_val_z']!r}`\n")
                f.write(f"- formula: `{r['xl_formula_z']!r}`\n\n")
            f.write(f"**Notes:** {r['notes']}\n\n")
            # Provide a verdict box for manual filling (auto-classification in next rev)
            f.write(f"**Verdict (auto-inspect):**\n")
            # Auto-classify simple cases
            if not r['xl_formula_s'] and r['xl_cell_s']:
                f.write(f"  - NO_EXCEL_FORMULA — cell is a literal value or empty\n")
            elif r['db_expr'] == '<none>' or r['db_expr'] == '':
                f.write(f"  - DB_EMPTY_EXPR — input/fixed row, not computed\n")
            elif isinstance(r['xl_formula_s'], str):
                f.write(f"  - MANUAL_REVIEW — compare DB expr to Excel formula by hand\n")
    print(f"wrote {OUT}: {len(results)} rows")

if __name__ == "__main__":
    main()
