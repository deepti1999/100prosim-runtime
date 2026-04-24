"""Build 6 curated DB-to-Excel mapping CSVs.

Strategy per table:
  LandUse (20)        — hand mapping from Round 1's Flächen dump.
  Gebäudewärme (26)   — hand mapping from 4. Verbrauch section 2.x.
  Verbrauch (151)     — automated by D-ref + name; reviewed via code position.
  Renewable (223)     — section-aware via 2. Erneuerbare row-by-row walk.
  Formula (760)       — deterministic: most Formulas map to the Excel cell of
                        their namesake data row (e.g. Formula[key='10.1', ft='status']
                        → the Excel cell used to compute Renewable[10.1].status).
  Bilanz              — enumerated directly from _S.xlsx!5. Bilanz.

Outputs:
  01_curated_mappings/landuse_to_excel.csv
  01_curated_mappings/gebaeudewaerme_to_excel.csv
  01_curated_mappings/verbrauch_to_excel.csv
  01_curated_mappings/renewable_to_excel.csv
  01_curated_mappings/formula_to_excel.csv
  01_curated_mappings/bilanz_to_excel.csv
"""
from __future__ import annotations
import os, sys, csv, unicodedata, re
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "landuse_project.settings")
import django
django.setup()

from openpyxl import load_workbook
from simulator.models import LandUse, VerbrauchData, RenewableData, GebaeudewaermeData, Formula
from simulator.ws_models import WS365Formula

SRC = ROOT / "docs" / "100prosim_d_250517_250517.1817m" / "_S.xlsx"
WS_SRC = ROOT / "docs" / "100prosim_d_250517_250517.1817m" / "WS.xlsm"
OUT_DIR = ROOT / "verification" / "formula_audit_full" / "01_curated_mappings"
OUT_DIR.mkdir(parents=True, exist_ok=True)

def norm(s):
    if s is None: return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = s.replace("ä","ae").replace("ö","oe").replace("ü","ue").replace("ß","ss")
    s = s.replace("Ä","ae").replace("Ö","oe").replace("Ü","ue")
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()

# --- LandUse — hand mapping derived from Round 1 Flächen dump ------
LANDUSE_MAP = {
    "LU_0":   ("1. Flächen", 8,  "Bodenfläche gesamt"),
    "LU_1":   ("1. Flächen", 9,  "Siedlung (Gebäude- & Freifläche)"),
    "LU_1.1": ("1. Flächen", 10, "Solare Dachflächen"),
    "LU_2":   ("1. Flächen", 12, "Landwirtschaftsfläche (LF)"),
    "LU_2.1": ("1. Flächen", 13, "Solare Freiflächen"),
    "LU_2.2": ("1. Flächen", 14, "Ackerland"),
    "LU_2.2.1": ("1. Flächen", 15, "Getreide-Anbaufl. (Stroh)"),
    "LU_2.2.2": ("1. Flächen", 16, "Energiepfl. (Biogas)"),
    "LU_2.2.3": ("1. Flächen", 17, "Energiepfl. (Pflanzenöl)"),
    "LU_2.2.4": ("1. Flächen", 18, "Energiepfl. (Ethanol)"),
    "LU_2.2.5": ("1. Flächen", 19, "Energiepfl. (Kurzumtr.)"),
    "LU_2.2.6": ("1. Flächen", 20, "(ohne energet. Relevanz)"),
    "LU_2.3":   ("1. Flächen", 22, "Dauergrünland"),
    "LU_2.3.1": ("1. Flächen", 23, "Grasschnitt (Biogas)"),
    "LU_2.4":   ("1. Flächen", 25, "(sonstige Nutzung)"),
    "LU_3":     ("1. Flächen", 27, "Waldfläche"),
    "LU_3.1":   ("1. Flächen", 28, "Forstfl. (u.a.Energieholz)"),
    "LU_3.2":   ("1. Flächen", 29, "(ohne forstwirtsch.Nutzung)"),
    "LU_4":     ("1. Flächen", 31, "(sonstige Flächen)"),
    "LU_5":     ("1. Flächen", 33, "Windenergie Flächenpotenzial"),
    "LU_6":     ("1. Flächen", 34, "Windparkfläche*"),
    "LU_6.1":   ("1. Flächen", 35, "Belegung (Potenzialflächen) *"),
}

# Status column = I, Ziel column = L for Flächen


def build_landuse_csv(wb_v):
    ws = wb_v["1. Flächen"]
    rows = []
    lus = list(LandUse.all_objects.filter(owner=None).order_by("code"))
    for lu in lus:
        m = LANDUSE_MAP.get(lu.code)
        if not m:
            rows.append({
                "db_code": lu.code, "db_name": lu.name, "db_unit": "ha",
                "excel_sheet": "", "excel_row": "", "excel_name_expected": "",
                "excel_cell_status": "", "excel_cell_ziel": "",
                "oos_reason": "NO_MAPPING_FOUND — manual review needed",
            })
            continue
        sheet, row, name_expected = m
        # Verify the expected name matches what Excel actually has
        actual = ws[f"AL{row}"].value or ws[f"E{row}"].value or ws[f"F{row}"].value or ""
        rows.append({
            "db_code": lu.code,
            "db_name": lu.name,
            "db_unit": "ha",
            "excel_sheet": sheet,
            "excel_row": row,
            "excel_name_expected": name_expected,
            "excel_cell_status": f"{sheet}!I{row}",
            "excel_cell_ziel": f"{sheet}!L{row}",
            "excel_actual_name": str(actual),
            "oos_reason": "",
        })
    out = OUT_DIR / "landuse_to_excel.csv"
    with open(out, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)
    print(f"  LandUse: {len(rows)} rows -> {out.name}")


# --- Gebäudewärme — maps to 4. Verbrauch section 2.x and nearby -----
# Based on VerbrauchData codes 2.0-2.10 + sub-hierarchy
GW_MAP = {
    "2.0":     ("4. Verbrauch", 8,   "Bedarfsniveau (statisch)"),
    "2.1":     ("4. Verbrauch", 11,  "davon Haushalte"),
    "2.1.1":   ("4. Verbrauch", 12,  "Wohnfläche pro Person"),
    "2.1.2":   ("4. Verbrauch", 13,  "Zieleinfluss Wohnflächen-Entwicklung"),
    "2.1.3":   ("4. Verbrauch", 14,  "Endenergie-Koeffizient"),
    "2.1.4":   ("4. Verbrauch", 15,  "Zieleinfluss Endanwendungs-Effizienz"),
    "2.2":     ("4. Verbrauch", 17,  "davon Handel/Dienstl."),
    "2.2.1":   ("4. Verbrauch", 18,  "Nutzvol. pro Pers/Anzahl Personen"),
    "2.2.2":   ("4. Verbrauch", 19,  "Zieleinfluss Vol./Anzahl-Entwicklung"),
    "2.2.3":   ("4. Verbrauch", 20,  "Endenergie-Koeffizient"),
    "2.2.4":   ("4. Verbrauch", 21,  "Zieleinfluss Endanwendungs-Effizienz"),
    "2.3":     ("4. Verbrauch", 23,  "davon GHD-Sockel/Industrie"),
    "2.4":     ("4. Verbrauch", 25,  "Endenergieverbrauch GW gesamt"),
    "2.5":     ("4. Verbrauch", 29,  "davon Warmwasser"),
    "2.5.1":   ("4. Verbrauch", 30,  "Warmwasser pro Person"),
    "2.5.2":   ("4. Verbrauch", 31,  "Zieleinfluss Warmwasser-Entwicklung"),
    "2.5.3":   ("4. Verbrauch", 32,  "= Res. Anteil Warmwasser an Gebäudew."),
    "2.6":     ("4. Verbrauch", 34,  "davon Raumwärme"),
    "2.6.1":   ("4. Verbrauch", 35,  "= Rechnerischer Anteil"),
    "2.7":     ("4. Verbrauch", 37,  "davon Brennstoffe"),
    "2.7.0":   ("4. Verbrauch", 38,  "= Endenergieverbrauch"),
    "2.8":     ("4. Verbrauch", 40,  "davon Wärme (30-100°C/verlustarm)"),
    "2.8.0":   ("4. Verbrauch", 41,  "= Endenergieverbrauch"),
    "2.9":     ("4. Verbrauch", 43,  "davon Strom (verlustarm nutzb.)"),
    "2.9.0":   ("4. Verbrauch", 44,  "= Endenergieverbrauch"),
    "2.9.1":   ("4. Verbrauch", 45,  "davon für Wärmepumpen"),
    "2.9.2":   ("4. Verbrauch", 46,  "="),
    "2.10":    ("4. Verbrauch", 48,  "Endenergieverbrauch GW gesamt"),
}


def build_gw_csv(wb_v):
    """Build the Gebäudewärme mapping CSV."""
    ws = wb_v["4. Verbrauch"]
    rows = []
    gws = list(GebaeudewaermeData.objects.all().order_by("code"))
    for g in gws:
        m = GW_MAP.get(g.code)
        if not m:
            rows.append({
                "db_code": g.code, "db_name": g.category or "",
                "db_unit": g.unit or "",
                "excel_sheet": "", "excel_row": "",
                "excel_name_expected": "",
                "excel_cell_status": "", "excel_cell_ziel": "",
                "oos_reason": "NO_MAPPING_FOUND",
            })
            continue
        sheet, row, name_expected = m
        actual = ws[f"AQ{row}"].value or ""
        rows.append({
            "db_code": g.code, "db_name": g.category or "",
            "db_unit": g.unit or "",
            "excel_sheet": sheet, "excel_row": row,
            "excel_name_expected": name_expected,
            "excel_cell_status": f"{sheet}!L{row}",
            "excel_cell_ziel": f"{sheet}!M{row}",
            "excel_actual_name": str(actual),
            "oos_reason": "",
        })
    out = OUT_DIR / "gebaeudewaerme_to_excel.csv"
    with open(out, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)
    print(f"  Gebäudewärme: {len(rows)} rows -> {out.name}")


# --- Verbrauch — automated by sheet + code position ---
# VerbrauchData codes use '1.x' (KLIK via 4. Verbrauch rows 4-16),
# '2.x' (GW via 4. Verbrauch rows 17+), '3.x' (PW), '4.x' (MA traktion detail
# in older seeds), '6.x' (MA via 4. Verbrauch or 6. Fossile), '10.x' (summary).
# Build an INDEX of (row → full hierarchical code) by walking the sheet.

def infer_verbrauch_codes(wb_v):
    """Walk 4. Verbrauch sheet and assign each data row a hierarchical code
    based on the indent depth of its name (columns E/F/G/H/I/J)."""
    ws = wb_v["4. Verbrauch"]
    # Stack maps indent-level → counter at that level. Reset downstream on shallower.
    # Sheet has a chapter header row — e.g. row 5 = "Wirtschaftliche Entwicklung" (chapter 1 = BIP),
    # row 8 = "Bedarfsniveau" (chapter 2). Chapter = E-level text.
    # Simpler: walk rows and for each, record row_num → (code_inferred, name_assembled).
    rows_info = []
    # levels map: "E"=0, "F"=1, "G"=2, "H"=3, "I"=4, "J"=5
    level_map = {"E": 0, "F": 1, "G": 2, "H": 3, "I": 4, "J": 5}
    counters = [0] * 6  # per-level counter
    chapter = 0  # top-level chapter (E content)
    for r in range(5, ws.max_row + 1):
        # Identify which indent level this row occupies
        indent_col = None
        first_text = None
        for col in ["E", "F", "G", "H", "I", "J"]:
            v = ws[f"{col}{r}"].value
            if v is None or not str(v).strip():
                continue
            s = str(v).strip()
            if s in ("*", "=", ">"):
                continue
            if indent_col is None:
                indent_col = col
                first_text = s
                break
        # Status / Ziel values
        L = ws[f"L{r}"].value
        M = ws[f"M{r}"].value
        has_value = L is not None or M is not None
        # Full name from hierarchy cols
        name_parts = []
        for col in ["E", "F", "G", "H", "I", "J"]:
            v = ws[f"{col}{r}"].value
            if v is not None and str(v).strip():
                name_parts.append(str(v).strip())
        full = " ".join(name_parts) if name_parts else None

        rows_info.append({
            "row": r, "indent_col": indent_col,
            "first_text": first_text, "full": full,
            "L": L, "M": M, "has_value": has_value,
        })
    return rows_info


def build_verbrauch_csv(wb_v):
    """Map each VerbrauchData row to its 4. Verbrauch or 6. Fossile row.

    Uses the chapter structure: codes 1.x (KLIK/BIP), 2.x (GW), 3.x (PW),
    4.x (NE/MA detail — older), 5.x (renewable-ratio), 6.x (MA), 10.x (summary).
    Sheet 4. Verbrauch has chapter boundaries at rows 5 (BIP), 8 (Bedarfsniveau),
    and per-chapter sub-sections.

    For this curated mapping, rely on the `VerbrauchData.category` field
    heuristics + row-by-row walk.
    """
    rows = []
    vs = list(VerbrauchData.all_objects.filter(owner=None).order_by("code"))
    # Use a lookup table built from the dump files plus indent tracking.
    # For initial pass, rely on VerbrauchData having the same position as a
    # row in 4. Verbrauch keyed by the numeric code part.
    # Scratch approach: scan 4. Verbrauch for rows with a matching category
    # label — that's essentially Round 1's fuzzy matcher re-wrapped but
    # restricted to sheet 4. Let's do a two-pass: exact on category name,
    # fallback to position.
    ws = wb_v["4. Verbrauch"]
    ws_f = wb_v["6. Fossile"] if "6. Fossile" in wb_v.sheetnames else None
    ws_b = wb_v["3. Bedarfsniveau"] if "3. Bedarfsniveau" in wb_v.sheetnames else None

    # Build name→(sheet,row) dict for 4. Verbrauch, 6. Fossile, 3. Bedarfsniveau
    name_idx = {}
    for sn in ["3. Bedarfsniveau", "4. Verbrauch", "6. Fossile"]:
        if sn not in wb_v.sheetnames: continue
        _ws = wb_v[sn]
        for r in range(5, _ws.max_row + 1):
            nm_parts = []
            for col in ["E", "F", "G", "H", "I", "J"]:
                v = _ws[f"{col}{r}"].value
                if v is not None and str(v).strip() and str(v).strip() not in ("*", "="):
                    nm_parts.append(str(v).strip())
            if not nm_parts:
                continue
            full = " ".join(nm_parts)
            key = norm(full)
            if key and key not in name_idx:
                name_idx[key] = (sn, r, full)

    for v in vs:
        cat = v.category or ""
        # Try exact match by normalized category
        k = norm(cat)
        if k in name_idx:
            sheet, row, excel_name = name_idx[k]
            rows.append({
                "db_code": v.code, "db_name": cat,
                "db_unit": v.unit or "",
                "excel_sheet": sheet, "excel_row": row,
                "excel_name_expected": excel_name,
                "excel_cell_status": f"{sheet}!L{row}",
                "excel_cell_ziel": f"{sheet}!M{row}",
                "oos_reason": "",
            })
        else:
            # Placeholder — mark OOS with reason
            rows.append({
                "db_code": v.code, "db_name": cat,
                "db_unit": v.unit or "",
                "excel_sheet": "", "excel_row": "",
                "excel_name_expected": "",
                "excel_cell_status": "", "excel_cell_ziel": "",
                "oos_reason": f"NO_NAME_MATCH in _S 3./4./6. sheets (category={cat!r})",
            })

    out = OUT_DIR / "verbrauch_to_excel.csv"
    with open(out, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)
    print(f"  Verbrauch: {len(rows)} rows -> {out.name}")


# --- Renewable — section-aware mapping ---
# _S.xlsx!2. Erneuerbare has chapter structure:
#   Solarenergie (row 5+)
#   Biogas (row 60+)
#   Biomasse
#   Holz / Stroh / Biodiesel / Bioethanol / Methan
#   Wasserkraft
#   Abwärme
#   Zusammenfassung (row 220+ — the 10.x aggregates)
# DB RenewableData has (category, subcategory, code) — category encodes
# the section.

def build_renewable_csv(wb_v):
    """Build section-aware mapping using (category + subcategory) to scope the
    Excel row search."""
    ws = wb_v["2. Erneuerbare"]
    rows_out = []
    rens = list(RenewableData.all_objects.filter(owner=None).order_by("code"))

    # Build a section-indexed dict of name → (row, excel_name)
    # Section boundaries in 2. Erneuerbare (approximate):
    SECTION_BOUNDS = [
        ("Solar", 5, 30),
        ("Solar/Dachflächen", 6, 11),
        ("Solar/Freiflächen", 12, 17),
        ("Biogas", 60, 105),
        ("Biomasse", 106, 150),
        ("Biogene Brennstoffe (flüssig)", 120, 145),
        ("Holz", 150, 175),
        ("Stroh", 175, 185),
        ("Wasserkraft", 185, 205),
        ("Abwärme", 205, 220),
        ("Zusammenfassung", 220, 293),
    ]

    # Build (row → (full_name, excel_cell_status, excel_cell_ziel))
    row_by_name_in_section = {}  # (section_hint, norm_name) -> row
    for (hint, r_lo, r_hi) in SECTION_BOUNDS:
        for r in range(r_lo, min(r_hi, ws.max_row) + 1):
            nm_parts = []
            for col in ["E", "F", "G", "H", "I", "J"]:
                v = ws[f"{col}{r}"].value
                if v is not None and str(v).strip() and str(v).strip() not in ("*", "="):
                    nm_parts.append(str(v).strip())
            if not nm_parts: continue
            full = " ".join(nm_parts)
            k = norm(full)
            if k:
                # Only register the FIRST occurrence in this section
                key = (hint, k)
                if key not in row_by_name_in_section:
                    row_by_name_in_section[key] = (r, full)
    # Also global name index (for rows outside the hinted section)
    global_idx = {}
    for r in range(5, ws.max_row + 1):
        nm_parts = []
        for col in ["E", "F", "G", "H", "I", "J"]:
            v = ws[f"{col}{r}"].value
            if v is not None and str(v).strip() and str(v).strip() not in ("*", "="):
                nm_parts.append(str(v).strip())
        if nm_parts:
            k = norm(" ".join(nm_parts))
            if k not in global_idx:
                global_idx[k] = (r, " ".join(nm_parts))

    # Section hint rules per DB category
    CAT_TO_HINT = {
        "Solar": "Solar",
        "Biogas": "Biogas",
        "Biomasse": "Biomasse",
        "Biogene Brennstoffe (flüssig)": "Biogene Brennstoffe (flüssig)",
        "Biogene Brennstoffe (gasförmig)": "Biogas",
        "Biogene Brennstoffe (fest)": "Holz",
        "Wasserkraft": "Wasserkraft",
        "Abwärme": "Abwärme",
        "Zusammenfassung": "Zusammenfassung",
    }

    for r in rens:
        cat = (r.category or "").strip()
        sub = (r.subcategory or "").strip()
        name = (r.name or "").strip()
        hint = CAT_TO_HINT.get(cat, None)
        # Try section-scoped first
        k = norm(name)
        matched_row = None; matched_name = None
        if hint and k:
            for (h, kk), (row, full) in row_by_name_in_section.items():
                if h == hint and kk == k:
                    matched_row = row
                    matched_name = full
                    break
        # Fallback: global index
        if not matched_row and k in global_idx:
            matched_row, matched_name = global_idx[k]
        # Fallback 2: match by code-prefix + name substring (for sub-hierarchy)
        if not matched_row:
            # Try matching on substring
            for kk, (row, full) in global_idx.items():
                if k and (k in kk or kk in k):
                    matched_row = row
                    matched_name = full
                    break
        if matched_row:
            rows_out.append({
                "db_code": r.code, "db_name": name,
                "db_category": cat, "db_subcategory": sub,
                "db_unit": r.unit or "",
                "excel_sheet": "2. Erneuerbare",
                "excel_row": matched_row,
                "excel_name_expected": matched_name,
                "excel_cell_status": f"2. Erneuerbare!L{matched_row}",
                "excel_cell_ziel": f"2. Erneuerbare!M{matched_row}",
                "oos_reason": "",
            })
        else:
            rows_out.append({
                "db_code": r.code, "db_name": name,
                "db_category": cat, "db_subcategory": sub,
                "db_unit": r.unit or "",
                "excel_sheet": "", "excel_row": "",
                "excel_name_expected": "",
                "excel_cell_status": "", "excel_cell_ziel": "",
                "oos_reason": f"NO_SECTION_MATCH (cat={cat}, sub={sub}, name={name!r})",
            })

    out = OUT_DIR / "renewable_to_excel.csv"
    with open(out, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows_out[0].keys()))
        w.writeheader(); w.writerows(rows_out)
    print(f"  Renewable: {len(rows_out)} rows -> {out.name}")


# --- Bilanz — enumerate every cell with a formula or cached value ---
def build_bilanz_csv(wb_v, wb_f):
    """Enumerate every non-empty cell on _S.xlsx!5. Bilanz with its formula
    and cached value. This becomes the authoritative Bilanz cell inventory."""
    ws = wb_v["5. Bilanz"]
    ws_f = wb_f["5. Bilanz"]
    rows = []
    from openpyxl.utils import get_column_letter
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            f = ws_f.cell(r, c).value
            if v is None and f is None:
                continue
            col = get_column_letter(c)
            rows.append({
                "excel_sheet": "5. Bilanz",
                "excel_ref": f"{col}{r}",
                "cached_value": v,
                "formula_text": f if isinstance(f, str) and f.startswith("=") else "",
                "row": r, "col": col,
            })
    out = OUT_DIR / "bilanz_to_excel.csv"
    with open(out, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)
    print(f"  Bilanz cells: {len(rows)} rows -> {out.name}")


# --- Formula — deterministic: Formula → corresponding data row's Excel cell ---
def build_formula_csv(wb_v):
    """For each Formula row, the Excel counterpart is the cell that COMPUTES
    the same output. For renewable/verbrauch/landuse Formula rows with a
    plain-code key (e.g. '10.1', 'V_2.9.0', 'LU_2.1'), the cell is the
    corresponding data row's status or ziel cell.

    For Formula rows that are Django-internal helpers (LANDUSE_CHANGE_RATIO
    etc.), they are OOS — marked as such.

    For WS constants + WS formulas, they map to named ranges and Zeitreihen
    cells respectively."""
    rows_out = []

    # Load the previously-built landuse, verbrauch, renewable mappings
    def load_csv(name):
        p = OUT_DIR / name
        if not p.exists():
            return {}
        with open(p, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            return {r["db_code"]: r for r in reader}

    lu_map = load_csv("landuse_to_excel.csv")
    v_map = load_csv("verbrauch_to_excel.csv")
    r_map = load_csv("renewable_to_excel.csv")
    g_map = load_csv("gebaeudewaerme_to_excel.csv")

    # WS named ranges (already known from Round 1)
    WS_CONST_MAP = {
        "WS_ETA_STROM_GAS":         ("WS.xlsm", "1.Jahresbilanz_Strom", "N33", "EtaStromGas"),
        "WS_ETA_GAS_STROM":         ("WS.xlsm", "1.Jahresbilanz_Strom", "S33", "EtaRückverstromung"),
        "WS_ABREGELUNG_THRESHOLD":  ("WS.xlsm", "1.Jahresbilanz_Strom", "N32", "Abregelung — DEAD in our code per F006"),
    }

    for f in Formula.objects.all().order_by("category", "key"):
        key = (f.key or "").strip()
        cat = f.category or ""
        ft = f.formula_type or "status"

        entry = {
            "formula_id": f.id,
            "formula_key": key,
            "formula_category": cat,
            "formula_type": ft,
            "is_fixed": f.is_fixed,
            "db_expression": (f.expression or "").strip(),
            "excel_book": "",
            "excel_sheet": "",
            "excel_cell": "",
            "resolution_note": "",
            "oos_reason": "",
        }

        # Dispatch by category
        if cat == "landuse":
            # The three helpers LANDUSE_CHANGE_RATIO etc. are Django-internal
            entry["oos_reason"] = (
                "HELPER — Django-internal ratio/percent helper; Excel encodes the "
                "same logic inline in the L/M/O columns of 1. Flächen per-row (J/O "
                "columns for %-of-parent + change-ratio)."
            )
        elif cat == "ws_constant":
            if key in WS_CONST_MAP:
                book, sheet, cell, note = WS_CONST_MAP[key]
                entry["excel_book"] = book
                entry["excel_sheet"] = sheet
                entry["excel_cell"] = cell
                entry["resolution_note"] = note
            else:
                entry["oos_reason"] = "WS_CONST_UNMAPPED"
        elif cat == "ws":
            entry["excel_book"] = "WS.xlsm"
            entry["excel_sheet"] = "Zeitreihen Kalkulation"
            # WS formulas evaluated across rows 158..521; representative = row 158
            entry["excel_cell"] = "(rows 158-521, one per day)"
            entry["resolution_note"] = (
                "Column varies by key: einspeich→P, abregelung→Q, mangel_last→R, "
                "brennstoff_ausgleich→S, speicher_ausgl_strom→T, "
                "ausspeich_rueckverstr→U, ausspeich_gas→V, ladezust_brutto→W, etc. "
                "Covered in §5 of Round 1; daily input columns C/D/E/F verified 0 drift."
            )
        elif cat == "renewable":
            # Strip _target/_ziel_target suffix to find the data-row code
            code = re.sub(r"(_target|_ziel_target|_ziel)$", "", key)
            mapped = r_map.get(code)
            if mapped and mapped.get("excel_cell_status"):
                entry["excel_book"] = "_S.xlsx"
                entry["excel_sheet"] = mapped["excel_sheet"]
                entry["excel_cell"] = mapped["excel_cell_ziel"].split("!", 1)[-1] if ft == "ziel" else mapped["excel_cell_status"].split("!", 1)[-1]
                entry["resolution_note"] = f"Via renewable_to_excel.csv ({mapped['db_name']!r})"
            else:
                entry["oos_reason"] = f"Renewable code {code!r} not in renewable_to_excel.csv"
        elif cat == "verbrauch":
            code = re.sub(r"(_target|_ziel_target|_ziel)$", "", key)
            code = re.sub(r"^V_", "", code)
            mapped = v_map.get(code)
            if mapped and mapped.get("excel_cell_status"):
                entry["excel_book"] = "_S.xlsx"
                entry["excel_sheet"] = mapped["excel_sheet"]
                entry["excel_cell"] = mapped["excel_cell_ziel"].split("!", 1)[-1] if ft == "ziel" else mapped["excel_cell_status"].split("!", 1)[-1]
                entry["resolution_note"] = f"Via verbrauch_to_excel.csv ({mapped['db_name']!r})"
            else:
                # Try gebäudewärme
                gm = g_map.get(code)
                if gm and gm.get("excel_cell_status"):
                    entry["excel_book"] = "_S.xlsx"
                    entry["excel_sheet"] = gm["excel_sheet"]
                    entry["excel_cell"] = gm["excel_cell_ziel"].split("!", 1)[-1] if ft == "ziel" else gm["excel_cell_status"].split("!", 1)[-1]
                    entry["resolution_note"] = f"Via gebaeudewaerme_to_excel.csv ({gm['db_name']!r})"
                else:
                    entry["oos_reason"] = f"Verbrauch code {code!r} not in verbrauch/gebäudewärme mappings"
        else:
            entry["oos_reason"] = f"Unknown category {cat!r}"

        rows_out.append(entry)

    # Also append WS365Formula rows (21)
    for f in WS365Formula.objects.all().order_by("order"):
        entry = {
            "formula_id": f"ws365_{f.id}",
            "formula_key": f.column_name,
            "formula_category": "ws365",
            "formula_type": f.stage,
            "is_fixed": False,
            "db_expression": (f.expression or "").strip(),
            "excel_book": "WS.xlsm",
            "excel_sheet": "Zeitreihen Kalkulation",
            "excel_cell": "(rows 158-521)",
            "resolution_note": (
                "Daily 365-row chain. Column P (einspeich), Q (abregelung), R "
                "(mangel_last), S (brennstoff_ausgleich), T (speicher_ausgl_strom), "
                "U (ausspeich_rueckverstr), V (ausspeich_gas), W (ladezust_brutto)."
            ),
            "oos_reason": "",
        }
        rows_out.append(entry)

    out = OUT_DIR / "formula_to_excel.csv"
    with open(out, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows_out[0].keys()))
        w.writeheader(); w.writerows(rows_out)
    print(f"  Formula: {len(rows_out)} rows -> {out.name}")


def main():
    wb_v = load_workbook(SRC, data_only=True)
    wb_f = load_workbook(SRC, data_only=False)

    build_landuse_csv(wb_v)
    build_gw_csv(wb_v)
    build_verbrauch_csv(wb_v)
    build_renewable_csv(wb_v)
    build_bilanz_csv(wb_v, wb_f)
    build_formula_csv(wb_v)  # depends on the others

if __name__ == "__main__":
    main()
