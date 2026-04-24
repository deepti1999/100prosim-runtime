"""Extract hierarchical codes from _S.xlsx sheets.

The _S.xlsx sheets don't explicitly store DB-style codes (LU_2.1,
9.3.1, etc.). They use indentation: columns E/F/G/H/I/J hold
hierarchical name text where deeper columns = deeper nesting level.

This script walks each sheet and reconstructs a hierarchical code
for each data row by tracking the most-recent non-empty cell per
column level. A row whose first-indented column is G implies the
row is a level-3 child of the most recent level-2 (F) parent.

Outputs: 01_curated_mappings/excel_rows/<sheet>.csv
with columns: sheet, row, col_E, col_F, col_G, col_H, col_I, col_J,
full_name, inferred_code, status_L_or_I, ziel_M_or_L, d_ref
"""
from __future__ import annotations
import csv
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
SRC = ROOT / "docs" / "100prosim_d_250517_250517.1817m" / "_S.xlsx"
OUT_DIR = ROOT / "verification" / "formula_audit_full" / "01_curated_mappings" / "excel_rows"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Sheet → column config:
#   name_cols: hierarchical name columns, outermost to innermost
#   status_col: column holding Status value
#   ziel_col:   column holding Ziel value
#   unit_col:   column holding unit string
#   dref_col:   column holding D.xlsx row reference (e.g. '1.64')
#   start_row:  first data row (exclusive of headers)
#   end_row:    last plausible data row
SHEETS = {
    "1. Flächen": {
        "name_cols": ["E", "F", "G", "H"],
        "status_col": "I",
        "ziel_col": "L",
        "unit_col": "I",
        "dref_col": "AG",
        "start_row": 5,
        "end_row": 36,
    },
    "2. Erneuerbare": {
        "name_cols": ["E", "F", "G", "H", "I", "J"],
        "status_col": "L",
        "ziel_col": "M",
        "unit_col": "K",
        "dref_col": "AV",
        "start_row": 5,
        "end_row": 293,
    },
    "3. Bedarfsniveau": {
        "name_cols": ["E", "F", "G", "H"],
        "status_col": "L",
        "ziel_col": "M",
        "unit_col": "K",
        "dref_col": "BA",
        "start_row": 5,
        "end_row": 53,
    },
    "4. Verbrauch": {
        "name_cols": ["E", "F", "G", "H", "I", "J"],
        "status_col": "L",
        "ziel_col": "M",
        "unit_col": "K",
        "dref_col": "AM",
        "start_row": 5,
        "end_row": 213,
    },
    "5. Bilanz": {
        "name_cols": ["F", "G"],  # simpler layout — row-major
        "status_col": None,  # varies per sector
        "ziel_col": None,
        "unit_col": None,
        "dref_col": None,
        "start_row": 5,
        "end_row": 71,
    },
    "6. Fossile": {
        "name_cols": ["E", "F", "G", "H", "I", "J"],
        "status_col": "L",
        "ziel_col": "M",
        "unit_col": "K",
        "dref_col": None,
        "start_row": 5,
        "end_row": 93,
    },
    "7. Verbrauch Status": {
        "name_cols": ["F", "G", "H", "I", "J"],
        "status_col": None,  # matrix layout; different shape
        "ziel_col": None,
        "unit_col": None,
        "dref_col": None,
        "start_row": 5,
        "end_row": 101,
    },
    "8. Kennzahlen": {
        "name_cols": ["F", "G", "H", "I"],
        "status_col": None,
        "ziel_col": None,
        "unit_col": None,
        "dref_col": None,
        "start_row": 5,
        "end_row": 128,
    },
}


def extract(sheet_name: str, layout: dict, wb_v, wb_f):
    ws = wb_v[sheet_name]
    ws_f = wb_f[sheet_name]
    out_rows = []

    # track most-recent name per column level (for code inference)
    name_stack = {c: None for c in layout["name_cols"]}

    for r in range(layout["start_row"], min(layout["end_row"], ws.max_row) + 1):
        # Find the deepest non-empty name column for this row
        name_parts = []
        row_first_col = None
        for col in layout["name_cols"]:
            v = ws[f"{col}{r}"].value
            if v is not None and str(v).strip() and str(v).strip() != "*" and str(v).strip() != "=":
                # If this row overwrites a parent's value, reset name_stack at deeper levels
                if row_first_col is None:
                    row_first_col = col
                name_stack[col] = str(v).strip()
                # Clear deeper levels (new sibling invalidates previous children)
                clear = False
                for c2 in layout["name_cols"]:
                    if clear:
                        name_stack[c2] = None
                    if c2 == col:
                        clear = True

        # Assemble full name = stack top-down
        full_name_parts = []
        for c in layout["name_cols"]:
            v = name_stack[c]
            if v:
                full_name_parts.append(v)
        full_name = " / ".join(full_name_parts) if full_name_parts else None

        # Get status / ziel if applicable
        status_v = None
        ziel_v = None
        if layout.get("status_col"):
            status_v = ws[f"{layout['status_col']}{r}"].value
        if layout.get("ziel_col"):
            ziel_v = ws[f"{layout['ziel_col']}{r}"].value

        # d_ref
        d_ref = None
        if layout.get("dref_col"):
            dv = ws[f"{layout['dref_col']}{r}"].value
            if dv is not None:
                d_ref = str(dv).strip()

        # Skip empty rows (no name, no values)
        if full_name is None and status_v is None and ziel_v is None:
            continue

        unit = None
        if layout.get("unit_col"):
            uv = ws[f"{layout['unit_col']}{r}"].value
            if uv is not None and not isinstance(uv, (int, float)):
                unit = str(uv).strip()

        # Also: row_first_col tells us the indent level
        out_rows.append({
            "sheet": sheet_name,
            "row": r,
            "row_indent_col": row_first_col,
            "full_name": full_name,
            "status": status_v,
            "ziel": ziel_v,
            "unit": unit,
            "d_ref": d_ref,
            "ref_status": f"{sheet_name}!{layout['status_col']}{r}" if layout.get("status_col") else "",
            "ref_ziel": f"{sheet_name}!{layout['ziel_col']}{r}" if layout.get("ziel_col") else "",
        })

    return out_rows


def main():
    wb_v = load_workbook(SRC, data_only=True)
    wb_f = load_workbook(SRC, data_only=False)

    for sname, layout in SHEETS.items():
        rows = extract(sname, layout, wb_v, wb_f)
        safe = sname.replace(" ", "_").replace(".", "")
        out = OUT_DIR / f"{safe}.csv"
        with open(out, "w", encoding="utf-8", newline="") as f:
            if rows:
                w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
                w.writeheader()
                w.writerows(rows)
        print(f"  {sname}: {len(rows)} rows -> {out.name}")

if __name__ == "__main__":
    main()
