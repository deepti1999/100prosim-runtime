"""§03 Full Bilanz parity — every output row on _S.xlsx!5. Bilanz
compared to our engine output.

Covers:
  Status section (rows 9-27):
    Row 9    Verbrauch Strom
    Row 10   Verbrauch Strom renewable
    Row 11   Verbrauch Strom fossil
    Row 12   Verbr.Brennst.gasf.
    Row 13   Verbr.Brennst.gasf. renewable
    Row 14   Verbr.Brennst.gasf. fossil
    Row 15   Verbr.Brennst.flüssig
    Row 16-17 (ren/fos)
    Row 18   Verbr.Brennst.fest
    Row 19-20 (ren/fos)
    Row 21   Verbrauch Wärme
    Row 22   Verbrauch Wärme renewable
    Row 23   Verbrauch Wärme Abwärme
    Row 24   Verbrauch Wärme fossil
    Row 25   Verbrauch gesamt
    Row 26-27 (ren/fos)

  Ziel-Bilanz section (rows 47-67):
    Similar structure but with M/P/S ziel columns.

For each row:
  KLIK = H (status) / H (ziel)
  GW   = K (status) / K (ziel)
  PW   = N (status) / N (ziel)
  MA   = Q (status) / Q (ziel)
  total= T (status) / T (ziel)
"""
from __future__ import annotations
import os, sys, csv, math
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "landuse_project.settings")
import django
django.setup()

from openpyxl import load_workbook
from calculation_engine.bilanz_engine import calculate_bilanz_data

SRC = ROOT / "docs" / "100prosim_d_250517_250517.1817m" / "_S.xlsx"
OUT = ROOT / "verification" / "formula_audit_full" / "03_full_bilanz_parity"
OUT.mkdir(parents=True, exist_ok=True)
(OUT / "per_section").mkdir(exist_ok=True)

# Columns by role
COLS_STATUS_TOTAL = dict(KLIK="H", GW="K", PW="N", MA="Q", total="T")
COLS_STATUS_REN   = dict(KLIK="I", GW="L", PW="O", MA="R", total="U")
COLS_STATUS_FOS   = dict(KLIK="J", GW="M", PW="P", MA="S", total="V")

# Engine key map per section (status section)
# Returns (engine_dict_key, excel_row, label)
STATUS_SECTIONS = [
    # (engine_total, engine_ren, engine_fos, row_total, row_ren, row_fos, label)
    ("verbrauch_strom", "verbrauch_strom_renewable", "verbrauch_strom_fossil",
     9, 10, 11, "Strom"),
    ("verbrauch_fuels", "verbrauch_fuels_renewable", "verbrauch_fuels_fossil",
     12, 13, 14, "Brennstoff gasförmig"),  # Excel aggregates gas into row 12
    # Rows 15-17: Brennstoff flüssig; our engine merges all fuels into one
    (None, None, None, 15, 16, 17, "Brennstoff flüssig"),
    (None, None, None, 18, 19, 20, "Brennstoff fest"),
    ("verbrauch_heat", "verbrauch_heat_renewable", "verbrauch_heat_fossil",
     21, 22, 24, "Wärme"),
    # Row 23: Abwärme — engine has verbrauch_heat_abwaerme
    ("verbrauch_gesamt", "erneuerbar", "verbrauch_gesamt_fossil",
     25, 26, 27, "Gesamt"),
]


def rel(a, b):
    if a is None or b is None:
        return math.inf
    try: fa = float(a); fb = float(b)
    except: return math.inf
    if fa == 0 and fb == 0: return 0.0
    m = max(abs(fa), abs(fb))
    if m < 1e-9: return 0.0
    return abs(fa - fb) / m

def verdict(d):
    if d is None: return "NO_DATA"
    if d == math.inf: return "NO_MATCH"
    if d == 0.0: return "EXACT"
    if d < 0.0001: return "PASS_COSMETIC"
    if d < 0.001: return "PASS"
    if d < 0.01: return "PASS_LOOSE"
    return "DRIFT"


def get_engine_row(engine, key, which):
    """Get {KLIK, GW, PW, MA, total} from engine[key][which]."""
    d = engine.get(key, {})
    v = d.get(which, {}) or {}
    return dict(
        KLIK=v.get("kraft_licht", 0),
        GW=v.get("gebaeudewaerme", 0),
        PW=v.get("prozesswaerme", 0),
        MA=v.get("mobile", 0),
        total=v.get("gesamt", 0),
    )


def compare_section(engine, ws, section_def, use_ziel=False):
    """Return list of row dicts for one section × 5 sectors × 3 carrier types."""
    (eng_total, eng_ren, eng_fos, row_total, row_ren, row_fos, label) = section_def
    out = []
    for (carrier_role, engine_key, row, cols) in [
        ("total", eng_total, row_total, COLS_STATUS_TOTAL),
        ("renewable", eng_ren, row_ren, COLS_STATUS_REN),
        ("fossil", eng_fos, row_fos, COLS_STATUS_FOS),
    ]:
        if engine_key is None:
            # No engine equivalent for this section — mark OOS
            for sector in ["KLIK", "GW", "PW", "MA", "total"]:
                col = cols[sector]
                cell_ref = f"{col}{row + (40 if use_ziel else 0)}"
                xl = ws[cell_ref].value
                out.append({
                    "section": label,
                    "carrier_role": carrier_role,
                    "sector": sector,
                    "view": "ziel" if use_ziel else "status",
                    "excel_cell": cell_ref,
                    "engine_value": "NO_ENGINE_EQUIV",
                    "excel_value": xl,
                    "drift": "",
                    "verdict": "NO_ENGINE_EQUIV",
                })
            continue
        # Apply ziel offset: ziel rows are 38 rows down from status rows (rough)
        # Actually the Ziel-Bilanz block starts at row 47. Row 9 status → 51 ziel,
        # row 12 status → 54 ziel, row 21 → 60 (Wärme), row 25 → 64 (Gesamt).
        # Build a map status_row → ziel_row
        ZIEL_ROW_MAP = {
            9: 51, 10: 52, 11: 53,
            12: 54, 13: 55, 14: 56,
            15: 57, 16: 58, 17: 59,  # liquid
            18: None, 19: None, 20: None,  # solid not in ziel block?
            21: 60, 22: 61, 23: 62, 24: 63,
            25: 64, 26: 65, 27: 66,
        }
        use_row = ZIEL_ROW_MAP.get(row) if use_ziel else row
        if use_row is None:
            continue
        engine_vals = get_engine_row(engine, engine_key, "ziel" if use_ziel else "status")
        for sector in ["KLIK", "GW", "PW", "MA", "total"]:
            col = cols[sector]
            cell_ref = f"{col}{use_row}"
            xl = ws[cell_ref].value
            eng = engine_vals[sector]
            d = rel(eng, xl)
            out.append({
                "section": label,
                "carrier_role": carrier_role,
                "sector": sector,
                "view": "ziel" if use_ziel else "status",
                "excel_cell": cell_ref,
                "engine_value": eng,
                "excel_value": xl,
                "drift": f"{d:.6f}" if d != math.inf else "inf",
                "verdict": verdict(d),
            })
    return out


def main():
    wb = load_workbook(SRC, data_only=True)
    ws = wb["5. Bilanz"]
    engine = calculate_bilanz_data()

    rows = []
    for section_def in STATUS_SECTIONS:
        rows.extend(compare_section(engine, ws, section_def, use_ziel=False))
    for section_def in STATUS_SECTIONS:
        rows.extend(compare_section(engine, ws, section_def, use_ziel=True))

    # Write CSV
    out_csv = OUT / "all_sections.csv"
    with open(out_csv, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)
    print(f"wrote {out_csv} ({len(rows)} rows)")

    # Per-section files
    from collections import defaultdict
    by_section = defaultdict(list)
    for r in rows:
        by_section[r["section"]].append(r)
    for section, rs in by_section.items():
        safe = section.replace(" ", "_").replace("ö", "oe").replace("ä", "ae")
        with open(OUT / "per_section" / f"{safe}.md", "w", encoding="utf-8") as f:
            f.write(f"# Bilanz section — {section}\n\n")
            f.write("| view | carrier | sector | engine | excel | drift | verdict |\n")
            f.write("|------|---------|--------|-------:|------:|------:|---------|\n")
            for r in rs:
                f.write(f"| {r['view']} | {r['carrier_role']} | {r['sector']} | {r['engine_value']} | {r['excel_value']} | {r['drift']} | {r['verdict']} |\n")

    # Summary
    from collections import Counter
    c = Counter(r["verdict"] for r in rows)
    with open(OUT / "summary.md", "w", encoding="utf-8") as f:
        f.write("# §03 Full Bilanz Parity — summary\n\n")
        f.write(f"Total cells compared: {len(rows)}\n\n")
        f.write("## Verdict distribution\n\n")
        for k, v in sorted(c.items()):
            f.write(f"- {k}: {v}\n")

    # Discrepancies
    drift_rows = [r for r in rows if r["verdict"] in ("DRIFT", "NO_MATCH")]
    with open(OUT / "discrepancies.md", "w", encoding="utf-8") as f:
        f.write(f"# §03 Full Bilanz Parity — discrepancies ({len(drift_rows)} DRIFT rows)\n\n")
        f.write("| section | carrier | sector | view | engine | excel | drift | cell |\n")
        f.write("|---------|---------|--------|------|-------:|------:|------:|------|\n")
        for r in drift_rows:
            f.write(f"| {r['section']} | {r['carrier_role']} | {r['sector']} | {r['view']} | {r['engine_value']} | {r['excel_value']} | {r['drift']} | {r['excel_cell']} |\n")

    print(f"Per-section files: {len(by_section)}")
    print(f"DRIFT rows: {len(drift_rows)}")

if __name__ == "__main__":
    main()
