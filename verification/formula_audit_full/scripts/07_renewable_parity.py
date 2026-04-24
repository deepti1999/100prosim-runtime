"""§04 Renewable section-aware value parity.

Use `01_curated_mappings/renewable_to_excel.csv` to compare every
RenewableData row to its mapped Excel cell. Check status + ziel
columns with 0.1 % tolerance + scale-factor search.

Outputs:
  04_renewable_section_aware/per_row_parity.csv
  04_renewable_section_aware/f005_resolution.md
  04_renewable_section_aware/discrepancies.md
  04_renewable_section_aware/summary.md
"""
from __future__ import annotations
import os, sys, csv, math
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "landuse_project.settings")
import django
django.setup()

from openpyxl import load_workbook
from simulator.models import RenewableData

MAP_CSV = ROOT / "verification" / "formula_audit_full" / "01_curated_mappings" / "renewable_to_excel.csv"
SRC = ROOT / "docs" / "100prosim_d_250517_250517.1817m" / "_S.xlsx"
OUT = ROOT / "verification" / "formula_audit_full" / "04_renewable_section_aware"
OUT.mkdir(parents=True, exist_ok=True)

TOL = 0.001  # 0.1%
TOL_COS = 0.0001  # 0.01%

def rel(a, b):
    if a is None or b is None: return math.inf
    try: fa = float(a); fb = float(b)
    except: return math.inf
    if fa == 0 and fb == 0: return 0.0
    m = max(abs(fa), abs(fb))
    if m < 1e-9: return 0.0
    return abs(fa - fb) / m

def try_scales(db, xl):
    if db is None or xl is None: return (1.0, rel(db, xl))
    try: fa = float(db)
    except: return (1.0, math.inf)
    best = (1.0, rel(fa, xl))
    for sc in (1000.0, 0.001, 10000.0, 0.0001, 100.0, 0.01):
        d = rel(fa * sc, xl)
        if d < best[1]:
            best = (sc, d)
    return best

def verdict(d, sc):
    if d is None: return "NO_DATA"
    if d == math.inf: return "NO_MATCH"
    if d == 0.0: return "EXACT"
    if d < TOL_COS: return "PASS_COSMETIC"
    if d < TOL: return "PASS"
    if sc and sc != 1.0: return f"DRIFT_SCALE_{sc}"
    return "DRIFT"


def main():
    wb = load_workbook(SRC, data_only=True)
    ws = wb["2. Erneuerbare"]

    # Load curated mapping
    with open(MAP_CSV, encoding="utf-8") as f:
        mapping = list(csv.DictReader(f))

    # Load DB values keyed by code
    db = {r.code: r for r in RenewableData.all_objects.filter(owner=None)}

    results = []
    for m in mapping:
        code = m["db_code"]
        r = db.get(code)
        if not r:
            continue
        entry = {
            "db_code": code,
            "db_name": r.name or "",
            "db_category": m.get("db_category", ""),
            "db_subcategory": m.get("db_subcategory", ""),
            "db_unit": r.unit or "",
            "excel_sheet": m["excel_sheet"],
            "excel_row": m["excel_row"],
            "excel_name_expected": m["excel_name_expected"],
            "db_status": float(r.status_value) if r.status_value is not None else None,
            "db_ziel": float(r.target_value) if r.target_value is not None else None,
            "xl_status": None,
            "xl_ziel": None,
            "drift_status": None,
            "drift_ziel": None,
            "scale_status": 1.0,
            "scale_ziel": 1.0,
            "verdict_status": "NO_MAPPING",
            "verdict_ziel": "NO_MAPPING",
            "oos_reason": m.get("oos_reason", ""),
        }
        if entry["oos_reason"]:
            entry["verdict_status"] = "NO_EXCEL_CELL_DOCUMENTED"
            entry["verdict_ziel"] = "NO_EXCEL_CELL_DOCUMENTED"
            results.append(entry); continue

        # Read Excel cells
        cs = m["excel_cell_status"].split("!", 1)[-1] if m["excel_cell_status"] else None
        cz = m["excel_cell_ziel"].split("!", 1)[-1] if m["excel_cell_ziel"] else None
        if cs:
            entry["xl_status"] = ws[cs].value
        if cz:
            entry["xl_ziel"] = ws[cz].value

        # Compare
        sc_s, d_s = try_scales(entry["db_status"], entry["xl_status"])
        sc_z, d_z = try_scales(entry["db_ziel"], entry["xl_ziel"])
        entry["scale_status"] = sc_s
        entry["scale_ziel"] = sc_z
        entry["drift_status"] = d_s
        entry["drift_ziel"] = d_z
        entry["verdict_status"] = verdict(d_s, sc_s)
        entry["verdict_ziel"] = verdict(d_z, sc_z)
        results.append(entry)

    # Write CSV
    out_csv = OUT / "per_row_parity.csv"
    with open(out_csv, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(results[0].keys()))
        w.writeheader(); w.writerows(results)
    print(f"wrote {out_csv} ({len(results)} rows)")

    # Summary
    from collections import Counter
    vs = Counter(r["verdict_status"] for r in results)
    vz = Counter(r["verdict_ziel"] for r in results)
    with open(OUT / "summary.md", "w", encoding="utf-8") as f:
        f.write("# §04 Renewable Section-Aware Parity — summary\n\n")
        f.write(f"Total Renewable rows compared: {len(results)}\n\n")
        f.write("## Status column verdict\n\n")
        for k, v in sorted(vs.items()):
            f.write(f"- {k}: {v}\n")
        f.write("\n## Ziel column verdict\n\n")
        for k, v in sorted(vz.items()):
            f.write(f"- {k}: {v}\n")

    # Discrepancies
    drift_rows = [r for r in results
                  if r["verdict_status"].startswith("DRIFT") or r["verdict_ziel"].startswith("DRIFT")]
    with open(OUT / "discrepancies.md", "w", encoding="utf-8") as f:
        f.write(f"# §04 Renewable Section-Aware Parity — discrepancies\n\n")
        f.write(f"Total DRIFT rows: {len(drift_rows)}\n\n")
        f.write("| code | name | unit | db_status | xl_status | drift_s | db_ziel | xl_ziel | drift_z |\n")
        f.write("|------|------|------|-----------|-----------|---------|---------|---------|---------|\n")
        for r in drift_rows:
            name = (r['db_name'] or "")[:40]
            ds = f"{r['drift_status']:.4f}" if r['drift_status'] is not None and r['drift_status'] != math.inf else "n/a"
            dz = f"{r['drift_ziel']:.4f}" if r['drift_ziel'] is not None and r['drift_ziel'] != math.inf else "n/a"
            f.write(f"| {r['db_code']} | {name} | {r['db_unit']} | {r['db_status']} | {r['xl_status']} | {ds} | {r['db_ziel']} | {r['xl_ziel']} | {dz} |\n")

    print(f"DRIFT rows: {len(drift_rows)}")

if __name__ == "__main__":
    main()
