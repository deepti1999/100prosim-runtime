"""Tests for manage.py import_excel_provenance (Phase A §2.3).

These tests skip when D.xlsx isn't available (CI-friendly), and exercise
the canonical Pascal-bundle file when it is. They do NOT use the full
seed; each test creates a small set of rows it knows the import command
should touch.
"""

from __future__ import annotations

import glob
import json
import os
import tempfile
from io import StringIO
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase
from openpyxl import Workbook

from simulator.models import (
    GebaeudewaermeData,
    LandUse,
    RenewableData,
    VerbrauchData,
)
from simulator.management.commands.import_excel_provenance import Command, _normalize_source_ref_code


_D_XLSX_CANDIDATES = sorted(glob.glob("docs/100prosim_d_*/D.xlsx"))
D_XLSX = _D_XLSX_CANDIDATES[0] if _D_XLSX_CANDIDATES else None


def _need_d(testcase: TestCase) -> None:
    if D_XLSX is None:
        testcase.skipTest("D.xlsx not present in docs/100prosim_d_*/ — skipping")


# ---------------------------------------------------------------------------
# Failure modes (no D.xlsx needed)
# ---------------------------------------------------------------------------


class TestImportFailsLoud(TestCase):
    """SR D3: import command fails loud on bad input."""

    def test_fails_loud_on_missing_file(self):
        with self.assertRaises(CommandError):
            call_command(
                "import_excel_provenance",
                "/definitely/does/not/exist.xlsx",
                stdout=StringIO(),
                stderr=StringIO(),
            )

    def test_fails_loud_on_non_xlsx_file(self):
        with tempfile.NamedTemporaryFile(suffix=".txt", delete=False) as f:
            f.write(b"this is not an xlsx workbook")
            path = f.name
        try:
            with self.assertRaises(CommandError):
                call_command("import_excel_provenance", path, stdout=StringIO(), stderr=StringIO())
        finally:
            os.unlink(path)


class TestExtractSourcesWorkbookLayouts(TestCase):
    """White-box coverage for 9.Quellen source extraction variants."""

    def test_extract_sources_maps_adjacent_url_row_back_to_description_ref(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "9.Quellen"

        # Split layout observed in D.xlsm:
        # row 1 = description, row 2 = URL-only continuation
        ws.cell(row=1, column=2).value = "9.224"
        ws.cell(row=1, column=4).value = (
            'STATISTISCHE ÄMTER DES BUNDES UND DER LÄNDER: "Regionaldatenbank Deutschland"; '
            "Online-Angebot GENESIS."
        )
        ws.cell(row=2, column=2).value = "9.225"
        ws.cell(row=2, column=4).value = "https://www.regionalstatistik.de/genesis/online/logon"

        sources = Command()._extract_sources(ws)

        self.assertEqual(
            sources.get("9.224"),
            "https://www.regionalstatistik.de/genesis/online/logon",
            "description row must inherit the URL stored on the next row",
        )
        self.assertEqual(
            sources.get("9.225"),
            "https://www.regionalstatistik.de/genesis/online/logon",
            "URL row should still map to its own code",
        )

    def test_extract_sources_falls_back_to_d_prefixed_code_when_cached_col2_is_not_9_ref(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "9.Quellen"

        ws.cell(row=1, column=1).value = "D.9.4"
        ws.cell(row=1, column=2).value = 6
        ws.cell(row=1, column=4).value = "AGENTUR ERNEUERBARE ENERGIEN: Quelle"
        ws.cell(row=2, column=1).value = "D.9.5"
        ws.cell(row=2, column=2).value = "9.5"
        ws.cell(row=2, column=4).value = "https://example.com/aee"

        catalog = Command()._extract_source_catalog(ws)

        self.assertEqual(catalog["9.4"]["description"], "AGENTUR ERNEUERBARE ENERGIEN: Quelle")
        self.assertEqual(catalog["9.4"]["url"], "https://example.com/aee")

    def test_extract_source_refs_accepts_9_ref_with_page_suffix_inside_brackets(self):
        refs = Command()._extract_source_refs(
            'Quelle laut Studie [9.182 Seite 9] und Potenzial [9.85, S. 21].',
            {
                "9.182": {"description": "Freiflächen-Leitfaden", "url": "https://example.com/182"},
                "9.85": {"description": "Wind-Potenzial", "url": "https://example.com/85"},
            },
        )

        self.assertEqual(
            refs,
            [
                {"code": "9.182", "description": "Freiflächen-Leitfaden", "url": "https://example.com/182"},
                {"code": "9.85", "description": "Wind-Potenzial", "url": "https://example.com/85"},
            ],
        )


class TestSourceRefNormalization(TestCase):
    def test_normalize_source_ref_code_uses_col2_when_it_is_already_canonical(self):
        self.assertEqual(_normalize_source_ref_code("9.224", "D.9.224"), "9.224")

    def test_normalize_source_ref_code_falls_back_to_col1_d_prefix(self):
        self.assertEqual(_normalize_source_ref_code(6, "D.9.4"), "9.4")

    def test_fails_loud_on_xlsx_missing_required_sheets(self):
        from openpyxl import Workbook

        wb = Workbook()
        wb.active.title = "wrong_sheet_name"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        wb.save(path)
        try:
            with self.assertRaises(CommandError):
                call_command("import_excel_provenance", path, stdout=StringIO(), stderr=StringIO())
        finally:
            os.unlink(path)


# ---------------------------------------------------------------------------
# Dry-run is read-only
# ---------------------------------------------------------------------------


class TestImportDryRun(TestCase):
    def setUp(self):
        _need_d(self)
        self.lu = LandUse.objects.create(
            code="LU_0",
            name="Bodenfläche gesamt",
            status_ha=35759529.0,
            target_ha=35759529.0,
        )

    def test_dry_run_default_does_not_write_anything(self):
        # No --apply means dry-run.
        out = StringIO()
        call_command("import_excel_provenance", D_XLSX, stdout=out, stderr=StringIO())
        self.lu.refresh_from_db()
        self.assertIsNone(self.lu.source_url, "dry-run must not populate source_url")
        self.assertIsNone(self.lu.notes_assumption, "dry-run must not populate notes_assumption")
        # origin is 'internal' (the model default), not yet 'd_xlsx'
        self.assertEqual(self.lu.origin, "internal")
        # Output should make it clear nothing was written
        self.assertIn("dry-run", out.getvalue().lower())


# ---------------------------------------------------------------------------
# Apply writes provenance for matched rows
# ---------------------------------------------------------------------------


class TestImportApply(TestCase):
    def setUp(self):
        _need_d(self)
        self.lu = LandUse.objects.create(
            code="LU_0",
            name="Bodenfläche gesamt",
            status_ha=35759529.0,
            target_ha=35759529.0,
        )
        self.lu_wald = LandUse.objects.create(
            code="LU_3",
            name="Waldfläche",
            status_ha=11538455.0,
            target_ha=11538455.0,
        )
        self.lu_ohne_forst = LandUse.objects.create(
            code="LU_3.2",
            name="(ohne forstwirtsch.Nutzung)",
            status_ha=1671822.0,
            target_ha=2165455.0,
            parent=self.lu_wald,
            quelle="[27][28]",
        )
        self.lu_dach = LandUse.objects.create(
            code="LU_1.1",
            name="Solare Dachflächen",
            status_ha=34243.0,
            target_ha=199398.0,
        )

    def test_apply_classifies_known_row_as_d_xlsx(self):
        call_command("import_excel_provenance", D_XLSX, "--apply", stdout=StringIO(), stderr=StringIO())
        self.lu.refresh_from_db()
        self.assertEqual(self.lu.origin, "d_xlsx", "Bodenfläche gesamt is in D.xlsx — origin must be d_xlsx")

    def test_apply_populates_notes_assumption(self):
        call_command("import_excel_provenance", D_XLSX, "--apply", stdout=StringIO(), stderr=StringIO())
        self.lu.refresh_from_db()
        self.assertIsNotNone(self.lu.notes_assumption, "matched row must get assumption text")
        # Bodenfläche gesamt has STATUS-Ansatz / ZIEL-Ansatz adjacent rows
        self.assertGreater(len(self.lu.notes_assumption or ""), 30, "assumption text should be substantive")

    def test_apply_populates_source_refs_from_9_quellen(self):
        call_command("import_excel_provenance", D_XLSX, "--apply", stdout=StringIO(), stderr=StringIO())
        self.lu.refresh_from_db()
        self.assertTrue(self.lu.source_refs, "matched row should get structured 9.Quellen refs")
        first_ref = self.lu.source_refs[0]
        self.assertEqual(first_ref["code"], "9.224")
        self.assertIn("Regionaldatenbank Deutschland", first_ref.get("description") or "")
        self.assertEqual(
            first_ref.get("url"),
            "https://www.regionalstatistik.de/genesis/online/logon",
        )

    def test_apply_preserves_status_and_ziel_ref_context_from_d_sheet(self):
        call_command("import_excel_provenance", D_XLSX, "--apply", stdout=StringIO(), stderr=StringIO())
        self.lu_dach.refresh_from_db()
        refs = self.lu_dach.source_refs or []
        self.assertTrue(refs, "Solare Dachflächen should get D-sheet source refs")
        self.assertTrue(
            any(ref.get("section") == "status" and ref.get("label") == "Solarthermie Kollektorfläche" for ref in refs),
            f"expected status-side source label from D-sheet, got: {refs}",
        )
        self.assertTrue(
            any(ref.get("section") == "status" and ref.get("label") == "PV-Leistung bauliche Anl." for ref in refs),
            f"expected second status-side source label from D-sheet, got: {refs}",
        )
        self.assertTrue(
            any(ref.get("section") == "ziel" and ref.get("code") == "9.17" for ref in refs),
            f"expected ziel-side source ref 9.17 from D-sheet, got: {refs}",
        )
        self.assertTrue(
            any(ref.get("section") == "ziel" and ref.get("code") == "9.19" for ref in refs),
            f"expected ziel-side source ref 9.19 from D-sheet, got: {refs}",
        )

    def test_landuse_internal_quelle_rows_inherit_from_parent_instead_of_wrong_direct_match(self):
        call_command("import_excel_provenance", D_XLSX, "--apply", stdout=StringIO(), stderr=StringIO())
        self.lu_wald.refresh_from_db()
        self.lu_ohne_forst.refresh_from_db()

        self.assertEqual(self.lu_wald.origin, "d_xlsx")
        self.assertEqual(
            self.lu_ohne_forst.origin,
            "derived",
            "rows with bracket-only internal Quelle refs must inherit from parent, not get an unrelated direct D-sheet match",
        )
        self.assertIsNone(
            self.lu_ohne_forst.notes_assumption,
            "derived LandUse rows must not get a direct foreign note block",
        )
        self.assertTrue(self.lu_ohne_forst.source_refs, "derived LandUse row should still inherit source refs")
        self.assertEqual(
            self.lu_ohne_forst.source_refs[0].get("code"),
            self.lu_wald.source_refs[0].get("code"),
        )

    def test_apply_does_not_touch_value_columns(self):
        """Pascal HARD rule: provenance import is value-neutral."""
        before = (self.lu.status_ha, self.lu.target_ha)
        call_command("import_excel_provenance", D_XLSX, "--apply", stdout=StringIO(), stderr=StringIO())
        self.lu.refresh_from_db()
        after = (self.lu.status_ha, self.lu.target_ha)
        self.assertEqual(before, after, f"value columns must not change: before={before} after={after}")

    def test_apply_does_not_rename_code(self):
        """SR-007: no rename."""
        before_code = self.lu.code
        call_command("import_excel_provenance", D_XLSX, "--apply", stdout=StringIO(), stderr=StringIO())
        self.lu.refresh_from_db()
        self.assertEqual(self.lu.code, before_code, "code must not change during import")

    def test_apply_classifies_unmatched_row_as_internal(self):
        # Create a row whose label doesn't match anything in D.xlsx
        unmatched = LandUse.objects.create(
            code="LU_FAKE_TEST",
            name="zZz_FakeUnmatchedLabel_zZz",
            status_ha=1.0,
            target_ha=1.0,
        )
        call_command("import_excel_provenance", D_XLSX, "--apply", stdout=StringIO(), stderr=StringIO())
        unmatched.refresh_from_db()
        self.assertEqual(unmatched.origin, "internal", "unmatched row must be classified internal")
        self.assertIsNone(unmatched.source_url)
        self.assertIsNone(unmatched.notes_assumption)


# ---------------------------------------------------------------------------
# Idempotency
# ---------------------------------------------------------------------------


class TestImportIdempotent(TestCase):
    def setUp(self):
        _need_d(self)
        LandUse.objects.create(
            code="LU_0",
            name="Bodenfläche gesamt",
            status_ha=35759529.0,
            target_ha=35759529.0,
        )

    def test_second_apply_run_reports_zero_changed(self):
        call_command("import_excel_provenance", D_XLSX, "--apply", stdout=StringIO(), stderr=StringIO())
        out2 = StringIO()
        call_command("import_excel_provenance", D_XLSX, "--apply", stdout=out2, stderr=StringIO())
        # Second run should report no changes
        text = out2.getvalue().lower()
        self.assertIn("0 changed", text, f"second run must report 0 changed, output was:\n{out2.getvalue()}")


# ---------------------------------------------------------------------------
# Manifest + orphan CSV
# ---------------------------------------------------------------------------


class TestImportArtifacts(TestCase):
    def setUp(self):
        _need_d(self)
        # Matched row (in D.xlsx)
        LandUse.objects.create(
            code="LU_0",
            name="Bodenfläche gesamt",
            status_ha=35759529.0,
            target_ha=35759529.0,
        )
        # Unmatched row (forces an orphan CSV entry)
        LandUse.objects.create(
            code="LU_FAKE_ART",
            name="zZzFakeArtefactRowzZz",
            status_ha=1.0,
            target_ha=1.0,
        )

    def test_writes_manifest_after_apply(self):
        # Phase B (T65): manifest is per-region; default region=DE.
        manifest_path = Path("data/import/DE/d_xlsx.manifest.json")
        if manifest_path.exists():
            manifest_path.unlink()
        call_command("import_excel_provenance", D_XLSX, "--apply", stdout=StringIO(), stderr=StringIO())
        self.assertTrue(manifest_path.exists(), "manifest must be written after --apply")
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        self.assertIn("files", manifest)
        self.assertGreater(len(manifest["files"]), 0)
        f0 = manifest["files"][0]
        self.assertIn("file_hash", f0)
        self.assertIn("sheet_hashes", f0)
        self.assertEqual(f0["region_code"], "DE")

    def test_writes_orphan_csv_after_apply(self):
        # Phase B (T65): orphan CSV is per-region; default region=DE.
        orphan_path = Path("data/import/DE/orphan_classification.csv")
        if orphan_path.exists():
            orphan_path.unlink()
        call_command("import_excel_provenance", D_XLSX, "--apply", stdout=StringIO(), stderr=StringIO())
        self.assertTrue(orphan_path.exists(), "orphan CSV must be written after --apply")
        # Should have a header line + at least one data row
        text = orphan_path.read_text(encoding="utf-8").splitlines()
        self.assertGreater(len(text), 1)


# ---------------------------------------------------------------------------
# User workspace rows untouched (SR-005)
# ---------------------------------------------------------------------------


class TestImportPreservesUserWorkspace(TestCase):
    def setUp(self):
        _need_d(self)
        User = get_user_model()
        self.user = User.objects.create_user(username="testimportuser", password="x")
        # Base row (owner=NULL)
        self.base = LandUse.objects.create(
            code="LU_0",
            name="Bodenfläche gesamt",
            status_ha=35759529.0,
            target_ha=35759529.0,
        )
        # User-scoped row (owner=user, same code)
        self.user_row = LandUse.all_objects.create(
            owner=self.user,
            code="LU_0",
            name="User-modified copy",
            status_ha=99999.0,
        )

    def test_user_workspace_value_untouched_but_provenance_propagated(self):
        """SR-005: user's VALUE override stays. But the 3 provenance columns
        propagate from the matching base row by code (provenance describes the
        underlying parameter, not the user's edit)."""
        call_command("import_excel_provenance", D_XLSX, "--apply", stdout=StringIO(), stderr=StringIO())
        self.user_row.refresh_from_db()
        self.base.refresh_from_db()
        # Value untouched (the hard SR-005 invariant)
        self.assertEqual(self.user_row.status_ha, 99999.0)
        # Provenance copied from base (since both share code='LU_0')
        self.assertEqual(self.user_row.origin, self.base.origin)
        self.assertEqual(self.user_row.notes_assumption, self.base.notes_assumption)
        self.assertEqual(self.user_row.source_url, self.base.source_url)
        self.assertEqual(self.user_row.source_refs, self.base.source_refs)
