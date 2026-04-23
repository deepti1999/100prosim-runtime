"""Phase B (T65) — region-aware import_excel_provenance tests.

Verifies:
- Default xlsx_path resolves to data/import/<region>/D.xlsx so per-
  region Excel files live under their own subdir.
- Manifest + orphan CSV write to data/import/<region>/.
- Unknown region in DB fails loud.
- Provenance write only touches base rows of the requested region.
- Workspace propagation only touches user rows of the same region.
"""
from __future__ import annotations

import glob
import json
import tempfile
from io import StringIO
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase

from simulator.models import LandUse, Region


_D_XLSX_CANDIDATES = sorted(glob.glob("docs/100prosim_d_*/D.xlsx"))
D_XLSX = _D_XLSX_CANDIDATES[0] if _D_XLSX_CANDIDATES else None


def _need_d(testcase: TestCase) -> None:
    if D_XLSX is None:
        testcase.skipTest("D.xlsx not present in docs/100prosim_d_*/ — skipping")


class TestImportRegionFlagValidation(TestCase):
    """Region must exist in DB before the command will run."""

    def test_unknown_region_code_fails_loud(self):
        # Even with a valid xlsx path, a non-existent Region.code aborts.
        from openpyxl import Workbook

        wb = Workbook()
        wb.create_sheet("1.")
        wb.create_sheet("9.Quellen")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        wb.save(path)
        try:
            with self.assertRaises(CommandError):
                call_command(
                    "import_excel_provenance",
                    path,
                    "--region=ZZ_NOT_REAL",
                    stdout=StringIO(),
                    stderr=StringIO(),
                )
        finally:
            Path(path).unlink(missing_ok=True)

    def test_default_xlsx_path_uses_region_subdir(self):
        """When xlsx_path is omitted, command resolves to data/import/<region>/D.xlsx."""
        # Create a region whose folder doesn't have D.xlsx yet — we expect a
        # "file not found" error specifically.
        Region.objects.create(code="ZZTESTREGION", display_name="ZZ Test", active=True)
        try:
            call_command(
                "import_excel_provenance",
                "--region=ZZTESTREGION",
                stdout=StringIO(),
                stderr=StringIO(),
            )
            self.fail("Expected CommandError when default file missing")
        except CommandError as e:
            # The error message should mention the resolved per-region path
            self.assertIn("ZZTESTREGION", str(e), f"Got: {e}")


class TestImportArtifactsPerRegion(TestCase):
    """Manifest and orphan CSV write to data/import/<region>/."""

    def setUp(self):
        _need_d(self)
        # One matched + one orphan row so manifest + CSV both populate.
        LandUse.objects.create(
            code="LU_0",
            name="Bodenfläche gesamt",
            status_ha=35759529.0,
            target_ha=35759529.0,
        )
        LandUse.objects.create(
            code="LU_FAKE_REGION_TEST",
            name="zZzFakeRegionRowzZz",
            status_ha=1.0,
            target_ha=1.0,
        )

    def test_manifest_at_region_path(self):
        manifest_path = Path("data/import/DE/d_xlsx.manifest.json")
        # Don't unlink — we want the test to actually exercise the write path.
        call_command(
            "import_excel_provenance",
            D_XLSX,
            "--apply",
            "--region=DE",
            stdout=StringIO(),
            stderr=StringIO(),
        )
        self.assertTrue(
            manifest_path.exists(),
            f"manifest must be written to {manifest_path}",
        )
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        self.assertEqual(manifest["files"][0]["region_code"], "DE")

    def test_orphan_csv_at_region_path(self):
        orphan_path = Path("data/import/DE/orphan_classification.csv")
        call_command(
            "import_excel_provenance",
            D_XLSX,
            "--apply",
            "--region=DE",
            stdout=StringIO(),
            stderr=StringIO(),
        )
        self.assertTrue(
            orphan_path.exists(),
            f"orphan CSV must be written to {orphan_path}",
        )


class TestImportFiltersByRegion(TestCase):
    """Provenance write only touches base rows of the requested region."""

    def setUp(self):
        _need_d(self)
        self.de = Region.objects.get(code="DE")
        self.bb = Region.objects.create(code="BB", display_name="Brandenburg", active=True)
        # DE row matched in D.xlsx
        self.de_row = LandUse.all_objects.create(
            code="LU_0",
            name="Bodenfläche gesamt",
            status_ha=35759529.0,
            target_ha=35759529.0,
            region=self.de,
        )
        # BB row with the SAME label that D.xlsx would normally match.
        self.bb_row = LandUse.all_objects.create(
            code="LU_0",
            name="Bodenfläche gesamt",
            status_ha=999.0,
            target_ha=999.0,
            region=self.bb,
        )

    def test_only_DE_row_touched_when_region_DE(self):
        call_command(
            "import_excel_provenance",
            D_XLSX,
            "--apply",
            "--region=DE",
            stdout=StringIO(),
            stderr=StringIO(),
        )
        self.de_row.refresh_from_db()
        self.bb_row.refresh_from_db()
        self.assertEqual(self.de_row.origin, "d_xlsx", "DE row should be classified d_xlsx")
        self.assertEqual(
            self.bb_row.origin,
            "internal",
            "BB row must NOT be touched when --region=DE",
        )


class TestWorkspacePropagationPerRegion(TestCase):
    """Workspace propagation only touches user rows of the same region."""

    def setUp(self):
        _need_d(self)
        self.de = Region.objects.get(code="DE")
        self.bb = Region.objects.create(code="BB", display_name="Brandenburg", active=True)
        # Base DE + base BB
        LandUse.all_objects.create(
            code="LU_0",
            name="Bodenfläche gesamt",
            status_ha=35759529.0,
            target_ha=35759529.0,
            region=self.de,
        )
        LandUse.all_objects.create(
            code="LU_0",
            name="Bodenfläche gesamt",
            status_ha=1.0,
            target_ha=1.0,
            region=self.bb,
        )
        # User has both DE and BB workspace rows for the same code
        User = get_user_model()
        self.user = User.objects.create_user(
            username="phaseb_propagate_user", password="x"
        )
        self.user_de = LandUse.all_objects.create(
            owner=self.user, region=self.de, code="LU_0", name="user DE", status_ha=42.0
        )
        self.user_bb = LandUse.all_objects.create(
            owner=self.user, region=self.bb, code="LU_0", name="user BB", status_ha=84.0
        )

    def test_only_user_DE_row_gets_propagated_when_region_DE(self):
        call_command(
            "import_excel_provenance",
            D_XLSX,
            "--apply",
            "--region=DE",
            stdout=StringIO(),
            stderr=StringIO(),
        )
        self.user_de.refresh_from_db()
        self.user_bb.refresh_from_db()
        # DE workspace row got the d_xlsx provenance from base DE
        self.assertEqual(self.user_de.origin, "d_xlsx")
        # BB workspace row stayed at internal default
        self.assertEqual(self.user_bb.origin, "internal")
        # Hard SR-005: values untouched on both
        self.assertEqual(self.user_de.status_ha, 42.0)
        self.assertEqual(self.user_bb.status_ha, 84.0)
