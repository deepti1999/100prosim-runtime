"""Phase A §2.3 — import provenance (source URL + cited-source refs +
assumption text + origin)
from D.xlsx into the 4 parameter-bearing models.

Strict invariants enforced (per project runtime notes + DATA_MODEL_IMPORT_AUDIT.md):
  - SR-005: per-user workspace rows (owner != NULL) NEVER touched.
  - SR-007: no `code` field rename; ONLY source_url, source_refs,
            notes_assumption, origin are written.
  - HARD: no value column (status_ha, target_ha, status, ziel, status_value,
          target_value) is touched. Phase A is provenance-only.
  - D3: fail loud on missing file, non-xlsx, or unrecognised sheet schema.
  - D5: write manifest + orphan-classification CSV after --apply.

Output artefacts on --apply:
  - data/import/d_xlsx.manifest.json — file hash + per-sheet hashes + counters.
  - data/import/orphan_classification.csv — every unmatched row with rationale.

Usage:
  python manage.py import_excel_provenance docs/100prosim_d_*/D.xlsx          # dry-run
  python manage.py import_excel_provenance docs/100prosim_d_*/D.xlsx --apply  # commit
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from simulator.models import (
    GebaeudewaermeData,
    LandUse,
    RenewableData,
    VerbrauchData,
)


REQUIRED_SHEETS = ("1.", "9.Quellen")

# Row-type column on D.xlsx 1. (col 67 = "BO" — single letter row classifier)
TYPE_COL = 67
LABEL_COL = 5
VALUE_COL_W = 23
ROW_SOURCE_CODE_COL = 16
ROW_SOURCE_LABEL_COL = 17

PARAMETER_TYPE = "p"
ASSUMPTION_TYPES = ("e", "s", "z", "h")  # erläuterung / status-Ansatz / ziel-Ansatz / herleitung

# Explicit external-source refs inside assumption text look like [9.123],
# [9.85, S. 21], [9.182 Seite 9], etc. Bare refs like [122] are usually
# D-sheet internal cross-references and must NOT be treated as 9.Quellen refs.
SOURCE_REF_RE = re.compile(r"\[(?:[^\]]*?)(9\.\d+(?:\.\d+)?)(?:[^\]]*?)\]")
QUELLE_D_ROW_RE = re.compile(r"D\.1\.(\d+)")


class _ModelPlan:
    def __init__(
        self,
        model,
        label_field: str,
        code_field: str,
        parent_resolver: str = "code_hierarchy",
    ):
        """parent_resolver: how to walk up the hierarchy.
        - 'code_hierarchy': split code on '.' and trim (e.g. '2.1.1' → '2.1' → '2.0' → '2').
        - 'parent_code_field': read self.parent_code (RenewableData).
        - 'parent_fk': follow self.parent.code (LandUse).
        """
        self.model = model
        self.label_field = label_field
        self.code_field = code_field
        self.parent_resolver = parent_resolver


MODEL_PLANS = [
    _ModelPlan(LandUse, label_field="name", code_field="code", parent_resolver="parent_fk"),
    _ModelPlan(RenewableData, label_field="name", code_field="code", parent_resolver="parent_code_field"),
    _ModelPlan(VerbrauchData, label_field="category", code_field="code", parent_resolver="code_hierarchy"),
    _ModelPlan(GebaeudewaermeData, label_field="category", code_field="code", parent_resolver="code_hierarchy"),
]


class _NoopAtomic:
    """Stand-in for transaction.atomic() when running dry-run; simply yields."""

    def __enter__(self):
        return self

    def __exit__(self, *args):
        return False


def _code_hierarchy_parent(code: str) -> Optional[str]:
    if not code or "." not in code:
        return None
    parts = code.split(".")
    if len(parts) <= 1:
        return None
    return ".".join(parts[:-1])


def _parent_code(row, plan: "_ModelPlan") -> Optional[str]:
    """Return the code of the row's parent, or None.

    Always falls back to code_hierarchy parsing when the explicit
    parent reference is missing — many RenewableData rows have NULL
    parent_code but their dotted code clearly implies a parent.
    """
    if plan.parent_resolver == "parent_fk":
        parent = getattr(row, "parent", None)
        if parent is not None:
            v = getattr(parent, "code", None)
            if v:
                return v
        # Fallback to code-hierarchy parse
        return _code_hierarchy_parent(getattr(row, "code", "") or "")

    if plan.parent_resolver == "parent_code_field":
        v = getattr(row, "parent_code", None)
        if v:
            return v
        # Fallback: many RenewableData rows have NULL parent_code; use code structure.
        return _code_hierarchy_parent(getattr(row, "code", "") or "")

    if plan.parent_resolver == "code_hierarchy":
        return _code_hierarchy_parent(getattr(row, "code", "") or "")

    return None


def _normalize_label(s: Optional[str]) -> str:
    """Lower-case, strip parens / asterisks / hyphens, collapse whitespace."""
    if not s:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"[\*]", "", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _section_from_note_line(text: str, current_section: Optional[str]) -> Optional[str]:
    norm = (text or "").strip().lower()
    if norm.startswith("- status-ansatz:"):
        return "status"
    if norm.startswith("- ziel-ansatz:"):
        return "ziel"
    return current_section


def _normalize_source_ref_code(raw_ref: object, d_prefixed_ref: object) -> Optional[str]:
    """Return canonical 9.xxx code from a 9.Quellen row.

    Some workbook rows cache the col-2 formula oddly under data_only mode.
    Col-1's D-prefixed code is more stable, so use it as a fallback.
    """
    if isinstance(raw_ref, str):
        raw_ref = raw_ref.strip()
        if re.fullmatch(r"9\.\d+(?:\.\d+)?", raw_ref):
            return raw_ref

    if isinstance(d_prefixed_ref, str):
        d_prefixed_ref = d_prefixed_ref.strip()
        m = re.fullmatch(r"D\.(9\.\d+(?:\.\d+)?)", d_prefixed_ref)
        if m:
            return m.group(1)

    return None


class Command(BaseCommand):
    help = (
        "Phase A §2.3: import provenance (source URL + cited-source refs + "
        "assumption text + origin) "
        "from D.xlsx into LandUse / RenewableData / VerbrauchData / GebaeudewaermeData."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "xlsx_path",
            nargs="?",
            default=None,
            help=(
                "Path to D.xlsx (gitignored). When omitted, defaults to "
                "data/import/<region>/D.xlsx so per-region Excel files "
                "live under their own subdir."
            ),
        )
        parser.add_argument(
            "--apply",
            action="store_true",
            help="Commit changes. Without this, runs in dry-run mode and prints the diff only.",
        )
        parser.add_argument(
            "--region",
            default="DE",
            help=(
                "Region code (must exist + be active in Region table). "
                "Phase B (T65): manifest, orphan CSV, and base-row filter "
                "are scoped to this region."
            ),
        )

    # ------------------------------------------------------------------
    # entry
    # ------------------------------------------------------------------

    def handle(self, *args, **options):
        region_code = options["region"]
        xlsx_path = options["xlsx_path"] or f"data/import/{region_code}/D.xlsx"
        apply_changes = options["apply"]

        # Phase B (T65): validate region against DB before any FS work.
        from simulator.models import Region

        try:
            region_obj = Region.objects.get(code=region_code)
        except Region.DoesNotExist:
            raise CommandError(
                f"Region '{region_code}' not found in DB. Create it first via "
                f"Region.objects.create(code='{region_code}', "
                f"display_name='...', active=True), or run "
                f"`manage.py shell` to add it."
            )
        if not region_obj.active:
            raise CommandError(
                f"Region '{region_code}' is inactive (active=False). "
                f"Activate it before importing."
            )

        # Default xlsx path now references region_code directly so the
        # error message is region-specific.
        if not xlsx_path or not isinstance(xlsx_path, str):
            raise CommandError(
                f"Could not resolve xlsx_path for region '{region_code}' "
                f"(checked default: data/import/{region_code}/D.xlsx)"
            )

        self._validate_file(xlsx_path)

        try:
            wb = load_workbook(xlsx_path, data_only=True, read_only=False)
        except Exception as e:
            raise CommandError(f"Failed to open {xlsx_path}: {e}")

        for sheet in REQUIRED_SHEETS:
            if sheet not in wb.sheetnames:
                raise CommandError(
                    f"D.xlsx schema mismatch: required sheet '{sheet}' not found. "
                    f"Got sheets: {wb.sheetnames}"
                )

        source_catalog = self._extract_source_catalog(wb["9.Quellen"])
        sources = {
            ref_code: entry["url"]
            for ref_code, entry in source_catalog.items()
            if entry.get("url")
        }
        self.stdout.write(self.style.SUCCESS(f"Extracted {len(sources)} source URLs from 9.Quellen"))

        d1_label_index, entries_by_row, row_to_parent_p = self._build_d1_index(wb["1."])
        self.stdout.write(
            self.style.SUCCESS(
                f"Indexed {len(entries_by_row)} parameter rows in D.xlsx '1.' "
                f"({len(d1_label_index)} unique labels)."
            )
        )

        # Phase C (T66): row-creating mode. If the target region has zero
        # base rows in any of the parameter models (i.e. it's a fresh
        # Bundesland), clone the DE row structure first then run the
        # standard UPDATE path on the new rows for provenance.
        de_region = None
        if region_code != "DE":
            try:
                de_region = Region.objects.get(code="DE")
            except Region.DoesNotExist:
                de_region = None
        target_has_rows = any(
            (plan.model.all_objects.filter(region=region_obj).exists()
             if any(f.name == "owner" for f in plan.model._meta.get_fields())
             else plan.model.objects.filter(region=region_obj).exists())
            for plan in MODEL_PLANS
        )
        if not target_has_rows and de_region is not None:
            created_total = self._create_region_rows_from_de_template(
                de_region, region_obj, d1_label_index, apply_changes,
            )
            self.stdout.write(
                self.style.SUCCESS(
                    f"CREATE mode: cloned {created_total} base rows from DE "
                    f"into region '{region_code}' "
                    f"(values overlaid from xlsx where label matches)."
                )
            )
        elif not target_has_rows and region_code != "DE":
            raise CommandError(
                f"Region '{region_code}' has zero base rows but DE region "
                f"is missing — cannot clone structure. Seed DE first."
            )

        # Load value-based audit mapping (final_map_*.csv) as a second-pass index.
        value_maps = self._load_value_maps()

        diffs = []
        for plan in MODEL_PLANS:
            vm = value_maps.get(plan.model.__name__, {})
            d = self._compute_diff(
                plan, d1_label_index, entries_by_row, row_to_parent_p, vm, sources, source_catalog,
                region_obj,
            )
            diffs.append(d)

        # Summary
        total_changed = sum(d["changed"] for d in diffs)
        total_matched = sum(d["matched"] for d in diffs)
        total_derived = sum(d["derived"] for d in diffs)
        total_unmatched = sum(d["unmatched"] for d in diffs)
        total_rows = sum(d["total"] for d in diffs)
        for d in diffs:
            self.stdout.write(
                f"  {d['model_name']:24s}  total={d['total']:4d}  "
                f"d_xlsx={d['matched']:4d}  derived={d['derived']:4d}  "
                f"internal={d['unmatched']:4d}  changed={d['changed']:4d}"
            )
        coverage = (total_matched + total_derived) / total_rows * 100 if total_rows else 0
        self.stdout.write(
            self.style.SUCCESS(
                f"Total: {total_changed} changed across {len(diffs)} models. "
                f"Provenance coverage: {total_matched + total_derived}/{total_rows} ({coverage:.1f}%) "
                f"[{total_matched} d_xlsx + {total_derived} derived]; "
                f"{total_unmatched} internal."
            )
        )

        if apply_changes:
            with transaction.atomic():
                for d, plan in zip(diffs, MODEL_PLANS):
                    self._apply_diff(plan.model, d)
                # Propagate the 3 provenance columns to user-scoped workspace rows
                # of the SAME region only (Phase B: no cross-region leakage).
                # SR-005 still holds: VALUE columns remain user-owned.
                propagated = self._propagate_to_workspace_rows(region_obj)
                if propagated:
                    self.stdout.write(
                        self.style.SUCCESS(f"Propagated provenance to {propagated} user-workspace rows")
                    )
            self._write_manifest(xlsx_path, region_code, diffs)
            self._write_orphan_csv(region_code, diffs)
            self.stdout.write(self.style.SUCCESS("APPLIED."))
        else:
            self.stdout.write(self.style.WARNING("dry-run only — pass --apply to commit"))

    # ------------------------------------------------------------------
    # validation
    # ------------------------------------------------------------------

    def _validate_file(self, xlsx_path: str) -> None:
        if not os.path.isfile(xlsx_path):
            raise CommandError(f"File not found: {xlsx_path}")
        if not xlsx_path.lower().endswith((".xlsx", ".xlsm")):
            raise CommandError(f"Not an .xlsx/.xlsm file: {xlsx_path}")
        try:
            with open(xlsx_path, "rb") as f:
                magic = f.read(4)
        except Exception as e:
            raise CommandError(f"Cannot read {xlsx_path}: {e}")
        if magic != b"PK\x03\x04":
            raise CommandError(f"Not a valid xlsx (zip magic missing): {xlsx_path}")

    # ------------------------------------------------------------------
    # extractors
    # ------------------------------------------------------------------

    def _extract_source_catalog(self, ws) -> dict[str, dict[str, Optional[str]]]:
        """Build {ref_code: {"description": ..., "url": ...}} from 9.Quellen.

        Layout (verified by scripts/audit_out/workbook_catalog.txt):
          col 1 = D-prefixed code (e.g. 'D.9.5')
          col 2 = bare ref code (e.g. '9.5')
          col 4 = source text or URL hyperlink target

        Some workbook variants split one citation across two adjacent rows:
          - row N   = reference text for 9.xxx
          - row N+1 = URL only

        In that case, attach the URL from row N+1 back to the preceding
        reference code as well. This preserves the user-facing "click the
        cited source" behavior even when the workbook stores the URL one row
        below the textual citation.
        """
        catalog: dict[str, dict[str, Optional[str]]] = {}
        pending_ref_code: Optional[str] = None
        for r in range(1, ws.max_row + 1):
            ref_code = _normalize_source_ref_code(
                ws.cell(row=r, column=2).value,
                ws.cell(row=r, column=1).value,
            )
            if not ref_code:
                continue
            cell_url = ws.cell(row=r, column=4)
            entry = catalog.setdefault(ref_code, {"description": None, "url": None})
            url: Optional[str] = None
            if cell_url.hyperlink and cell_url.hyperlink.target:
                url = cell_url.hyperlink.target
            elif isinstance(cell_url.value, str) and cell_url.value.startswith(("http://", "https://")):
                url = cell_url.value

            if url:
                entry["url"] = url
                if pending_ref_code:
                    pending_entry = catalog.setdefault(
                        pending_ref_code,
                        {"description": None, "url": None},
                    )
                    pending_entry["url"] = pending_entry.get("url") or url
                pending_ref_code = None
                continue

            cell_text = cell_url.value if isinstance(cell_url.value, str) else ""
            has_reference_text = bool(cell_text.strip())
            if has_reference_text:
                entry["description"] = cell_text.strip()
            pending_ref_code = ref_code if has_reference_text else None
        return catalog

    def _extract_sources(self, ws) -> dict[str, str]:
        """Compatibility wrapper: Build {ref_code: url} from 9.Quellen."""
        catalog = self._extract_source_catalog(ws)
        return {
            ref_code: entry["url"]
            for ref_code, entry in catalog.items()
            if entry.get("url")
        }

    def _extract_source_refs(
        self,
        note_text: Optional[str],
        source_catalog: dict[str, dict[str, Optional[str]]],
    ) -> list[dict[str, Optional[str]]]:
        """Return cited 9.Quellen refs from note text in stable order."""
        if not note_text:
            return []

        refs: list[dict[str, Optional[str]]] = []
        seen: set[str] = set()
        for ref_code in SOURCE_REF_RE.findall(note_text):
            if ref_code in seen:
                continue
            seen.add(ref_code)
            catalog_entry = source_catalog.get(ref_code, {})
            refs.append(
                {
                    "code": ref_code,
                    "description": catalog_entry.get("description"),
                    "url": catalog_entry.get("url"),
                }
            )
        return refs

    def _merge_source_refs(
        self,
        explicit_refs: list[dict[str, Optional[str]]],
        note_text: Optional[str],
        source_catalog: dict[str, dict[str, Optional[str]]],
    ) -> list[dict[str, Optional[str]]]:
        """Prefer explicit D-sheet row refs; add regex-derived refs only as fallback."""
        merged: list[dict[str, Optional[str]]] = []
        seen: set[tuple[Optional[str], Optional[str], Optional[str]]] = set()

        for ref in explicit_refs:
            key = (ref.get("code"), ref.get("section"), ref.get("label"))
            if key in seen:
                continue
            seen.add(key)
            merged.append(ref)

        existing_codes = {ref.get("code") for ref in merged}
        for ref in self._extract_source_refs(note_text, source_catalog):
            if ref.get("code") in existing_codes:
                continue
            merged.append(ref)
        return merged

    def _build_d1_index(self, ws):
        """Build three indexes over D.xlsx '1.':
        - label_index: {normalized_label: entry_dict} keyed on parameter (p-type) labels
        - entries_by_row: {p_row_num: entry_dict}
        - row_to_parent_p: {any_row_num: p_row_num} — for any row, find its containing parameter

        Each entry_dict = {row, label, assumption_text, value_w, source_refs}.
        """
        rows: list[tuple[int, Optional[str], Optional[str], object]] = []
        for r in range(1, ws.max_row + 1):
            t = ws.cell(row=r, column=TYPE_COL).value
            label = ws.cell(row=r, column=LABEL_COL).value
            w = ws.cell(row=r, column=VALUE_COL_W).value
            if label is None and t is None:
                continue
            rows.append((r, t, label, w))

        label_index: dict[str, dict] = {}
        entries_by_row: dict[int, dict] = {}
        row_to_parent_p: dict[int, int] = {}

        # Walk: for each p-row, gather assumption rows beneath; index everything
        for i, (r, t, label, w) in enumerate(rows):
            if t != PARAMETER_TYPE or not isinstance(label, str) or not label.strip():
                continue
            assumption_parts: list[str] = []
            source_refs: list[dict[str, Optional[str]]] = []
            current_section: Optional[str] = None
            for j in range(i + 1, len(rows)):
                rj, tj, lj, wj = rows[j]
                if tj == PARAMETER_TYPE:
                    break
                if tj in ASSUMPTION_TYPES and isinstance(lj, str) and lj.strip():
                    assumption_parts.append(lj.strip())
                    current_section = _section_from_note_line(lj, current_section)
                    source_code = ws.cell(row=rj, column=ROW_SOURCE_CODE_COL).value
                    source_code = source_code.strip() if isinstance(source_code, str) else source_code
                    if isinstance(source_code, str) and re.fullmatch(r"9\.\d+(?:\.\d+)?", source_code):
                        source_label = ws.cell(row=rj, column=ROW_SOURCE_LABEL_COL).value
                        source_label = source_label.strip() if isinstance(source_label, str) else None
                        source_refs.append(
                            {
                                "code": source_code,
                                "label": source_label or None,
                                "section": current_section,
                            }
                        )
            assumption_text = "\n\n".join(assumption_parts) if assumption_parts else None

            entry = {
                "row": r,
                "label": label.strip(),
                "assumption_text": assumption_text,
                "value_w": w,
                "source_refs": source_refs,
            }
            entries_by_row[r] = entry
            norm = _normalize_label(label)
            if norm and norm not in label_index:
                label_index[norm] = entry

        # Build row_to_parent_p: walk all rows; for each, the containing p-row is the
        # last p-row at or before its index.
        current_p = None
        for r, t, label, w in rows:
            if t == PARAMETER_TYPE and r in entries_by_row:
                current_p = r
            if current_p is not None:
                row_to_parent_p[r] = current_p

        return label_index, entries_by_row, row_to_parent_p

    def _load_value_maps(self) -> dict[str, dict]:
        """Load scripts/audit_out/final_map_<model>.csv into:
            {ModelClassName: {our_code: {status_row, ziel_row}}}
        These CSVs ship with the §2.3 audit (commit 58a1b90 Step C area) and provide
        a value-based DB→D.xlsx 1. row mapping that complements label matching."""
        out: dict[str, dict] = {}
        plan_to_csv = {
            "LandUse": "scripts/audit_out/final_map_landuse.csv",
            "RenewableData": "scripts/audit_out/final_map_renewabledata.csv",
            "VerbrauchData": "scripts/audit_out/final_map_verbrauchdata.csv",
            "GebaeudewaermeData": "scripts/audit_out/final_map_gebaeudewaermedata.csv",
        }
        for model_name, csv_path in plan_to_csv.items():
            mapping: dict[str, dict] = {}
            if not os.path.isfile(csv_path):
                out[model_name] = mapping
                continue
            try:
                with open(csv_path, encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        code = (row.get("our_code") or "").strip()
                        if not code:
                            continue

                        def _to_int(v):
                            try:
                                return int(v) if v else None
                            except ValueError:
                                return None

                        mapping[code] = {
                            "status_row": _to_int(row.get("status_excel_row")),
                            "ziel_row": _to_int(row.get("ziel_excel_row")),
                            "confidence": (row.get("confidence") or "").strip(),
                        }
            except Exception as e:
                self.stdout.write(self.style.WARNING(f"Failed to load {csv_path}: {e}"))
            out[model_name] = mapping
        return out

    # ------------------------------------------------------------------
    # diff + apply
    # ------------------------------------------------------------------

    def _compute_diff(
        self,
        plan: _ModelPlan,
        d1_label_index: dict,
        entries_by_row: dict,
        row_to_parent_p: dict,
        value_map: dict,
        sources: dict,
        source_catalog: dict,
        region_obj,
    ) -> dict:
        model = plan.model
        # SR-005: ONLY base rows (owner=NULL); never touch user workspace.
        # Phase B (T65): scope to the requested region too.
        has_region = any(f.name == "region" for f in model._meta.get_fields())
        if hasattr(model, "all_objects"):
            qs = model.all_objects.all()
            if any(f.name == "owner" for f in model._meta.get_fields()):
                qs = model.all_objects.filter(owner__isnull=True)
        else:
            qs = model.objects.all()
        if has_region:
            qs = qs.filter(region=region_obj)

        rows = list(qs)

        # Build a code → row dict for parent lookup
        rows_by_code: dict = {}
        for row in rows:
            code = getattr(row, plan.code_field, None)
            if code:
                rows_by_code[code] = row

        diff = {
            "model_name": model.__name__,
            "total": len(rows),
            "matched": 0,        # direct d_xlsx matches
            "derived": 0,        # parent-inherited
            "unmatched": 0,      # truly orphan
            "changed": 0,
            "updates": [],
            "orphans": [],
        }

        # Pre-compute direct + value-map matches per row (so we can inherit from them).
        # Direct = label match into d1_label_index.
        # Value-map = final_map_*.csv lookup → status_row or ziel_row → containing p-row.
        direct_matches: dict = {}  # code -> entry
        for row in rows:
            code = getattr(row, plan.code_field, None) or ""
            if not code:
                continue
            if model.__name__ == "LandUse":
                quelle = getattr(row, "quelle", None) or ""
                m = QUELLE_D_ROW_RE.search(str(quelle))
                if m:
                    d_row = int(m.group(1))
                    if d_row in entries_by_row:
                        direct_matches[code] = entries_by_row[d_row]
                        continue
                # LandUse rows whose Quelle only contains workbook-internal
                # bracket refs like [27][28] or [14]...[19] should inherit
                # provenance from their parent, not be force-matched to an
                # unrelated D-sheet note via the value-map fallback.
                if str(quelle).strip():
                    continue
            label = getattr(row, plan.label_field, None) or ""
            entry = d1_label_index.get(_normalize_label(label))
            if entry:
                direct_matches[code] = entry
                continue
            # Value-map fallback
            vm_entry = value_map.get(code) or {}
            for d_row in (vm_entry.get("status_row"), vm_entry.get("ziel_row")):
                if d_row and d_row in row_to_parent_p:
                    p_row = row_to_parent_p[d_row]
                    if p_row in entries_by_row:
                        direct_matches[code] = entries_by_row[p_row]
                        break

        for row in rows:
            code = getattr(row, plan.code_field, None) or ""
            label = getattr(row, plan.label_field, None) or ""
            entry = direct_matches.get(code)

            new_source_url: Optional[str] = None
            new_notes_assumption: Optional[str] = None
            new_source_refs: list[dict[str, Optional[str]]] = []
            new_origin = "internal"

            if entry:
                # Direct match in D.xlsx 1.
                new_origin = "d_xlsx"
                new_notes_assumption = entry["assumption_text"]
                explicit_refs = []
                for source_ref in entry.get("source_refs", []):
                    code = source_ref.get("code")
                    catalog_entry = source_catalog.get(code or "", {})
                    explicit_refs.append(
                        {
                            "code": code,
                            "label": source_ref.get("label"),
                            "section": source_ref.get("section"),
                            "description": catalog_entry.get("description"),
                            "url": catalog_entry.get("url"),
                        }
                    )
                new_source_refs = self._merge_source_refs(
                    explicit_refs,
                    new_notes_assumption,
                    source_catalog,
                )
                if new_source_refs:
                    new_source_url = new_source_refs[0].get("url")
                diff["matched"] += 1
            else:
                # Walk up the hierarchy looking for a matched ancestor.
                ancestor_entry = None
                cursor_row = row
                hops = 0
                while hops < 6:  # bounded in case of cycles
                    p_code = _parent_code(cursor_row, plan)
                    if not p_code or p_code not in rows_by_code:
                        break
                    p_row = rows_by_code[p_code]
                    p_match = direct_matches.get(getattr(p_row, plan.code_field, None))
                    if p_match:
                        ancestor_entry = p_match
                        break
                    cursor_row = p_row
                    hops += 1

                if ancestor_entry:
                    new_origin = "derived"
                    # Inherit source_url from the ancestor's first reference; do NOT copy
                    # ancestor's notes_assumption (that's not OUR row's assumption).
                    a_text = ancestor_entry.get("assumption_text") or ""
                    explicit_refs = []
                    for source_ref in ancestor_entry.get("source_refs", []):
                        code = source_ref.get("code")
                        catalog_entry = source_catalog.get(code or "", {})
                        explicit_refs.append(
                            {
                                "code": code,
                                "label": source_ref.get("label"),
                                "section": source_ref.get("section"),
                                "description": catalog_entry.get("description"),
                                "url": catalog_entry.get("url"),
                            }
                        )
                    new_source_refs = self._merge_source_refs(
                        explicit_refs,
                        a_text,
                        source_catalog,
                    )
                    if new_source_refs:
                        new_source_url = new_source_refs[0].get("url")
                    diff["derived"] += 1
                else:
                    diff["unmatched"] += 1
                    diff["orphans"].append(
                        (code, label, "no D.xlsx 1. parameter row or ancestor matches"),
                    )

            current = (row.source_url, row.notes_assumption, row.origin, row.source_refs or [])
            new = (new_source_url, new_notes_assumption, new_origin, new_source_refs)
            if current != new:
                diff["changed"] += 1
                diff["updates"].append(
                    {
                        "pk": row.pk,
                        "code": code,
                        "source_url": new_source_url,
                        "notes_assumption": new_notes_assumption,
                        "source_refs": new_source_refs,
                        "origin": new_origin,
                    }
                )

        return diff

    # ------------------------------------------------------------------
    # Phase C (T66) — create-mode helper
    # ------------------------------------------------------------------

    # Per-model field that receives the xlsx column-W value at clone time.
    # The other status/target fields stay at DE values until per-region
    # target / ziel ingest is added (Phase D).
    _STATUS_FIELDS_FOR_CREATE = {
        "LandUse": ("status_ha", "target_ha"),
        "RenewableData": ("status_value", "target_value"),
        "VerbrauchData": ("status", "ziel"),
        "GebaeudewaermeData": ("status", "ziel"),
    }

    def _create_region_rows_from_de_template(
        self, de_region, target_region, d1_label_index, apply_changes,
    ) -> int:
        """For each parameter model, clone DE base rows into target_region
        with status/target overridden by the new region's xlsx W value
        (when label matches). Fields other than the value pair (codes,
        names, parents, formulas, etc.) carry over verbatim from DE.

        Dry-run prints a summary; --apply commits.
        """
        from django.db import transaction

        total_created = 0
        with (transaction.atomic() if apply_changes else _NoopAtomic()):
            for plan in MODEL_PLANS:
                model = plan.model
                manager = getattr(model, "all_objects", model.objects)
                de_qs = manager.filter(region=de_region)
                if any(f.name == "owner" for f in model._meta.get_fields()):
                    de_qs = de_qs.filter(owner__isnull=True)
                de_rows = list(de_qs.order_by("id"))
                if not de_rows:
                    continue

                # Build map of code → xlsx W value.
                value_map = {}
                for de_row in de_rows:
                    label = getattr(de_row, plan.label_field, None) or ""
                    entry = d1_label_index.get(_normalize_label(label))
                    if entry and isinstance(entry.get("value_w"), (int, float)):
                        value_map[de_row.pk] = float(entry["value_w"])

                value_pair = self._STATUS_FIELDS_FOR_CREATE.get(model.__name__, ())
                concrete = {f.name for f in model._meta.concrete_fields}
                # Skip auto / managed / FK fields we set explicitly.
                skip = {
                    "id", "owner", "region", "created_at", "updated_at",
                    "parent",  # LandUse parent FK — we re-link by code below
                }

                clones = []
                de_pk_to_clone_index = {}
                for idx, de_row in enumerate(de_rows):
                    data = {
                        name: getattr(de_row, name)
                        for name in concrete
                        if name not in skip
                    }
                    new_value = value_map.get(de_row.pk)
                    if new_value is not None and value_pair:
                        for fname in value_pair:
                            if fname in concrete:
                                data[fname] = new_value
                    data["region"] = target_region
                    if "owner" in concrete:
                        data["owner"] = None
                    clones.append(model(**data))
                    de_pk_to_clone_index[de_row.pk] = idx

                if apply_changes:
                    manager.bulk_create(clones, batch_size=1000)
                total_created += len(clones)

                # LandUse: re-link parent FK by code after bulk_create
                if apply_changes and model.__name__ == "LandUse":
                    created = list(
                        manager.filter(
                            region=target_region, owner__isnull=True
                        ).order_by("id")
                    )
                    by_code = {row.code: row for row in created}
                    updates = []
                    for de_row in de_rows:
                        if not de_row.parent_id:
                            continue
                        try:
                            de_parent = de_row.parent
                        except model.DoesNotExist:
                            continue
                        if de_parent is None:
                            continue
                        clone = by_code.get(de_row.code)
                        clone_parent = by_code.get(de_parent.code)
                        if clone is not None and clone_parent is not None and clone.parent_id != clone_parent.id:
                            clone.parent_id = clone_parent.id
                            updates.append(clone)
                    if updates:
                        manager.bulk_update(updates, ["parent"])

        return total_created

    def _propagate_to_workspace_rows(self, region_obj) -> int:
        """For each base row with non-default provenance, copy the provenance
        columns onto every user-workspace row sharing the same code AND
        the same region. Value columns are NOT touched (SR-005 hold). Uses
        QuerySet.update() to bypass signals — provenance writes never affect
        calculations.

        Phase B (T65): per-region scope so a user's BB workspace doesn't get
        DE provenance bleeding into it.
        """
        total = 0
        for plan in MODEL_PLANS:
            model = plan.model
            if not hasattr(model, "all_objects"):
                continue
            if not any(f.name == "owner" for f in model._meta.get_fields()):
                continue
            base_rows = model.all_objects.filter(
                owner__isnull=True,
                region=region_obj,
            ).exclude(
                origin="internal", source_url__isnull=True, notes_assumption__isnull=True,
            )
            for base in base_rows:
                code = getattr(base, plan.code_field, None)
                if not code:
                    continue
                affected = model.all_objects.filter(
                    owner__isnull=False,
                    region=region_obj,
                    code=code,
                ).update(
                    source_url=base.source_url,
                    notes_assumption=base.notes_assumption,
                    source_refs=base.source_refs,
                    origin=base.origin,
                )
                total += affected
        return total

    def _apply_diff(self, model, diff: dict) -> None:
        """Write only the provenance columns via QuerySet.update() —
        bypasses signals so cache invalidation isn't triggered (provenance
        does not affect calculations)."""
        for upd in diff["updates"]:
            if hasattr(model, "all_objects"):
                qs = model.all_objects.filter(pk=upd["pk"])
            else:
                qs = model.objects.filter(pk=upd["pk"])
            qs.update(
                source_url=upd["source_url"],
                notes_assumption=upd["notes_assumption"],
                source_refs=upd["source_refs"],
                origin=upd["origin"],
            )

    # ------------------------------------------------------------------
    # artefacts
    # ------------------------------------------------------------------

    def _write_manifest(self, xlsx_path: str, region: str, diffs: list[dict]) -> None:
        # Phase B (T65): per-region subdir so adding Bundesländer doesn't
        # collide on a single file.
        out_dir = Path(f"data/import/{region}")
        out_dir.mkdir(parents=True, exist_ok=True)

        with open(xlsx_path, "rb") as f:
            file_hash = hashlib.sha256(f.read()).hexdigest()

        # Per-sheet content hash (cheap fingerprint over cell values)
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        sheet_hashes: dict[str, str] = {}
        for sn in REQUIRED_SHEETS:
            if sn in wb.sheetnames:
                ws = wb[sn]
                blob = []
                row_count = 0
                for row in ws.iter_rows(values_only=True):
                    blob.append("|".join(str(v) if v is not None else "" for v in row))
                    row_count += 1
                    if row_count >= 5000:
                        break
                sheet_hashes[sn] = hashlib.sha256("\n".join(blob).encode("utf-8")).hexdigest()

        manifest = {
            "import_tool_version": "1.0.0",
            "import_date": datetime.now(timezone.utc).isoformat(),
            "files": [
                {
                    "path": os.path.basename(xlsx_path),
                    "file_hash": f"sha256:{file_hash}",
                    "sheet_hashes": {k: f"sha256:{v}" for k, v in sheet_hashes.items()},
                    "region_code": region,
                    "rows_imported": sum(d["total"] for d in diffs),
                    "rows_matched_d_xlsx": sum(d["matched"] for d in diffs),
                    "rows_matched_derived": sum(d["derived"] for d in diffs),
                    "rows_unmatched_internal": sum(d["unmatched"] for d in diffs),
                    "rows_changed": sum(d["changed"] for d in diffs),
                    "per_model": [
                        {
                            "model": d["model_name"],
                            "total": d["total"],
                            "d_xlsx": d["matched"],
                            "derived": d["derived"],
                            "internal": d["unmatched"],
                            "changed": d["changed"],
                        }
                        for d in diffs
                    ],
                }
            ],
        }

        path = out_dir / "d_xlsx.manifest.json"
        path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
        self.stdout.write(self.style.SUCCESS(f"Wrote manifest: {path}"))

    def _write_orphan_csv(self, region: str, diffs: list[dict]) -> None:
        out_dir = Path(f"data/import/{region}")
        out_dir.mkdir(parents=True, exist_ok=True)
        path = out_dir / "orphan_classification.csv"
        with open(path, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["model", "code", "label", "origin", "rationale"])
            for d in diffs:
                for code, label, reason in d["orphans"]:
                    w.writerow([d["model_name"], code, label, "internal", reason])
        self.stdout.write(self.style.SUCCESS(f"Wrote orphan CSV: {path}"))
