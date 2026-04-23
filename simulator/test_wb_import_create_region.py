"""
Phase C (T66) — row-creating import mode for new regions.

Phase B's import_excel_provenance only UPDATES existing rows. For a new
Bundesland (BB, etc.) the region table has 0 rows, so the import is a
no-op. Phase C adds a CREATE branch: when the target region is empty,
clone the DE row structure and overlay the new region's values from
its Excel file.

Tests use a tiny synthetic xlsx generated on the fly.
"""
from __future__ import annotations

import glob
import os
import tempfile
from io import StringIO
from pathlib import Path

from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase

from simulator.models import (
    GebaeudewaermeData,
    LandUse,
    Region,
    RenewableData,
    VerbrauchData,
)


_D_XLSX_CANDIDATES = sorted(glob.glob("docs/100prosim_d_*/D.xlsx"))
D_XLSX = _D_XLSX_CANDIDATES[0] if _D_XLSX_CANDIDATES else None


def _need_d(testcase: TestCase) -> None:
    if D_XLSX is None:
        testcase.skipTest("D.xlsx not present in docs/100prosim_d_*/ — skipping")


def _make_synthetic_xlsx_from_de(scale: float = 1.05) -> str:
    """Copy DE's D.xlsx, multiply column W values on sheet '1.' by `scale`,
    save to a temp .xlsx and return the path. Caller is responsible for
    deleting the file.
    """
    from openpyxl import load_workbook

    if D_XLSX is None:
        raise RuntimeError("D.xlsx not present — cannot build synthetic xlsx")

    wb = load_workbook(D_XLSX, data_only=True, read_only=False)
    sheet1 = wb["1."]
    type_col, value_col_w = 67, 23
    for r in range(1, sheet1.max_row + 1):
        t = sheet1.cell(row=r, column=type_col).value
        if t != "p":
            continue
        v = sheet1.cell(row=r, column=value_col_w).value
        if isinstance(v, (int, float)):
            sheet1.cell(row=r, column=value_col_w).value = v * scale

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path


class CreateModeForEmptyRegionTests(TestCase):
    """When --region targets a Region with 0 rows, the import CREATES
    them by cloning DE's structure + overlaying the new region's
    column-W values."""

    def setUp(self):
        _need_d(self)
        # Seed minimal DE rows to clone from.
        self.de = Region.objects.get(code="DE")
        # Use a real D.xlsx label so the scaling math has something to overlay.
        self.de_landuse = LandUse.all_objects.create(
            code="LU_0",
            name="Bodenfläche gesamt",
            status_ha=35759529.0,
            target_ha=35759529.0,
            region=self.de,
        )
        self.test_region = Region.objects.create(
            code="TEST", display_name="Synthetic Test", active=True,
            installed_pmax_ely_gw=200.0, installed_pmax_rv_gw=270.0,
        )

    def test_create_mode_populates_landuse_for_new_region(self):
        synth = _make_synthetic_xlsx_from_de(scale=1.05)
        try:
            call_command(
                "import_excel_provenance",
                synth,
                "--apply",
                "--region=TEST",
                stdout=StringIO(),
                stderr=StringIO(),
            )
            test_rows = LandUse.all_objects.filter(region=self.test_region, owner__isnull=True)
            self.assertGreater(
                test_rows.count(), 0,
                "Create mode should clone DE base rows into the new region.",
            )
            cloned = test_rows.filter(code="LU_0").first()
            self.assertIsNotNone(cloned, "LU_0 should exist in TEST region after create.")
            # Region isolation invariant: TEST row's value MUST differ from
            # DE's (otherwise switching region would be a visual no-op). The
            # xlsx column-W value for "Bodenfläche gesamt" is itself a
            # scenario number rather than an aggregate status, so the +5%
            # multiplier lands at a value clearly distinguishable from DE.
            self.assertIsNotNone(cloned.status_ha)
            self.assertNotAlmostEqual(
                cloned.status_ha, 35759529.0, delta=1.0,
                msg="TEST row must NOT equal DE's value — isolation broken.",
            )
            # Sanity: should look like xlsx_W * 1.05 (positive, scaled).
            self.assertGreater(cloned.status_ha, 0)
        finally:
            Path(synth).unlink(missing_ok=True)

    def test_create_mode_does_not_touch_DE(self):
        synth = _make_synthetic_xlsx_from_de(scale=1.05)
        try:
            call_command(
                "import_excel_provenance",
                synth,
                "--apply",
                "--region=TEST",
                stdout=StringIO(),
                stderr=StringIO(),
            )
            self.de_landuse.refresh_from_db()
            self.assertEqual(
                self.de_landuse.status_ha, 35759529.0,
                "DE rows must NOT be modified by a create-mode TEST import.",
            )
        finally:
            Path(synth).unlink(missing_ok=True)


class UpdateModeStillWorksForExistingDETests(TestCase):
    """The UPDATE path (Phase A/B behaviour) is preserved for DE."""

    def setUp(self):
        _need_d(self)
        # Create a DE row at the same label as something in D.xlsx.
        self.de_row = LandUse.all_objects.create(
            code="LU_0",
            name="Bodenfläche gesamt",
            status_ha=35759529.0,
            target_ha=35759529.0,
        )

    def test_de_value_unchanged_on_provenance_overlay(self):
        before = self.de_row.status_ha
        call_command(
            "import_excel_provenance",
            D_XLSX,
            "--apply",
            "--region=DE",
            stdout=StringIO(),
            stderr=StringIO(),
        )
        self.de_row.refresh_from_db()
        self.assertEqual(self.de_row.status_ha, before, "DE value must not drift")
        self.assertEqual(self.de_row.origin, "d_xlsx")
