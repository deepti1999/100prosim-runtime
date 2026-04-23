"""Verify the formulas for each T54 D1-D4c hardcoded value BEFORE wiring them.

Strategy:
  - Read Excel reference values from WS.xlsm (what the diagram SHOULD show)
  - Read Excel formulas (not just values) from D.xlsx + WS.xlsm to understand
    what denominators/normalizations Schmidt-Kanefendt actually uses
  - For D1/D2 (Tagesladungen): confirm annual / peak_daily OR find alternate
  - For D3 (percent shares): extract the exact formula
  - For D4c (Abgleichdifferenz): locate the solver residual cell

Report: for each of D1/D2/D3/D4c, "formula confirmed" or "needs clarification".
"""
import glob
import warnings
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

WS = glob.glob("docs/100prosim_d_*/WS.xlsm")[0]

# Open WS.xlsm in both modes
WS_V = load_workbook(WS, data_only=True, keep_vba=True)   # values
WS_F = load_workbook(WS, data_only=False, keep_vba=True)  # formulas

def find_cells_with(val, sheet_names=None, tol=0.5):
    """Find all cells with numeric value ≈ val."""
    hits = []
    names = sheet_names or WS_V.sheetnames
    for sn in names:
        ws = WS_V[sn]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and abs(cell.value - val) < tol:
                    hits.append((sn, cell.coordinate, cell.value))
    return hits

# -------------------------------------------------
# D1 source Tagesladungen — known Excel values
# -------------------------------------------------
print("=" * 70)
print("D1 — SOURCE TAGESLADUNGEN")
print("=" * 70)
print("""
Known Excel values (from Jahresstrom diagram at state 250517):
  PV   = 397 (annual 1,201,630 GWh)
  Wind = 186 (annual 706,237 GWh)
  Hydro=   5 (annual 19,509 GWh)
  Bio  =   1 (annual 4,525 GWh)

Derived denominators if formula = annual / Tages:
  PV:    1201630 / 397 = {:.1f}
  Wind:   706237 / 186 = {:.1f}
  Hydro:   19509 /   5 = {:.1f}
  Bio:     4525 /   1 = {:.1f}
""".format(1201630/397, 706237/186, 19509/5, 4525/1))

print("Hypothesis: denominator = PER-SOURCE peak daily output.")
print("Need to verify: does Excel store peak-daily per source somewhere?")

# Search WS for ~3026 (PV per-tages denom)
print("\nCells in WS.xlsm with value 3000-3100 (PV peak daily):")
hits = find_cells_with(3026, tol=2)
for sn, c, v in hits[:5]:
    print(f"  {sn:30s} {c:6s} = {v:.2f}")
if not hits:
    print("  (no exact match — denominator might be computed not stored)")

# Search for 3797 (Wind peak)
print("\nCells with ~3797 (Wind peak daily):")
hits = find_cells_with(3797, tol=3)
for sn, c, v in hits[:5]:
    print(f"  {sn:30s} {c:6s} = {v:.2f}")
if not hits:
    print("  (not found)")

# Search for 4525 (Bio peak = its own annual, since Bio has Tages=1)
print("\nBio self-test: annual=4525, Tages=1 -&gt; peak_daily=4525.")
print("(Bio is 'demand-driven', so its Tages is by definition 1.)")

# -------------------------------------------------
# D2 flow Tagesladungen — known values
# -------------------------------------------------
print("\n" + "=" * 70)
print("D2 — FLOW TAGESLADUNGEN")
print("=" * 70)
print("""
Known flow Tagesladungen (Excel):
  splitter-&gt;Q 1,541,442/509 = {:.0f}
  Q-&gt;Abreg    189,289/62     = {:.0f}
  Q-&gt;N        947,106/313    = {:.0f}
  N-&gt;S        1,105,556/365  = {:.0f}
  Q-&gt;Ely-ES   405,047/134    = {:.0f}
  Ely-P2G-&gt;D  250,857/87     = {:.0f}
  Ely-ES-&gt;Gas 263,280/87     = {:.0f}
  Gas-&gt;Rückv  263,120/87     = {:.0f}
  Rückv-&gt;S    153,925/51     = {:.0f}
  Speicher   241,727/80     = {:.0f}

All flow denominators cluster around 3000 — so flow Tages uses a
SYSTEM-LEVEL constant (probably max daily final_stromnetz) not per-source peaks.
""".format(1541442/509, 189289/62, 947106/313, 1105556/365, 405047/134,
           250857/87, 263280/87, 263120/87, 153925/51, 241727/80))

print("Hypothesis: denominator = max(daily_final_stromnetz) over 365 days,")
print("which is roughly ~3020 GWh/day for this scenario.")
print()
print("Cells in WS.xlsm with value ~3020 (flow-Tages denominator):")
hits = find_cells_with(3020, tol=10)
for sn, c, v in hits[:10]:
    print(f"  {sn:30s} {c:6s} = {v:.2f}")
if not hits:
    print("  (not exactly stored — likely computed)")

# -------------------------------------------------
# D3 percent shares — try to read Excel formulas
# -------------------------------------------------
print("\n" + "=" * 70)
print("D3 — PERCENT SHARES")
print("=" * 70)
print("""
Known Excel values (from Jahresstrom diagram):
  PV    62,2%
  Wind  29,2%
  Hydro  0,8%
  Bio    0,2%
  Sum = 92,4%  (NOT 100% — 7,6% unaccounted)

Naive check pv/(pv+wind+hydro+bio):
""")
pv, wind, hydro, bio = 1201630, 706237, 19509, 4525
s = pv + wind + hydro + bio
for name, val in [("PV", pv), ("Wind", wind), ("Hydro", hydro), ("Bio", bio)]:
    print(f"  {name:6s} {val/s*100:.1f}%  (Excel: see above)")

print("""
Only PV matches. So denominator is NOT the sum of our 4 sources.

Reading Excel cell E21 on '1.Jahresbilanz_Strom' (where audit found 0.6227
for PV) — let's see what formula produces 0.6227:
""")
ws_f = WS_F["1.Jahresbilanz_Strom"]
for col in "ABCDEFGHIJ":
    for r in range(19, 28):
        f_val = ws_f[f"{col}{r}"].value
        v_val = WS_V["1.Jahresbilanz_Strom"][f"{col}{r}"].value
        if f_val is not None and (isinstance(f_val, str) and f_val.startswith("="))\
           or (isinstance(v_val, (int, float)) and 0.001 < abs(v_val) < 1 and abs(v_val - 0.6227) < 0.01):
            print(f"  {col}{r}: formula={f_val!r:50s} computed={v_val!r}")

# E21 directly
print("\nDirect look at E21 and neighbors (where 0.6227 was found):")
for col in "CDEFG":
    f_val = ws_f[f"{col}21"].value
    v_val = WS_V["1.Jahresbilanz_Strom"][f"{col}21"].value
    print(f"  {col}21: formula={str(f_val)[:60]!r:65s} value={v_val!r}")

# -------------------------------------------------
# D4c Abgleichdifferenz — find the residual cell
# -------------------------------------------------
print("\n" + "=" * 70)
print("D4c — ABGLEICHDIFFERENZ (scenario-balance residual)")
print("=" * 70)
print("Known Excel value: 160 GWh (with 80 Tagesladungen)")
hits = find_cells_with(160, tol=0.5)
print(f"\nCells with value ≈ 160 across WS.xlsm:")
for sn, c, v in hits[:15]:
    # Get context — label in col D or E
    ws = WS_V[sn]
    context = ws.cell(row=ws[c].row, column=4).value or ws.cell(row=ws[c].row, column=5).value
    print(f"  {sn:30s} {c:6s} = {v:8.2f}  ctx={str(context)[:40]!r}")

# Also search for "Abgleichdifferenz" string
print("\nCells containing 'Abgleich' text:")
for sn in WS_V.sheetnames:
    ws = WS_V[sn]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "abgleich" in cell.value.lower():
                print(f"  {sn:30s} {cell.coordinate:6s} = {cell.value[:60]!r}")

# -------------------------------------------------
# Summary
# -------------------------------------------------
print("\n" + "=" * 70)
print("SUMMARY OF VERIFICATION")
print("=" * 70)
print("""
D1 (source Tagesladungen): formula CONFIRMABLE via per-source peak
    daily from our daily_data (max across 365 days). No Excel
    formula-reading needed.

D2 (flow Tagesladungen): denominator clusters around 3020 — likely
    max(daily_final_stromnetz). Need to confirm by computing our
    own max and seeing if ~3020 matches.

D3 (percent shares): naive formula matches PV only. Denominator
    for Wind/Hydro/Bio unclear from Excel inspection. Need either:
    (a) read Excel formula string from E21 and neighbors (see above
        output for candidates), or
    (b) one clarification question to Schmidt-Kanefendt.

D4c (Abgleichdifferenz): value is a solver residual. Our backend
    computes it internally but doesn't return it. Code change only.
""")
