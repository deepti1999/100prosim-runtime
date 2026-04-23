"""Round-2 audit: actually find row labels in sheet 1, look for Tagesladungen
in WS.xlsm, and look for percent shares in 8.Kennzahlen / 9.Quellen / I_BS sheets."""
from openpyxl import load_workbook
import warnings, glob
warnings.filterwarnings('ignore')

D = load_workbook(glob.glob('docs/100prosim_d_*/D.xlsx')[0], data_only=True)
WS = load_workbook(glob.glob('docs/100prosim_d_*/WS.xlsm')[0], data_only=True, keep_vba=True)

# ---- Sheet "1." structure: what's in column E (the labels we saw)? ----
print('=== Sheet "1." column E samples (likely the parameter labels) ===')
ws = D['1.']
n = 0
for r in range(1, ws.max_row+1):
    v = ws.cell(row=r, column=5).value
    if v is not None and isinstance(v, str) and len(v.strip()) > 3:
        print(f'  E{r:4d}: {str(v)[:90]!r}')
        n += 1
        if n >= 30: break

# ---- Look at I_S sheet (likely Input-Status, where actual parameter VALUES live) ----
print('\n=== Sheet "I_S" column E samples ===')
ws = D['I_S']
n = 0
for r in range(1, ws.max_row+1):
    v = ws.cell(row=r, column=5).value
    if v is not None:
        # Show row and value if it's a number or label
        next_v = ws.cell(row=r, column=6).value if ws.max_column >= 6 else None
        print(f'  E{r:4d}: {str(v)[:60]!r:65s} F{r}: {str(next_v)[:30]!r}')
        n += 1
        if n >= 25: break

# ---- WS.xlsm: look for Tagesladungen ----
print('\n=== WS.xlsm — find "Tagesladungen" or related cells ===')
for sname in WS.sheetnames:
    ws = WS[sname]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                s = str(cell.value).lower()
                if 'tageslad' in s or 'ø tag' in s or 'ladung' in s or 'tagesladung' in s:
                    print(f'  {sname:30s} {cell.coordinate:6s} = {str(cell.value)[:80]!r}')

# ---- 8.Kennzahlen and 9.Quellen — look for percent share values (62.2 / 29.2 / 0.8 / 0.2) ----
print('\n=== Searching for the % share values 62, 29, 0.8, 0.2 across D.xlsx ===')
TARGETS = [62.2, 29.2, 0.8, 0.2, 0.622, 0.292, 0.008, 0.002]
for sname in D.sheetnames:
    ws = D[sname]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                for t in TARGETS:
                    if abs(cell.value - t) < 0.001:
                        # Get surrounding context
                        label_e = ws.cell(row=cell.row, column=5).value
                        print(f'  {sname:25s} {cell.coordinate:6s} = {cell.value:8.4f}  label_E={str(label_e)[:60]!r}')
                        break

# ---- WS.xlsm: look for the absolute Tagesladungen numbers (397, 186, 5, 1) ----
print('\n=== WS.xlsm — search for known Tagesladungen integers ===')
TAGES = [397, 186, 5, 1, 509, 313, 365, 62, 87, 51, 80, 134]
for sname in WS.sheetnames:
    ws = WS[sname]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value in TAGES:
                # Only report if surrounded by suggestive context
                label_e = ws.cell(row=cell.row, column=5).value
                label_l = ws.cell(row=cell.row, column=12).value
                ctx = str(label_e)[:30] if label_e else (str(label_l)[:30] if label_l else '')
                print(f'  {sname:30s} {cell.coordinate:6s} = {cell.value:6}  ctx={ctx!r}')
