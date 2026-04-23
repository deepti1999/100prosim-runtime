"""Catalog every Excel workbook in docs/100prosim_d_*/.

For each workbook, prints:
  - file size
  - sheet names with dimensions (max_row x max_col)
  - per sheet: sample col-A labels (first 5 non-empty)
  - per sheet: total non-empty cells
  - hyperlink counts (per sheet, totals)
  - cell comment counts (per sheet, totals)
  - external link references (workbook-to-workbook)
  - VBA macro presence (.xlsm only)

Output is structured text designed to be transcribed into
docs/stakeholder/WORKBOOK_CATALOG.md.

Run from repo root:
    python scripts/catalog_workbooks.py > scripts/audit_out/workbook_catalog.txt
"""
from __future__ import annotations

import glob
import os
import warnings
import zipfile
from typing import Iterable

from openpyxl import load_workbook

warnings.filterwarnings("ignore")


def head_labels(ws, col_idx: int = 1, max_take: int = 5) -> list[str]:
    """First N non-empty values of a column, capped at 80 chars each."""
    out: list[str] = []
    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
        if not row:
            continue
        v = row[0]
        if v is None or str(v).strip() == "":
            continue
        s = str(v).strip()
        out.append(s[:80] + "…" if len(s) > 80 else s)
        if len(out) >= max_take:
            break
    return out


def count_non_empty(ws) -> int:
    n = 0
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is not None and str(v).strip() != "":
                n += 1
    return n


def count_links_and_comments(ws) -> tuple[int, int]:
    links = 0
    comments = 0
    for row in ws.iter_rows():
        for cell in row:
            if getattr(cell, "hyperlink", None) is not None:
                links += 1
            if getattr(cell, "comment", None) is not None:
                comments += 1
    return links, comments


def external_links(path: str) -> list[str]:
    out: list[str] = []
    try:
        with zipfile.ZipFile(path) as z:
            for n in z.namelist():
                if "externalLink" in n and n.endswith(".xml") and "rels" not in n:
                    out.append(os.path.basename(n))
                if n == "xl/externalLinks/_rels/" and not out:
                    pass
    except Exception:
        pass
    return out


def vba_present(path: str) -> bool:
    if not path.endswith(".xlsm"):
        return False
    try:
        with zipfile.ZipFile(path) as z:
            return any(n.endswith("vbaProject.bin") for n in z.namelist())
    except Exception:
        return False


def catalog(path: str) -> None:
    base = os.path.basename(path)
    size_mb = os.path.getsize(path) / 1024 / 1024
    print("=" * 78)
    print(f"FILE: {base}    ({size_mb:.2f} MB)    [{path}]")
    print("=" * 78)

    keep_vba = path.endswith(".xlsm")
    try:
        wb = load_workbook(path, data_only=True, keep_vba=keep_vba, read_only=False)
    except Exception as e:
        print(f"  ! Cannot open: {e}")
        return

    print(f"sheet_count: {len(wb.sheetnames)}")
    if keep_vba:
        print(f"vba_macros: {'YES' if vba_present(path) else 'no'}")

    ext = external_links(path)
    if ext:
        print(f"external_link_files: {ext}")

    total_links = 0
    total_comments = 0
    total_cells = 0
    for sname in wb.sheetnames:
        ws = wb[sname]
        try:
            mr = ws.max_row or 0
            mc = ws.max_column or 0
        except Exception:
            mr, mc = 0, 0

        cells = count_non_empty(ws) if mr <= 5000 and mc <= 100 else -1
        if cells == -1:
            cells_display = "(skip — large)"
        else:
            cells_display = f"{cells:>6}"
            total_cells += cells

        try:
            links, comments = count_links_and_comments(ws)
        except Exception:
            links, comments = 0, 0
        total_links += links
        total_comments += comments

        labels = head_labels(ws, col_idx=1)

        print(f"  - {sname:20s}  shape=({mr:>5} x {mc:>3})  cells={cells_display}  links={links:>3}  comments={comments:>4}")
        if labels:
            for lab in labels:
                print(f"        col-A head: {lab}")

    print(f"TOTALS:  cells={total_cells}  hyperlinks={total_links}  comments={total_comments}")
    print()


def main() -> None:
    candidates: list[str] = []
    for ext in ("*.xlsx", "*.xlsm"):
        candidates.extend(sorted(glob.glob(os.path.join("docs", "100prosim_d_*", ext))))
    if not candidates:
        print("No workbooks found under docs/100prosim_d_*/")
        return
    print(f"Found {len(candidates)} workbooks.\n")
    for p in candidates:
        catalog(p)


if __name__ == "__main__":
    main()
