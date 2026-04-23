"""Audit the 100prosim Excel files for traceable source references.

Checks each .xlsx / .xlsm in docs/100prosim_d_*/ for:
  - hyperlinks (cell-level external links — sources/papers/URLs)
  - cell comments (annotations / assumption notes)
  - external link references (linked workbooks)
  - sheet names that hint at sources/notes/assumptions
"""
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import warnings
import glob
import os

warnings.filterwarnings('ignore')

NS = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
      'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}

def audit_file(path):
    print(f'\n=== {os.path.basename(path)} ===')
    try:
        wb = load_workbook(path, data_only=True, keep_vba=path.endswith('.xlsm'))
    except Exception as e:
        print(f'  ! Cannot open: {e}')
        return
    print(f'  Sheets ({len(wb.sheetnames)}): {wb.sheetnames[:12]}')
    if len(wb.sheetnames) > 12:
        print(f'  ... +{len(wb.sheetnames)-12} more')

    # 1. Look for source-related sheet names
    src_sheets = [s for s in wb.sheetnames
                  if any(k in s.lower() for k in ('quell', 'source', 'literatur', 'annahm', 'ref', 'beleg', 'herkunft'))]
    if src_sheets:
        print(f'  Source-named sheets: {src_sheets}')

    # 2. Hyperlink count + comment count per sheet
    total_links = 0; total_comments = 0
    sheets_with_links = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        links = 0; comments = 0
        try:
            # openpyxl exposes hyperlinks via ws._hyperlinks
            for row in ws.iter_rows():
                for cell in row:
                    if cell.hyperlink is not None:
                        links += 1
                    if cell.comment is not None:
                        comments += 1
        except Exception:
            pass
        if links or comments:
            sheets_with_links.append((sname, links, comments))
        total_links += links
        total_comments += comments
    print(f'  TOTAL hyperlinks: {total_links}, TOTAL comments: {total_comments}')

    # 3. External link references (linked workbooks)
    try:
        with zipfile.ZipFile(path) as z:
            ext_links = [n for n in z.namelist() if 'externalLink' in n]
            if ext_links:
                print(f'  External-workbook links: {len(ext_links)}')
                for el in ext_links[:5]:
                    if el.endswith('.xml') and 'rels' not in el:
                        try:
                            data = z.read(el).decode('utf-8', errors='ignore')
                            # find href targets
                            root = ET.fromstring(data)
                            for ref in root.iter():
                                if 'externalBook' in ref.tag or 'sheetNames' in ref.tag:
                                    pass
                        except Exception:
                            pass
    except Exception:
        pass

    if sheets_with_links:
        print(f'  Sheets with links/comments:')
        for sname, lnk, cmt in sheets_with_links[:8]:
            print(f'    {sname:30s}  links={lnk}  comments={cmt}')
        # Sample first hyperlink target from each linked sheet
        for sname, lnk, cmt in sheets_with_links[:3]:
            if lnk == 0:
                continue
            ws = wb[sname]
            samples = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.hyperlink and len(samples) < 3:
                        target = cell.hyperlink.target if cell.hyperlink.target else cell.hyperlink.location
                        samples.append((cell.coordinate, target))
                if len(samples) >= 3: break
            for c, t in samples:
                print(f'    sample link in {sname}: {c} -> {t[:80] if t else None}')

# Audit each main file
for f in sorted(glob.glob('docs/100prosim_d_*/D.xlsx')) + \
         sorted(glob.glob('docs/100prosim_d_*/C.xlsx')) + \
         sorted(glob.glob('docs/100prosim_d_*/MH.xlsx')) + \
         sorted(glob.glob('docs/100prosim_d_*/_S.xlsx')) + \
         sorted(glob.glob('docs/100prosim_d_*/WS.xlsm')) + \
         sorted(glob.glob('docs/100prosim_d_*/_100prosim.xlsm')):
    audit_file(f)
