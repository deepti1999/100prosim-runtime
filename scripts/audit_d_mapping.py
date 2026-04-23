"""Deeper audit of D.xlsx to map our DB tables to Excel sheets, and find
the T54-blocked values (Tagesladungen, percent shares)."""
from openpyxl import load_workbook
import warnings, glob
warnings.filterwarnings('ignore')

path = glob.glob('docs/100prosim_d_*/D.xlsx')[0]
wb = load_workbook(path, data_only=True)

print('=== ALL SHEET NAMES with dimensions ===')
for s in wb.sheetnames:
    ws = wb[s]
    print(f'  {s:30s} {ws.max_row:5d}r × {ws.max_column:4d}c')

# Look for sheets that look like our DB tables
TABLE_HINTS = {
    'LandUse / Flächen': ['LU_', 'Fläche', 'Land', 'Lan', 'Flä'],
    'Renewable / Erneuerbare': ['9.3', 'Wind', 'PV', 'Solar', 'Renew', 'Erneu'],
    'Verbrauch': ['Verbrauch', 'Endenergie', 'Verb'],
    'WS365 / Zeitreihen': ['WS', 'Zeitreihe', 'Jahresgang'],
    'Bilanz': ['Bilanz', 'Stromnetz', 'Jahresbilanz'],
    'Tagesladungen / percent share (T54 D1-D3)': ['Tageslad', '%', 'Anteil', 'Share'],
}

# Search all cells for keywords from our codebase
print('\n=== KEYWORD SEARCH across all sheets ===')
def find_keyword(kw, max_hits=5):
    hits = []
    for s in wb.sheetnames:
        ws = wb[s]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and kw.lower() in str(cell.value).lower():
                    hits.append((s, cell.coordinate, str(cell.value)[:60]))
                    if len(hits) >= max_hits: return hits
    return hits

for kw in ['LU_0', 'LU_2.1', 'LU_6', '9.3.1', '9.3.4', '10.1', '10.2',
           'KLIK', 'Gebäudewärme', 'Prozesswärme', 'Mobile Anwendungen',
           'Tagesladungen', 'Quellbezug', 'Annahme']:
    hits = find_keyword(kw, 3)
    print(f'  {kw!r}:')
    for s, c, v in hits:
        print(f'      {s:25s} {c:6s} = {v!r}')
    if not hits:
        print(f'      (not found)')

# Specifically check sheet '1.' for our parameter codes
print('\n=== Sheet "1." columns (first 30, with header text) ===')
ws = wb['1.']
# Try to find header row — look for row with most string-type cells
header_row_idx = 0
header_score = 0
for r in range(1, min(15, ws.max_row+1)):
    row = [c.value for c in ws[r]]
    score = sum(1 for v in row if isinstance(v, str) and 1 <= len(str(v)) <= 50)
    if score > header_score:
        header_score = score
        header_row_idx = r
print(f'  Likely header row: row {header_row_idx}')
for col in range(1, min(31, ws.max_column+1)):
    cell = ws.cell(row=header_row_idx, column=col)
    print(f'  col {col:3d} ({cell.coordinate}): {str(cell.value)[:60]!r}')
