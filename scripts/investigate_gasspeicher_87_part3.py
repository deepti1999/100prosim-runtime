"""Part 3 — Final triangulation.

Previous findings:
  - Excel "87" at L37/Q37 IS a formula: L36 * TLproEingabeEinheit
  - L36 = 263,970 (formula: =L28*N33, where N33 = 0.65 = Eta Strom→Gas)
  - Q36 = 263,811 (formula: ='Zeitreihen Kalkulation'!U152)
  - TLproEingabeEinheit = 0.0003293634487435486 (=S26/VerbrauchStrom)

Our simulator uses ely_branch_value (9.3.3 Optimales Solar) * ETA_STROM_GAS.
Let's see what ely_branch_value actually is in the seed + what Excel L28 is.
"""
import glob
import warnings
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

WS = glob.glob("docs/100prosim_d_*/WS.xlsm")[0]
WS_V = load_workbook(WS, data_only=True, keep_vba=True)
WS_F = load_workbook(WS, data_only=False, keep_vba=True)

print("=" * 70)
print("Q1 — What is L28 on 1.Jahresbilanz_Strom?")
print("=" * 70)
sheet = "1.Jahresbilanz_Strom"
ws_v = WS_V[sheet]
ws_f = WS_F[sheet]
for r in range(24, 32):
    for col in "ABCDEFGHIJKLMNOPQRS":
        v = ws_v[f"{col}{r}"].value
        f = ws_f[f"{col}{r}"].value
        if v is not None:
            is_f = isinstance(f, str) and f.startswith("=")
            kind = "F" if is_f else "L"
            v_str = f"{v:.2f}" if isinstance(v, (int, float)) else str(v)[:18]
            # Only print non-empty + interesting
            if isinstance(v, str) or isinstance(v, (int, float)) and abs(v) > 0.01:
                print(f"  {col}{r}: {v_str:>20s} [{kind}]  f={str(f)[:55]!r}")

print("\n" + "=" * 70)
print("Q2 — What is VerbrauchStrom (used in TLproEingabeEinheit formula)?")
print("=" * 70)
for name in WS_V.defined_names:
    if "verbrauchstrom" in name.lower():
        dn = WS_V.defined_names[name]
        ref = dn.attr_text
        print(f"  name={name!r}  refers={ref!r}")
        if "!" in ref:
            sheet_part, cell_part = ref.split("!", 1)
            sheet_part = sheet_part.strip("'")
            cell_part = cell_part.replace("$", "")
            try:
                val = WS_V[sheet_part][cell_part].value
                f = WS_F[sheet_part][cell_part].value
                print(f"  -> {sheet_part}!{cell_part} = {val!r}  f={f!r}")
            except Exception as e:
                print(f"  (lookup failed: {e})")

print("\n" + "=" * 70)
print("Q3 — What is S26 on 1.Jahresbilanz_Strom?")
print("=" * 70)
for r in [26]:
    for col in "ABCDEFGHIJKLMNOPQRSTUV":
        v = ws_v[f"{col}{r}"].value
        f = ws_f[f"{col}{r}"].value
        if v is not None:
            is_f = isinstance(f, str) and f.startswith("=")
            v_str = f"{v:.4f}" if isinstance(v, (int, float)) else str(v)[:20]
            print(f"  {col}{r}: {v_str:>20s}  f={str(f)[:55]!r}")

print("\n" + "=" * 70)
print("Q4 — Zeitreihen Kalkulation U152 — what's in this cell?")
print("=" * 70)
sheet = "Zeitreihen Kalkulation"
ws_zv = WS_V[sheet]
ws_zf = WS_F[sheet]
for r in range(150, 155):
    for col in "STUVW":
        v = ws_zv[f"{col}{r}"].value
        f = ws_zf[f"{col}{r}"].value
        if v is not None:
            v_str = f"{v:.2f}" if isinstance(v, (int, float)) else str(v)[:20]
            print(f"  {col}{r}: {v_str:>20s}  f={str(f)[:70]!r}")

print("\n" + "=" * 70)
print("Q5 — What does Excel call L28 (label for the row)?")
print("=" * 70)
# Labels are typically in columns A-F of the same row
sheet = "1.Jahresbilanz_Strom"
ws_v = WS_V[sheet]
for r in [25, 26, 27, 28, 29, 30]:
    for col in "ABCDEFGHIJK":
        v = ws_v[f"{col}{r}"].value
        if isinstance(v, str) and len(v) > 2:
            print(f"  label candidate row {r} col {col}: {v!r}")

print("\n" + "=" * 70)
print("Q6 — Dump rows 20-32 of 1.Jahresbilanz_Strom (surrounding context)")
print("=" * 70)
for r in range(20, 33):
    # Gather all non-empty cells
    row_str = f"  Row {r}: "
    cells = []
    for col in "ABCDEFGHIJKLMNOPQRS":
        v = ws_v[f"{col}{r}"].value
        if v is not None:
            if isinstance(v, (int, float)):
                cells.append(f"{col}={v:.1f}")
            else:
                s = str(v)[:18]
                cells.append(f"{col}={s!r}")
    if cells:
        print(row_str + " | ".join(cells[:10]))
