"""Investigate Gasspeicher Direktverbr 83 vs Excel diagram 87.

Questions:
  1. Where in WS.xlsm does '87' appear next to Gasspeicher Direktverbr?
  2. Does that cell carry a formula or just a literal?
  3. What does our simulator compute (flow_gasspeicher_direkt_tages)?
  4. Where does Excel itself compute the Gasspeicher Direktverbr value (250.857)?

Report: the mathematical truth + a recommendation.
"""
import glob
import warnings
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

WS = glob.glob("docs/100prosim_d_*/WS.xlsm")[0]
WS_V = load_workbook(WS, data_only=True, keep_vba=True)
WS_F = load_workbook(WS, data_only=False, keep_vba=True)

print("=" * 70)
print("Q1 — Find all cells in WS.xlsm with value == 87 (±0.5)")
print("=" * 70)
for sn in WS_V.sheetnames:
    ws_v = WS_V[sn]
    ws_f = WS_F[sn]
    for row in ws_v.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and abs(cell.value - 87) < 0.5:
                f_val = ws_f[cell.coordinate].value
                is_formula = isinstance(f_val, str) and f_val.startswith("=")
                label_col_d = ws_v.cell(row=cell.row, column=4).value
                label_col_e = ws_v.cell(row=cell.row, column=5).value
                label_col_c = ws_v.cell(row=cell.row, column=3).value
                label_col_b = ws_v.cell(row=cell.row, column=2).value
                ctx = label_col_d or label_col_e or label_col_c or label_col_b or "?"
                kind = "FORMULA" if is_formula else "LITERAL"
                print(f"  {sn:32s} {cell.coordinate:6s} = {cell.value:.2f}  [{kind}]  ctx={str(ctx)[:40]!r}")
                if is_formula:
                    print(f"    formula: {f_val}")

print("\n" + "=" * 70)
print("Q2 — Look at cell H37 on every sheet that has an H37")
print("=" * 70)
for sn in WS_V.sheetnames:
    ws_v = WS_V[sn]
    ws_f = WS_F[sn]
    if ws_v.max_row >= 37 and ws_v.max_column >= 8:
        v = ws_v["H37"].value
        f = ws_f["H37"].value
        if v is not None or f is not None:
            is_formula = isinstance(f, str) and f.startswith("=")
            kind = "FORMULA" if is_formula else "LITERAL"
            label = ws_v.cell(row=37, column=2).value or ws_v.cell(row=37, column=3).value or ws_v.cell(row=37, column=4).value
            print(f"  {sn:32s} H37  value={v!r:20s} kind={kind}  row_label={str(label)[:40]!r}")
            if is_formula:
                print(f"    formula: {f}")

print("\n" + "=" * 70)
print("Q3 — Find all cells with 'Gasspeicher Direktverbr' label")
print("=" * 70)
for sn in WS_V.sheetnames:
    ws_v = WS_V[sn]
    for row in ws_v.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "direktverbr" in cell.value.lower():
                print(f"  {sn:32s} {cell.coordinate:6s} = {cell.value!r}")
                # Look at adjacent cells (right 1-10 cols)
                for dc in range(1, 12):
                    adj = ws_v.cell(row=cell.row, column=cell.column + dc)
                    adj_f = WS_F[sn].cell(row=cell.row, column=cell.column + dc).value
                    if adj.value is not None:
                        is_formula = isinstance(adj_f, str) and adj_f.startswith("=")
                        kind = "F" if is_formula else "L"
                        print(f"    +{dc:2d} col: {adj.coordinate:6s} = {str(adj.value)[:30]!r:32s} [{kind}]")

print("\n" + "=" * 70)
print("Q4 — Try to compute 250857 / x for x in {2870, 2882, 2884, 3020, 3022, 3030}")
print("     to see what denominator would give exactly 87.")
print("=" * 70)
annual = 250857
for x in [2870, 2882, 2884, 2900, 3020, 3022, 3026, 3030, 3070]:
    print(f"  250857 / {x} = {annual/x:.3f}")

print("\nFor Tages = 87 exactly, need denominator = 250857 / 87 = {:.2f}".format(annual / 87))
print("For Tages = 83 exactly, need denominator = 250857 / 83 = {:.2f}".format(annual / 83))

print("\n" + "=" * 70)
print("Q5 — What is Excel's 'TLproEingabeEinheit' (our tl_factor)?")
print("     Looking at 1.Jahresbilanz_Strom for the diagram's tl_factor cell")
print("=" * 70)
try:
    sheet = "1.Jahresbilanz_Strom"
    if sheet in WS_V.sheetnames:
        ws = WS_V[sheet]
        ws_f = WS_F[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and (
                    "tlproe" in cell.value.lower() or "tles" in cell.value.lower()
                    or "tages" in cell.value.lower()
                ):
                    print(f"  {cell.coordinate:6s} label={cell.value!r}")
                    for dc in range(1, 8):
                        adj = ws.cell(row=cell.row, column=cell.column + dc)
                        adj_f = ws_f.cell(row=cell.row, column=cell.column + dc).value
                        if adj.value is not None:
                            print(f"    +{dc}: {adj.coordinate} = {adj.value!r}"
                                  f"  f={adj_f!r}")
except Exception as e:
    print(f"(error: {e})")
