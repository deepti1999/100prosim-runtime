"""Probe _S.xlsx VALUES (data_only=True) to understand layout per page-sheet.

For each app-page sheet, dump first ~60 rows showing:
  - col-E label (German parameter description)
  - cols F..AB sample values

Goal: figure out which columns hold the user-facing numeric values
(status / ziel) so the mapping script (Step C) can compare against
our DB rows.

Also: count how many rows have a label AND a numeric value.
"""
from __future__ import annotations

import glob
import os
import warnings

from openpyxl import load_workbook

warnings.filterwarnings("ignore")


def is_number(v) -> bool:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return True
    return False


def short(v, n: int = 22) -> str:
    if v is None:
        return ""
    s = str(v)
    return s if len(s) <= n else s[: n - 1] + "…"


def probe(ws, max_rows: int = 80) -> None:
    print(f"\n--- _S sheet: {ws.title}  ({ws.max_row} x {ws.max_column}) ---")
    # Try cols 1..30
    headers_at_top = []
    for col in range(1, min(ws.max_column, 30) + 1):
        v = ws.cell(row=1, column=col).value
        headers_at_top.append((col, v))

    # Print header line
    print("  HEAD:  ", " | ".join(f"{c}={short(v, 12)}" for c, v in headers_at_top[:20] if v is not None))

    for row in range(2, min(ws.max_row, max_rows) + 1):
        cells = [ws.cell(row=row, column=c).value for c in range(1, min(ws.max_column, 30) + 1)]
        # find label column (first text-bearing col 4..10)
        label_col = None
        label = None
        for c in range(4, 12):
            v = cells[c - 1]
            if isinstance(v, str) and v.strip() and not v.startswith("=") and not v[0].isdigit():
                label_col = c
                label = v.strip()
                break
        if label_col is None:
            continue
        # gather numeric cols among 4..30
        numeric_pairs = [
            (c, cells[c - 1])
            for c in range(4, min(len(cells), 30) + 1)
            if is_number(cells[c - 1])
        ][:6]
        if not numeric_pairs:
            continue
        nums_str = " ".join(f"col{c}={n:>10g}" for c, n in numeric_pairs)
        print(f"  row {row:3d}  L({label_col}): {short(label, 35):37s}  {nums_str}")


def main() -> None:
    paths = glob.glob("docs/100prosim_d_*/_S.xlsx")
    if not paths:
        print("not found")
        return
    p = paths[0]
    print(f"Loading {p} (data_only=True)...")
    wb = load_workbook(p, data_only=True)
    for s in [
        "1. Flächen",
        "2. Erneuerbare",
        "3. Bedarfsniveau",
        "4. Verbrauch",
        "5. Bilanz",
        "6. Fossile",
        "7. Verbrauch Status",
        "8. Kennzahlen",
    ]:
        if s in wb.sheetnames:
            probe(wb[s])
        else:
            print(f"!!! {s!r} not in {wb.sheetnames}")


if __name__ == "__main__":
    main()
