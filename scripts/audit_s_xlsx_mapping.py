"""Map our 420 DB rows to _S.xlsx (sheet, row).

Why _S.xlsx is the right substrate (vs D.xlsx alone):
  - _S.xlsx sheets are named exactly like our app pages:
        '1. Flächen', '2. Erneuerbare', '3. Bedarfsniveau', '4. Verbrauch',
        '5. Bilanz', '6. Fossile', '7. Verbrauch Status', '8. Kennzahlen'
  - Each sheet uses col E as the German parameter label, with status / ziel
    in dedicated columns. Convention varies per sheet (see PER_MODEL below).
  - _S.xlsx cells are mostly formulas referring to D.xlsx via external link
    [4]. Loading with data_only=False exposes those refs, giving us the
    automatic _S → D provenance chain.

Per-model mapping plan:
   Model              -> _S sheet            status_col  ziel_col   D-link prefix
   LandUse            -> "1. Flächen"        I (9)       L (12)     [4]1.
   RenewableData      -> "2. Erneuerbare"    L (12)      M (13)     [4]1.
   VerbrauchData      -> "4. Verbrauch"      L (12)      M (13)     [4]1.
   GebaeudewaermeData -> "3. Bedarfsniveau"  L (12)      M (13)     [4]1.
   (fall-back: try other page sheets if no match in primary)

For each DB row:
  1. Find candidate _S rows where col-E label closely matches our label.
  2. Tie-break with numeric closeness on status_col / ziel_col (with
     a small list of possible scale factors {1, 0.001, 0.0001, 1000, 10000}
     to handle GWh/TWh & ha drift).
  3. Once a row is picked, read the formula in status_col and ziel_col
     from a parallel data_only=False workbook to extract any external
     reference token like "'[4]1.'!U54" — that's D.xlsx provenance.
  4. Emit CSV row.

Output CSVs:
  scripts/audit_out/s_xlsx_map_landuse.csv
  scripts/audit_out/s_xlsx_map_renewabledata.csv
  scripts/audit_out/s_xlsx_map_verbrauchdata.csv
  scripts/audit_out/s_xlsx_map_gebaeudewaermedata.csv

Plus a summary JSON: scripts/audit_out/s_xlsx_map_summary.json

Run from repo root:
    python scripts/audit_s_xlsx_mapping.py
"""
from __future__ import annotations

import csv
import glob
import json
import os
import re
import warnings
from dataclasses import dataclass, field
from typing import Iterable, Optional

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

OUT_DIR = "scripts/audit_out"
os.makedirs(OUT_DIR, exist_ok=True)


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------


@dataclass
class ModelPlan:
    model: str
    seed_key: str  # key in the seed JSON
    primary_sheet: str
    fallback_sheets: list[str]
    status_col: int
    ziel_col: int
    label_field: str
    code_field: str
    status_field: str
    ziel_field: str

    def per_row_status_value(self, row: dict) -> Optional[float]:
        v = row["fields"].get(self.status_field)
        return float(v) if isinstance(v, (int, float)) else None

    def per_row_ziel_value(self, row: dict) -> Optional[float]:
        v = row["fields"].get(self.ziel_field)
        return float(v) if isinstance(v, (int, float)) else None

    def per_row_label(self, row: dict) -> str:
        v = row["fields"].get(self.label_field) or ""
        return str(v).strip()

    def per_row_code(self, row: dict) -> str:
        return str(row["fields"].get(self.code_field) or "")


PLANS: list[ModelPlan] = [
    ModelPlan(
        model="LandUse",
        seed_key="simulator.landuse",
        primary_sheet="1. Flächen",
        fallback_sheets=[],
        status_col=9,
        ziel_col=12,
        label_field="name",
        code_field="code",
        status_field="status_ha",
        ziel_field="target_ha",
    ),
    ModelPlan(
        model="RenewableData",
        seed_key="simulator.renewabledata",
        primary_sheet="2. Erneuerbare",
        fallback_sheets=["1. Flächen"],
        status_col=12,
        ziel_col=13,
        label_field="name",
        code_field="code",
        status_field="status_value",
        ziel_field="target_value",
    ),
    ModelPlan(
        model="VerbrauchData",
        seed_key="simulator.verbrauchdata",
        primary_sheet="4. Verbrauch",
        fallback_sheets=["3. Bedarfsniveau", "7. Verbrauch Status"],
        status_col=12,
        ziel_col=13,
        label_field="category",
        code_field="code",
        status_field="status",
        ziel_field="ziel",
    ),
    ModelPlan(
        model="GebaeudewaermeData",
        seed_key="simulator.gebaeudewaermedata",
        primary_sheet="3. Bedarfsniveau",
        fallback_sheets=["4. Verbrauch"],
        status_col=12,
        ziel_col=13,
        label_field="category",
        code_field="code",
        status_field="status",
        ziel_field="ziel",
    ),
]


# Possible scale factors between our DB value and the _S cell value.
# A negative power means we'd multiply our DB value by that factor to
# match _S; a positive means divide. We'll just check both directions.
SCALES = [1.0, 0.001, 0.0001, 1000.0, 10000.0]


EXT_REF_RE = re.compile(r"\[(\d+)\]([^!\s'\"]+)['\"]?!\$?([A-Z]+)\$?(\d+)")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def norm_label(s: str) -> str:
    if not s:
        return ""
    s = s.lower()
    s = re.sub(r"[^\wäöüß]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def label_overlap(a: str, b: str) -> float:
    na, nb = set(norm_label(a).split()), set(norm_label(b).split())
    if not na or not nb:
        return 0.0
    inter = na & nb
    return len(inter) / max(len(na), len(nb))


def value_close(a: Optional[float], b: Optional[float], rel: float = 0.005, abs_tol: float = 0.01) -> bool:
    if a is None or b is None:
        return False
    if abs(a) < 1e-9 and abs(b) < 1e-9:
        return True
    if abs(a - b) < abs_tol:
        return True
    denom = max(abs(a), abs(b))
    if denom == 0:
        return False
    return abs(a - b) / denom < rel


def value_close_with_scale(ours: Optional[float], theirs: Optional[float]) -> tuple[bool, float]:
    """Return (matched, scale_used). Tries SCALES + their inverses."""
    if ours is None or theirs is None:
        return False, 1.0
    if value_close(ours, theirs):
        return True, 1.0
    for s in SCALES:
        if s == 1.0:
            continue
        if value_close(ours * s, theirs):
            return True, s
        if value_close(ours, theirs * s):
            return True, 1.0 / s
    return False, 1.0


def extract_external_ref(formula: object) -> Optional[str]:
    if not isinstance(formula, str):
        return None
    if not formula.startswith("="):
        return None
    m = EXT_REF_RE.search(formula)
    if not m:
        return None
    book_id = m.group(1)
    sheet = m.group(2)
    col = m.group(3)
    row = m.group(4)
    return f"[{book_id}]{sheet}!{col}{row}"


# ---------------------------------------------------------------------------
# Load _S.xlsx — both views
# ---------------------------------------------------------------------------


def load_s_xlsx() -> tuple[dict, dict]:
    p = glob.glob("docs/100prosim_d_*/_S.xlsx")[0]
    print(f"Loading {p} (data_only=True)...")
    wb_v = load_workbook(p, data_only=True)
    print(f"Loading {p} (data_only=False)...")
    wb_f = load_workbook(p, data_only=False)
    return wb_v, wb_f


def index_sheet(ws_v, status_col: int, ziel_col: int) -> list[dict]:
    """Return per-row dicts for the sheet, filtering rows that have a label."""
    rows = []
    max_r = ws_v.max_row
    max_c = ws_v.max_column
    for r in range(1, max_r + 1):
        # try label in cols 4..10
        label = None
        label_col = None
        for c in range(4, min(11, max_c) + 1):
            v = ws_v.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() and not v.strip().startswith("=") and not v.strip()[0].isdigit():
                label = v.strip()
                label_col = c
                break
        if label is None:
            continue
        s_v = ws_v.cell(row=r, column=status_col).value if status_col <= max_c else None
        z_v = ws_v.cell(row=r, column=ziel_col).value if ziel_col <= max_c else None
        rows.append(
            {
                "row": r,
                "label": label,
                "label_col": label_col,
                "status_value": s_v if isinstance(s_v, (int, float)) else None,
                "ziel_value": z_v if isinstance(z_v, (int, float)) else None,
            }
        )
    return rows


# ---------------------------------------------------------------------------
# Mapping core
# ---------------------------------------------------------------------------


def find_match(plan: ModelPlan, db_row: dict, sheet_index: list[dict]) -> Optional[dict]:
    label = plan.per_row_label(db_row)
    s_ours = plan.per_row_status_value(db_row)
    z_ours = plan.per_row_ziel_value(db_row)

    candidates: list[tuple[float, dict, dict]] = []
    for sr in sheet_index:
        ov = label_overlap(label, sr["label"])
        s_match, s_scale = value_close_with_scale(s_ours, sr["status_value"])
        z_match, z_scale = value_close_with_scale(z_ours, sr["ziel_value"])
        score = 0.0
        if ov >= 0.6:
            score += 1.0
        elif ov >= 0.3:
            score += 0.5
        if s_match:
            score += 1.0
        if z_match:
            score += 1.0
        if score > 0:
            candidates.append(
                (
                    score,
                    sr,
                    {
                        "label_overlap": round(ov, 2),
                        "status_match": s_match,
                        "status_scale": s_scale,
                        "ziel_match": z_match,
                        "ziel_scale": z_scale,
                    },
                )
            )
    if not candidates:
        return None
    candidates.sort(key=lambda t: -t[0])
    score, sr, meta = candidates[0]
    return {"score": score, "sheet_row": sr, "match_meta": meta}


def confidence_label(score: float, has_status_match: bool, has_ziel_match: bool, has_label_match: bool) -> str:
    if has_status_match and has_ziel_match:
        return "HIGH"
    if has_status_match or has_ziel_match:
        return "MED"
    if has_label_match:
        return "LABEL_ONLY"
    return "NONE"


def map_model(plan: ModelPlan, seed: list[dict], wb_v, wb_f) -> tuple[list[dict], dict]:
    seed_rows = [r for r in seed if r["model"] == plan.seed_key]
    print(f"\n=== {plan.model}: {len(seed_rows)} rows ===")

    # Build sheet index for primary + fallback
    if plan.primary_sheet not in wb_v.sheetnames:
        print(f"  ! primary sheet {plan.primary_sheet!r} not in {wb_v.sheetnames}")
        return [], {"model": plan.model, "rows": len(seed_rows), "high": 0, "med": 0, "label_only": 0, "none": 0}

    sheets_to_try: list[tuple[str, list[dict]]] = []
    sheets_to_try.append((plan.primary_sheet, index_sheet(wb_v[plan.primary_sheet], plan.status_col, plan.ziel_col)))
    for fb in plan.fallback_sheets:
        if fb in wb_v.sheetnames:
            sheets_to_try.append((fb, index_sheet(wb_v[fb], plan.status_col, plan.ziel_col)))

    rows_out = []
    counts = {"HIGH": 0, "MED": 0, "LABEL_ONLY": 0, "NONE": 0}

    for db in seed_rows:
        code = plan.per_row_code(db)
        label = plan.per_row_label(db)
        s_ours = plan.per_row_status_value(db)
        z_ours = plan.per_row_ziel_value(db)

        best = None
        best_sheet = None
        for sheet_name, idx in sheets_to_try:
            r = find_match(plan, db, idx)
            if r is None:
                continue
            if best is None or r["score"] > best["score"]:
                best = r
                best_sheet = sheet_name

        if best is None:
            rows_out.append(
                {
                    "our_code": code,
                    "our_label": label,
                    "our_status": s_ours,
                    "our_ziel": z_ours,
                    "s_sheet": "",
                    "s_row": "",
                    "s_label": "",
                    "s_status_value": "",
                    "s_ziel_value": "",
                    "s_status_formula": "",
                    "s_ziel_formula": "",
                    "d_status_ref": "",
                    "d_ziel_ref": "",
                    "label_overlap": 0.0,
                    "status_match": False,
                    "status_scale": 1.0,
                    "ziel_match": False,
                    "ziel_scale": 1.0,
                    "confidence": "NONE",
                }
            )
            counts["NONE"] += 1
            continue

        sr = best["sheet_row"]
        meta = best["match_meta"]
        ws_f = wb_f[best_sheet]
        sf = ws_f.cell(row=sr["row"], column=plan.status_col).value
        zf = ws_f.cell(row=sr["row"], column=plan.ziel_col).value
        sf_str = sf if isinstance(sf, str) else ""
        zf_str = zf if isinstance(zf, str) else ""
        d_sref = extract_external_ref(sf) or ""
        d_zref = extract_external_ref(zf) or ""

        conf = confidence_label(
            best["score"],
            meta["status_match"],
            meta["ziel_match"],
            meta["label_overlap"] >= 0.6,
        )
        counts[conf] += 1

        rows_out.append(
            {
                "our_code": code,
                "our_label": label,
                "our_status": s_ours,
                "our_ziel": z_ours,
                "s_sheet": best_sheet,
                "s_row": sr["row"],
                "s_label": sr["label"],
                "s_status_value": sr["status_value"],
                "s_ziel_value": sr["ziel_value"],
                "s_status_formula": sf_str[:150],
                "s_ziel_formula": zf_str[:150],
                "d_status_ref": d_sref,
                "d_ziel_ref": d_zref,
                "label_overlap": meta["label_overlap"],
                "status_match": meta["status_match"],
                "status_scale": meta["status_scale"],
                "ziel_match": meta["ziel_match"],
                "ziel_scale": meta["ziel_scale"],
                "confidence": conf,
            }
        )

    print(
        f"  matched HIGH={counts['HIGH']:4d}  MED={counts['MED']:4d}  "
        f"LABEL_ONLY={counts['LABEL_ONLY']:4d}  NONE={counts['NONE']:4d}"
    )
    summary = {
        "model": plan.model,
        "rows": len(seed_rows),
        "high": counts["HIGH"],
        "med": counts["MED"],
        "label_only": counts["LABEL_ONLY"],
        "none": counts["NONE"],
        "primary_sheet": plan.primary_sheet,
        "fallback_sheets": plan.fallback_sheets,
        "status_col": plan.status_col,
        "ziel_col": plan.ziel_col,
    }
    return rows_out, summary


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------


CSV_FIELDS = [
    "our_code",
    "our_label",
    "our_status",
    "our_ziel",
    "s_sheet",
    "s_row",
    "s_label",
    "s_status_value",
    "s_ziel_value",
    "s_status_formula",
    "s_ziel_formula",
    "d_status_ref",
    "d_ziel_ref",
    "label_overlap",
    "status_match",
    "status_scale",
    "ziel_match",
    "ziel_scale",
    "confidence",
]


def emit_csv(plan: ModelPlan, rows: list[dict]) -> str:
    fname = os.path.join(OUT_DIR, f"s_xlsx_map_{plan.model.lower()}.csv")
    with open(fname, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    return fname


def main() -> None:
    with open("seed/sqlite_seed.json", encoding="utf-8") as f:
        seed = json.load(f)

    wb_v, wb_f = load_s_xlsx()
    summaries = []
    for plan in PLANS:
        rows, summary = map_model(plan, seed, wb_v, wb_f)
        path = emit_csv(plan, rows)
        summary["csv"] = path
        summaries.append(summary)
        print(f"  wrote {path}  ({len(rows)} rows)")

    # Aggregate summary
    total = sum(s["rows"] for s in summaries)
    high = sum(s["high"] for s in summaries)
    med = sum(s["med"] for s in summaries)
    lonly = sum(s["label_only"] for s in summaries)
    none = sum(s["none"] for s in summaries)
    agg = {
        "total_rows": total,
        "high": high,
        "med": med,
        "label_only": lonly,
        "none": none,
        "models": summaries,
        "scale_set_used": SCALES,
        "_S_external_link_map": {
            "[1]": "C.xlsx",
            "[2]": "WS.xlsm",
            "[3]": "BS.xlsx (NOT in bundle)",
            "[4]": "D.xlsx",
        },
    }
    out_json = os.path.join(OUT_DIR, "s_xlsx_map_summary.json")
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(agg, f, indent=2, default=str)
    print(f"\n--- AGGREGATE ---")
    print(f"  rows total:    {total}")
    print(f"  HIGH (s+z):    {high} ({100*high/total:.1f}%)")
    print(f"  MED (s|z):     {med} ({100*med/total:.1f}%)")
    print(f"  LABEL_ONLY:    {lonly} ({100*lonly/total:.1f}%)")
    print(f"  NONE:          {none} ({100*none/total:.1f}%)")
    print(f"  summary -> {out_json}")


if __name__ == "__main__":
    main()
