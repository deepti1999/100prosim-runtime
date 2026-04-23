"""Final mapping audit — loose tolerance (5%), scale-aware, label-tiebreak.

For each DB parameter row:
 1. Find all D.xlsx cells (row, col) where value * scale ≈ our status or ziel
    for scale ∈ {1, 1000, 0.001, 10000, 0.0001}.
 2. Filter to cells whose label shares ≥ 1 token with our label.
 3. Score (value-match + label-overlap); pick top 1.
 4. Emit mapping CSV with scale factor recorded.
"""
import json, glob, warnings, csv, os, re
from openpyxl import load_workbook
from collections import Counter
warnings.filterwarnings('ignore')

with open('seed/sqlite_seed.json', encoding='utf-8') as f:
    seed = json.load(f)

D = load_workbook(glob.glob('docs/100prosim_d_*/D.xlsx')[0], data_only=True)
ws = D['1.']
SCEN_COLS = [('U', 21), ('V', 22), ('W', 23), ('AG', 33), ('AN', 40)]
SCALES = [1, 1000, 0.001, 10000, 0.0001]

excel_rows = []
for r in range(1, ws.max_row + 1):
    label_e = ws.cell(row=r, column=5).value
    if not isinstance(label_e, str) or len(label_e.strip()) < 3:
        continue
    vals = {}
    for name, c in SCEN_COLS:
        v = ws.cell(row=r, column=c).value
        if isinstance(v, (int, float)):
            vals[name] = float(v)
    if vals:
        excel_rows.append({'row': r, 'label': label_e.strip()[:200], 'vals': vals})


def norm(s):
    if not s: return ''
    s = s.lower()
    s = re.sub(r'[^\wäöüß]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def close(a, b, rel=0.05):
    if a == 0 and b == 0: return True
    denom = max(abs(a), abs(b))
    return denom > 0 and abs(a - b) / denom < rel


def match_field(target, db_words, field_name):
    """For one DB field value (status or ziel), return best (excel_row, col, scale, overlap)."""
    if target is None or not isinstance(target, (int, float)) or target == 0:
        return None
    best = None
    best_score = -1
    for scale in SCALES:
        scaled = target * scale
        for er in excel_rows:
            for col_name, v in er['vals'].items():
                if not close(scaled, v):
                    continue
                ex_words = set(norm(er['label']).split())
                if not ex_words: continue
                overlap = len(db_words & ex_words)
                if overlap == 0: continue
                # Score: label overlap ratio + prefer scale=1 + prefer col W
                score = overlap / max(len(db_words) or 1, len(ex_words) or 1)
                if scale == 1: score += 0.1
                if col_name == 'W': score += 0.05
                if score > best_score:
                    best_score = score
                    best = (er, col_name, scale, overlap)
    return best


MODELS = {
    'simulator.landuse': ('name', 'status_ha', 'target_ha'),
    'simulator.renewabledata': (None, 'status_value', 'target_value'),
    'simulator.verbrauchdata': ('category', 'status', 'ziel'),
    'simulator.gebaeudewaermedata': ('category', 'status', 'ziel'),
}

os.makedirs('scripts/audit_out', exist_ok=True)
summary = {}

for model_name, (label_f, status_f, ziel_f) in MODELS.items():
    rows = [r for r in seed if r['model'] == model_name]
    short = model_name.split('.')[-1]
    csv_path = f'scripts/audit_out/final_map_{short}.csv'

    status_cols = Counter()
    ziel_cols = Counter()
    status_scales = Counter()
    ziel_scales = Counter()
    matched = 0
    status_only = 0
    ziel_only = 0
    both = 0
    zero_rows = 0
    unmatched = 0

    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['our_code', 'our_label', 'status_ours', 'ziel_ours',
                    'status_excel_row', 'status_col', 'status_scale', 'status_excel_label',
                    'ziel_excel_row', 'ziel_col', 'ziel_scale', 'ziel_excel_label',
                    'confidence'])
        for r in rows:
            fv = r['fields']
            code = fv.get('code', '')
            if short == 'renewabledata':
                label = (fv.get('subcategory') or '') + ' / ' + (fv.get('category') or '')
            else:
                label = fv.get(label_f or 'category', '') or ''
            s_ours = fv.get(status_f)
            z_ours = fv.get(ziel_f)
            db_words = set(norm(label).split())

            if (s_ours in (None, 0) and z_ours in (None, 0)):
                zero_rows += 1
                w.writerow([code, label[:60], s_ours, z_ours,
                            '', '', '', '', '', '', '', '', 'ZERO'])
                continue

            s_match = match_field(s_ours, db_words, 'status')
            z_match = match_field(z_ours, db_words, 'ziel')

            conf = 'LOW'
            if s_match and z_match:
                both += 1; conf = 'HIGH'
            elif s_match:
                status_only += 1; conf = 'MED'
            elif z_match:
                ziel_only += 1; conf = 'MED'
            else:
                unmatched += 1

            s_row = s_match[0]['row'] if s_match else ''
            s_col = s_match[1] if s_match else ''
            s_scale = s_match[2] if s_match else ''
            s_lbl = s_match[0]['label'][:50] if s_match else ''
            z_row = z_match[0]['row'] if z_match else ''
            z_col = z_match[1] if z_match else ''
            z_scale = z_match[2] if z_match else ''
            z_lbl = z_match[0]['label'][:50] if z_match else ''

            if s_match:
                status_cols[s_col] += 1
                status_scales[s_scale] += 1
                matched += 1
            if z_match:
                ziel_cols[z_col] += 1
                ziel_scales[z_scale] += 1

            w.writerow([code, label[:60], s_ours, z_ours,
                        s_row, s_col, s_scale, s_lbl,
                        z_row, z_col, z_scale, z_lbl,
                        conf])

    total = len(rows)
    summary[short] = {
        'total': total, 'both': both, 'status_only': status_only, 'ziel_only': ziel_only,
        'zero': zero_rows, 'unmatched': unmatched,
        'status_cols': dict(status_cols), 'ziel_cols': dict(ziel_cols),
        'status_scales': dict(status_scales), 'ziel_scales': dict(ziel_scales),
    }
    print(f'=== {short} (n={total}) ===')
    print(f'  HIGH (both match): {both}')
    print(f'  MED (status or ziel match): {status_only + ziel_only}')
    print(f'  LOW (no match): {unmatched}')
    print(f'  ZERO (no values): {zero_rows}')
    print(f'  status column source: {dict(status_cols)}')
    print(f'  ziel column source:   {dict(ziel_cols)}')
    print(f'  status scale factor:  {dict(status_scales)}')
    print(f'  ziel scale factor:    {dict(ziel_scales)}')
    print(f'  -> {csv_path}')
    print()

print('===== OVERALL =====')
tot_rows = sum(s['total'] for s in summary.values())
tot_high = sum(s['both'] for s in summary.values())
tot_med = sum(s['status_only'] + s['ziel_only'] for s in summary.values())
tot_zero = sum(s['zero'] for s in summary.values())
tot_low = sum(s['unmatched'] for s in summary.values())
print(f'  {tot_rows} rows total')
print(f'  HIGH: {tot_high} ({100*tot_high/tot_rows:.0f}%)')
print(f'  MED:  {tot_med} ({100*tot_med/tot_rows:.0f}%)')
print(f'  ZERO: {tot_zero} ({100*tot_zero/tot_rows:.0f}%)')
print(f'  LOW:  {tot_low} ({100*tot_low/tot_rows:.0f}%)')

import json as j
with open('scripts/audit_out/final_summary.json', 'w') as f:
    j.dump(summary, f, indent=2)
