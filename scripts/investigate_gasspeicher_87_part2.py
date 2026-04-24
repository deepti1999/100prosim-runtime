"""Follow-up — find TLproEingabeEinheit's value and L36/Q36 annuals.

Key findings from part 1:
  - Cell L37 on '1.Jahresbilanz_Strom' = 86.94, formula =L36*TLproEingabeEinheit
  - Cell Q37 on '1.Jahresbilanz_Strom' = 86.89, formula =Q36*TLproEingabeEinheit
  - There is NO H37 — that was wrong in HARDCODED_VALUES_TRACE §6

So "87" IS Excel-computed. We need to find:
  1. What is TLproEingabeEinheit's VALUE (named range)?
  2. What is L36 and Q36 (annual values feeding into 87)?
  3. What row-labels do L36/Q36 represent (Gasspeicher Direktverbr?)
  4. Compare: Excel's formula output (87) vs ours (83) — which is math-correct?
"""
import glob
import warnings
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

WS = glob.glob("docs/100prosim_d_*/WS.xlsm")[0]
WS_V = load_workbook(WS, data_only=True, keep_vba=True)
WS_F = load_workbook(WS, data_only=False, keep_vba=True)

print("=" * 70)
print("Q1 — Value of TLproEingabeEinheit (named range)")
print("=" * 70)
# Check defined names
for name in WS_V.defined_names:
    try:
        dn = WS_V.defined_names[name]
        if "tlpro" in name.lower() or "eingabe" in name.lower():
            print(f"  defined_name: {name!r}")
            print(f"    refers to: {dn.attr_text}")
    except Exception as e:
        print(f"  (error reading {name}: {e})")

# Sometimes the named range is defined in a single cell — just search for the value
print("\n  Searching all sheets for any cell with label 'TLproEingabeEinheit':")
for sn in WS_V.sheetnames:
    ws = WS_V[sn]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "tlproe" in cell.value.lower():
                print(f"    {sn:30s} {cell.coordinate:6s} = {cell.value!r}")
                for dc in range(1, 6):
                    adj = ws.cell(row=cell.row, column=cell.column + dc)
                    adj_f = WS_F[sn].cell(row=cell.row, column=cell.column + dc).value
                    if adj.value is not None:
                        print(f"      +{dc}: {adj.coordinate} = {adj.value!r}  f={adj_f!r}")

print("\n" + "=" * 70)
print("Q2 — Inspect 1.Jahresbilanz_Strom rows 32-42 columns I-R")
print("=" * 70)
sheet = "1.Jahresbilanz_Strom"
ws_v = WS_V[sheet]
ws_f = WS_F[sheet]
for r in range(32, 43):
    row_cells = []
    label = None
    # Find label col — usually A-E
    for col in "ABCDEF":
        cv = ws_v[f"{col}{r}"].value
        if isinstance(cv, str) and len(cv) > 3:
            label = cv
            break
    print(f"\n  Row {r} label: {label!r}")
    for col in "IJKLMNOPQR":
        v = ws_v[f"{col}{r}"].value
        f = ws_f[f"{col}{r}"].value
        if v is not None:
            is_f = isinstance(f, str) and f.startswith("=")
            kind = "F" if is_f else "L"
            v_str = f"{v:.2f}" if isinstance(v, (int, float)) else str(v)[:15]
            print(f"    {col}{r}: {v_str:>10s} [{kind}]  formula={str(f)[:60]!r}")

print("\n" + "=" * 70)
print("Q3 — Specifically find TLproEingabeEinheit cell")
print("=" * 70)
# Workbook-level defined names
for name in WS_V.defined_names:
    if "TLpro" in name or "TLPro" in name or "tlpro" in name.lower():
        dn = WS_V.defined_names[name]
        print(f"  name={name!r}")
        print(f"  refers: {dn.attr_text}")
        # Parse "'Sheet Name'!$X$Y"
        ref = dn.attr_text
        if "!" in ref:
            sheet_part, cell_part = ref.split("!", 1)
            sheet_part = sheet_part.strip("'")
            cell_part = cell_part.replace("$", "")
            try:
                val = WS_V[sheet_part][cell_part].value
                f = WS_F[sheet_part][cell_part].value
                print(f"  -> {sheet_part}!{cell_part} = {val!r}  formula={f!r}")
            except Exception as e:
                print(f"  (lookup failed: {e})")

print("\n" + "=" * 70)
print("Q4 — Our simulator's tl_factor vs Excel's TLproEingabeEinheit")
print("=" * 70)
print("""
Our formula: tl_factor = 365 / final_stromnetz
            where final_stromnetz ~ 1,105,519 GWh
            so tl_factor ~ 365/1105519 = 3.302e-4
            1/tl_factor ~ 3022 (days equivalent of final_stromnetz)

For Tages = 87 on annual 250,857: TLproEingabeEinheit = 87/250857 = 3.468e-4
                                  1/TLproEingabeEinheit = 2884

Excel's TLproEingabeEinheit is 3.468e-4, ours is 3.302e-4.
Difference: Excel uses denominator ~2884, we use ~3022.

Ratio: 3022/2884 = {:.4f}
      Our tl_factor is ~4.8% smaller than Excel's.

So Excel uses a SMALLER denominator -> LARGER Tagesladung value.
""".format(3022/2884))
