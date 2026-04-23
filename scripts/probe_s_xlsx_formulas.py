"""Probe _S.xlsx formulas to understand its link to D.xlsx.

For each app-page sheet (1. Flächen, 2. Erneuerbare, 3. Bedarfsniveau,
4. Verbrauch, 5. Bilanz, 6. Fossile, 7. Verbrauch Status, 8. Kennzahlen):

  - print first 30 rows with column A label
  - for each label-bearing row, sample what formula or value is in cols
    F..K (where the value or external ref typically lives)
  - count formulas with [D.xlsx] external reference
  - count formulas with internal cell references
  - count plain values (no formula)

Also dumps zip-level external link metadata for both _S.xlsx and D.xlsx.

Run from repo root:
    python scripts/probe_s_xlsx_formulas.py > scripts/audit_out/s_xlsx_probe.txt
"""
from __future__ import annotations

import glob
import os
import re
import warnings
import xml.etree.ElementTree as ET
import zipfile

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

EXT_RE = re.compile(r"\[(\d+)\]")  # matches [1], [2], ... in formula text


def dump_external_links(path: str) -> None:
    print(f"\n--- {os.path.basename(path)} external link targets ---")
    try:
        with zipfile.ZipFile(path) as z:
            # The .rels file maps externalLink<n>.xml IDs to actual filenames
            rels_files = [n for n in z.namelist() if n.startswith("xl/externalLinks/_rels/") and n.endswith(".rels")]
            for rels in sorted(rels_files):
                content = z.read(rels).decode("utf-8", errors="ignore")
                root = ET.fromstring(content)
                for rel in root:
                    target = rel.attrib.get("Target", "")
                    print(f"  {os.path.basename(rels):40s}  Target={target}")
    except Exception as e:
        print(f"  ! {e}")


def probe_sheet(wb_formulas, sheet_name: str, max_rows: int = 30) -> None:
    print(f"\n--- _S.xlsx sheet: {sheet_name} ---")
    ws = wb_formulas[sheet_name]
    formula_count_external = 0
    formula_count_internal = 0
    plain_value_count = 0
    empty_count = 0
    sampled_external_refs: set[str] = set()

    for row_idx in range(1, min(ws.max_row, max_rows) + 1):
        label = ws.cell(row=row_idx, column=1).value
        if label is None:
            continue
        cells_sample = []
        for col_idx in range(2, min(ws.max_column, 12) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            v = cell.value
            if v is None:
                continue
            cells_sample.append(f"col{col_idx}={str(v)[:50]}")
            if len(cells_sample) >= 4:
                break
        label_short = str(label)[:50]
        print(f"  row {row_idx:3d}  A={label_short!r:55s}  {' | '.join(cells_sample)}")

    # Now scan ALL rows for formula-vs-value distribution
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                empty_count += 1
            elif isinstance(v, str) and v.startswith("="):
                if "[" in v and "]" in v:
                    formula_count_external += 1
                    # Extract sample reference
                    m = EXT_RE.search(v)
                    if m and len(sampled_external_refs) < 10:
                        sampled_external_refs.add(v[:80])
                else:
                    formula_count_internal += 1
            else:
                plain_value_count += 1

    print(f"  -- TOTALS for sheet {sheet_name}:")
    print(f"     formulas with external [.] ref:  {formula_count_external}")
    print(f"     formulas internal (no [.] ref):  {formula_count_internal}")
    print(f"     plain values:                    {plain_value_count}")
    print(f"     empty cells:                     {empty_count}")
    if sampled_external_refs:
        print(f"  -- Sample external-ref formulas:")
        for f in list(sampled_external_refs)[:6]:
            print(f"     {f}")


def main() -> None:
    s_path = glob.glob("docs/100prosim_d_*/_S.xlsx")
    d_path = glob.glob("docs/100prosim_d_*/D.xlsx")
    if not s_path or not d_path:
        print("Workbooks not found")
        return
    s_path = s_path[0]
    d_path = d_path[0]

    dump_external_links(s_path)
    dump_external_links(d_path)

    print(f"\nLoading {s_path} with formulas (data_only=False)...")
    wb = load_workbook(s_path, data_only=False, keep_vba=False)

    target_sheets = [
        "1. Flächen",
        "2. Erneuerbare",
        "3. Bedarfsniveau",
        "4. Verbrauch",
        "5. Bilanz",
        "6. Fossile",
        "7. Verbrauch Status",
        "8. Kennzahlen",
    ]
    for s in target_sheets:
        if s in wb.sheetnames:
            probe_sheet(wb, s)
        else:
            print(f"\n!!! sheet {s!r} not found in _S.xlsx (have {wb.sheetnames})")


if __name__ == "__main__":
    main()
