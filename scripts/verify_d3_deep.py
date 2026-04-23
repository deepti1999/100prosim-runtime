"""Deep dive into D3 percent-share formulas — what are AE19/AE25/AE31/E13
for each source? And find Abgleichdifferenz formula."""
import glob, warnings
from openpyxl import load_workbook
warnings.filterwarnings("ignore")

WS = glob.glob("docs/100prosim_d_*/WS.xlsm")[0]
WS_V = load_workbook(WS, data_only=True, keep_vba=True)
WS_F = load_workbook(WS, data_only=False, keep_vba=True)

print("=" * 70)
print("D3 - percent share numerators + denominator")
print("=" * 70)

ws_f = WS_F["1.Jahresbilanz_Strom"]
ws_v = WS_V["1.Jahresbilanz_Strom"]

# Columns E and AE rows 12-33 — see the relationship between raw and adjusted
print("\nRaw (col E) vs adjusted (col AE), rows 12-33:")
print(f"{'row':>4} {'E_formula':45s} {'E_val':>14s} {'AE_formula':45s} {'AE_val':>14s}")
for r in range(12, 34):
    e_f = ws_f[f"E{r}"].value
    e_v = ws_v[f"E{r}"].value
    ae_f = ws_f[f"AE{r}"].value
    ae_v = ws_v[f"AE{r}"].value
    if e_f is None and ae_f is None: continue
    print(f"{r:>4} {str(e_f)[:42]!r:45s} {str(e_v)[:14]:>14s} {str(ae_f)[:42]!r:45s} {str(ae_v)[:14]:>14s}")

# Percent-share formulas for all 4 sources — find them all
print("\n\nAll percent-share formulas in column E (looking for =.../(...) patterns):")
for r in range(10, 40):
    f = ws_f[f"E{r}"].value
    v = ws_v[f"E{r}"].value
    if isinstance(f, str) and "/" in f and isinstance(v, (int, float)) and 0 < v < 1:
        print(f"  E{r}: {f!r:55s} -> {v:.4f}")

# Also row labels in col G/H/I
print("\nLabels in col G rows 12-34 (context):")
for r in range(12, 35):
    v = ws_v[f"G{r}"].value
    if v is not None:
        print(f"  G{r}: {str(v)[:70]!r}")

# Find AE column name-row (what IS AE?)
print("\nHeader row for AE column (rows 14-18):")
for r in range(14, 19):
    v = ws_v[f"AE{r}"].value
    if v is not None:
        print(f"  AE{r}: {str(v)[:80]!r}")
print("\nHeader at AD (left neighbor) and AF (right neighbor):")
for r in range(14, 19):
    ad = ws_v[f"AD{r}"].value
    af = ws_v[f"AF{r}"].value
    print(f"  AD{r}: {str(ad)[:40]!r:45s}  AF{r}: {str(af)[:40]!r:45s}")

# Abgleichdifferenz — the 160 GWh
print("\n\n" + "=" * 70)
print("D4c - Abgleichdifferenz formula")
print("=" * 70)
print("\nLabel at Q45 says 'Abgleichdifferenz'. Check surrounding cells for value + formula:")
for col in "OPQRS":
    for r in range(40, 50):
        f = ws_f[f"{col}{r}"].value
        v = ws_v[f"{col}{r}"].value
        if f is not None or v is not None:
            print(f"  {col}{r}: formula={str(f)[:55]!r:60s} value={str(v)[:20]!r}")
