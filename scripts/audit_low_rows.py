"""Diagnose why 317 rows are LOW confidence in value-match audit.

Possibilities:
 (a) Tolerance too tight — loosen to 5% and re-match
 (b) Our status/ziel are zero (nothing to match)
 (c) D.xlsx row is a "ZIEL-Ansatz" labeled row (one row per status OR ziel)
 (d) Our rows are computed/derived, not D.xlsx-sourced (origin='internal')
 (e) Unit/scale mismatch (ha vs m², GWh vs TWh)
"""
import json, glob, warnings, re
from openpyxl import load_workbook
from collections import Counter
warnings.filterwarnings('ignore')

with open('seed/sqlite_seed.json', encoding='utf-8') as f:
    seed = json.load(f)

D = load_workbook(glob.glob('docs/100prosim_d_*/D.xlsx')[0], data_only=True)
ws = D['1.']
SCEN_COLS = [('U', 21), ('V', 22), ('W', 23), ('AG', 33), ('AN', 40)]

excel_vals = []  # list of (row, col, value, label)
for r in range(1, ws.max_row + 1):
    label_e = ws.cell(row=r, column=5).value
    if not isinstance(label_e, str):
        continue
    for name, c in SCEN_COLS:
        v = ws.cell(row=r, column=c).value
        if isinstance(v, (int, float)):
            excel_vals.append((r, name, float(v), label_e.strip()[:80]))

print(f'D.xlsx: {len(excel_vals)} numeric cells across {len({e[0] for e in excel_vals})} rows')
print()


def close(a, b, rel=0.05):
    if a is None or b is None or not isinstance(a, (int, float)) or not isinstance(b, (int, float)):
        return False
    if a == 0 and b == 0: return True
    denom = max(abs(a), abs(b))
    if denom == 0: return False
    return abs(a - b) / denom < rel


def find_any(target, scales=(1, 1000, 0.001, 10000, 0.0001)):
    """Find D.xlsx matches at various scales (handles unit scale differences)."""
    if target is None or not isinstance(target, (int, float)): return []
    hits = []
    for scale in scales:
        scaled = target * scale
        for (r, col, v, lbl) in excel_vals:
            if close(scaled, v):
                hits.append((r, col, v, lbl, scale))
    return hits[:3]  # cap per target


# Categorise LOW rows
MODELS = {
    'simulator.landuse': ('name', 'status_ha', 'target_ha'),
    'simulator.renewabledata': (None, 'status_value', 'target_value'),
    'simulator.verbrauchdata': ('category', 'status', 'ziel'),
    'simulator.gebaeudewaermedata': ('category', 'status', 'ziel'),
}

for model_name, (label_f, status_f, ziel_f) in MODELS.items():
    rows = [r for r in seed if r['model'] == model_name]
    short = model_name.split('.')[-1]
    print(f'=== {short} (n={len(rows)}) ===')

    cat = Counter()
    for r in rows:
        fv = r['fields']
        s = fv.get(status_f)
        z = fv.get(ziel_f)
        # Category rules
        if (s in (None, 0) and z in (None, 0)):
            cat['both_zero_or_null'] += 1
            continue
        # Try loose match at multiple scales
        s_hits = find_any(s) if isinstance(s, (int, float)) and s != 0 else []
        z_hits = find_any(z) if isinstance(z, (int, float)) and z != 0 else []
        if s_hits or z_hits:
            # Check if scale factor needed
            scales_used = {h[4] for h in s_hits + z_hits}
            if scales_used == {1}:
                cat['loose_match_same_scale'] += 1
            else:
                cat[f'loose_match_scaled={scales_used}'] += 1
        else:
            cat['no_match_any_scale'] += 1

    for k, v in cat.most_common():
        print(f'  {k}: {v}')
    print()

# Sample 10 LOW rows per model — show what their values look like
print('=== Sample LOW rows (first 3 per model) ===')
for model_name, (label_f, status_f, ziel_f) in MODELS.items():
    rows = [r for r in seed if r['model'] == model_name]
    short = model_name.split('.')[-1]
    print(f'--- {short} ---')
    shown = 0
    for r in rows:
        fv = r['fields']
        s = fv.get(status_f)
        z = fv.get(ziel_f)
        if (s in (None, 0) and z in (None, 0)):
            continue
        s_hits = find_any(s, scales=(1,)) if isinstance(s, (int, float)) and s != 0 else []
        z_hits = find_any(z, scales=(1,)) if isinstance(z, (int, float)) and z != 0 else []
        if s_hits or z_hits:
            continue
        # This is a LOW row at strict tolerance and same scale
        code = fv.get('code', '')
        label = fv.get(label_f or 'category', '') or fv.get('subcategory', '')
        print(f'  {code} "{label[:40]}" status={s} ziel={z}')
        shown += 1
        if shown >= 3:
            break
    print()
