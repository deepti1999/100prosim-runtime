"""Round-3: read FORMULAS (not just values) from WS.xlsm to find the
Tagesladungen conversion, and look at I_S sheet for scenario parameters."""
from openpyxl import load_workbook
import warnings, glob
warnings.filterwarnings('ignore')

# Load WITH formulas (data_only=False)
WS_F = load_workbook(glob.glob('docs/100prosim_d_*/WS.xlsm')[0], data_only=False, keep_vba=True)
WS_V = load_workbook(glob.glob('docs/100prosim_d_*/WS.xlsm')[0], data_only=True, keep_vba=True)

# Tagesladung "Faktor" cell — J85 on 1.Jahresbilanz_Strom
print('=== WS.xlsm "Faktor Eingabe-Einheit > Tagesladung" cell context ===')
ws_f = WS_F['1.Jahresbilanz_Strom']
ws_v = WS_V['1.Jahresbilanz_Strom']
for col in ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
    for r in range(83, 92):
        coord = f'{col}{r}'
        f_val = ws_f[coord].value
        v_val = ws_v[coord].value
        if f_val is not None or v_val is not None:
            same = '' if f_val == v_val else f' (formula! computed={v_val!r})'
            print(f'  {coord:6s}: {str(f_val)[:90]!r}{same[:60]}')

# Find any cell whose VALUE is around 3000 (the suspected Tagesladung GWh)
print('\n=== Cells with values 2900-3100 (suspected per-Tagesladung GWh size) ===')
for sname in WS_V.sheetnames:
    ws = WS_V[sname]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and 2900 < cell.value < 3100:
                ctx_e = ws.cell(row=cell.row, column=5).value
                ctx_l = ws.cell(row=cell.row, column=12).value
                ctx = str(ctx_e or ctx_l or '')[:50]
                print(f'  {sname:30s} {cell.coordinate:6s} = {cell.value:8.2f}  ctx={ctx!r}')

# What's in cells near "Wert in Ø Tagesladungen" label (D42)?
print('\n=== WS.xlsm 1.Jahresbilanz_Strom — rows 38-46 columns C-F ===')
for r in range(38, 47):
    for col in ['C', 'D', 'E', 'F', 'G']:
        v = ws_v[f'{col}{r}'].value
        if v is not None:
            print(f'  {col}{r}: {str(v)[:60]!r}')

# Now map D.xlsx I_S sheet — the actual scenario parameters
print('\n=== D.xlsx "I_S" — sample non-zero rows in cols E-J ===')
D = load_workbook(glob.glob('docs/100prosim_d_*/D.xlsx')[0], data_only=True)
ws = D['I_S']
n = 0
for r in range(1, ws.max_row+1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(5, 11)]
    has_data = any(v is not None and v != 0 and v != '0' for v in row_vals)
    if has_data:
        # Look for the row label in column D or earlier
        labels = [ws.cell(row=r, column=c).value for c in range(1, 5)]
        labels_s = ' | '.join(str(l)[:25] for l in labels if l is not None)
        vals_s = ' | '.join(str(v)[:25] for v in row_vals if v is not None)
        print(f'  R{r:4d} labels={labels_s[:60]!r} vals={vals_s[:80]!r}')
        n += 1
        if n >= 25: break

# 8.Kennzahlen — likely where the percent shares live
print('\n=== D.xlsx "8.Kennzahlen" sample ===')
ws = D['8.Kennzahlen']
n = 0
for r in range(1, ws.max_row+1):
    label = ws.cell(row=r, column=5).value
    if label and isinstance(label, str) and len(label.strip()) > 3:
        # Sample values in columns G-L
        vals = []
        for c in range(7, 13):
            v = ws.cell(row=r, column=c).value
            if v is not None and v != '':
                vals.append(f'{ws.cell(row=r, column=c).coordinate}={str(v)[:20]}')
        if vals:
            print(f'  R{r:4d} E={str(label)[:40]!r:45s}  {", ".join(vals)[:80]}')
            n += 1
            if n >= 20: break
