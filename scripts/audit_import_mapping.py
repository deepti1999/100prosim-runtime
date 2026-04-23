"""Deep mapping audit — attempt to match our DB parameter rows to D.xlsx rows.

For each parameter-bearing model (LandUse, RenewableData, VerbrauchData,
GebaeudewaermeData), list every row with its code + label + status/target +
unit, then try to find a matching row in D.xlsx sheet '1.' by label-match.
Output CSV per model + gap reports.
"""
import json, glob, warnings, csv, os, re
from openpyxl import load_workbook
warnings.filterwarnings('ignore')

# Load seed data
with open('seed/sqlite_seed.json', encoding='utf-8') as f:
    seed = json.load(f)

# Load D.xlsx sheet 1. into a (row, label, has_value) list
D = load_workbook(glob.glob('docs/100prosim_d_*/D.xlsx')[0], data_only=True)
ws = D['1.']
excel_rows = []
for r in range(1, ws.max_row + 1):
    label_e = ws.cell(row=r, column=5).value
    if not isinstance(label_e, str) or len(label_e.strip()) < 3:
        continue
    sample_val = ws.cell(row=r, column=21).value  # col U — scenario value column
    if not isinstance(sample_val, (int, float)):
        # try next columns
        for c in (22, 23, 33, 40):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)):
                sample_val = v; break
    excel_rows.append({
        'row': r,
        'label': label_e.strip()[:120],
        'sample_value': sample_val if isinstance(sample_val, (int, float)) else None,
    })

print(f'D.xlsx sheet "1." — {len(excel_rows)} non-empty parameter-label rows')
print()

# Helper: normalise German labels for matching
def norm(s):
    if not s: return ''
    s = s.lower()
    s = re.sub(r'[^\wäöüß]', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def match_excel(db_label, max_hits=3):
    """Find best matches for db_label in excel_rows by shared-word overlap."""
    db_words = set(norm(db_label).split())
    if not db_words: return []
    scored = []
    for er in excel_rows:
        ex_words = set(norm(er['label']).split())
        if not ex_words: continue
        overlap = db_words & ex_words
        if overlap:
            # Score: #shared_words / max(db_words, ex_words)
            score = len(overlap) / max(len(db_words), len(ex_words))
            scored.append((score, er, overlap))
    scored.sort(key=lambda x: -x[0])
    return scored[:max_hits]

# For each parameter model, report mapping coverage
os.makedirs('scripts/audit_out', exist_ok=True)

MODELS = {
    'simulator.landuse': ['code', 'name', 'status_ha', 'target_ha'],
    'simulator.renewabledata': ['code', 'category', 'subcategory', 'status_value', 'target_value', 'unit'],
    'simulator.verbrauchdata': ['code', 'category', 'status', 'ziel', 'unit'],
    'simulator.gebaeudewaermedata': ['code', 'category', 'status', 'ziel', 'unit'],
}

summary = {}
for model_name, fields in MODELS.items():
    rows = [r for r in seed if r['model'] == model_name]
    if not rows: continue
    short = model_name.split('.')[-1]
    csv_path = f'scripts/audit_out/mapping_{short}.csv'
    matched = partial = unmatched = 0
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['our_code', 'our_label', 'match_score', 'excel_row', 'excel_label', 'status', 'notes'])
        for r in rows:
            fv = r['fields']
            code = fv.get('code', '')
            # Best label per model
            if short == 'landuse': label = fv.get('name', '')
            elif short == 'renewabledata': label = (fv.get('subcategory') or '') + ' / ' + (fv.get('category') or '')
            elif short == 'verbrauchdata': label = fv.get('category', '')
            elif short == 'gebaeudewaermedata': label = fv.get('category', '')
            else: label = ''
            if not label:
                unmatched += 1
                w.writerow([code, label, '', '', '', 'NO_LABEL', 'our row has no descriptive label'])
                continue
            hits = match_excel(label, max_hits=1)
            if not hits:
                unmatched += 1
                w.writerow([code, label, '0', '', '', 'NO_MATCH', ''])
                continue
            score, er, overlap = hits[0]
            status = 'MATCHED' if score >= 0.6 else ('PARTIAL' if score >= 0.3 else 'WEAK')
            if status == 'MATCHED': matched += 1
            elif status == 'PARTIAL': partial += 1
            else: unmatched += 1
            w.writerow([code, label[:80], f'{score:.2f}', er['row'], er['label'][:80], status, f"overlap: {','.join(sorted(overlap))[:50]}"])
    summary[short] = {'total': len(rows), 'matched': matched, 'partial': partial, 'unmatched': unmatched}
    print(f'{short:25s} total={len(rows):4d}  matched={matched:4d}  partial={partial:4d}  unmatched={unmatched:4d}  -> {csv_path}')

print()
print('Mapping-coverage percent:')
for short, s in summary.items():
    if s['total']:
        ok = (s['matched'] + s['partial']) / s['total'] * 100
        print(f'  {short:25s}  {ok:5.1f}% findable in D.xlsx (matched+partial) / {s["total"]} rows')
