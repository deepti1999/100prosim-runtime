"""Sample the source hyperlinks and parameter comments in D.xlsx."""
from openpyxl import load_workbook
import warnings, glob
warnings.filterwarnings('ignore')

path = glob.glob('docs/100prosim_d_*/D.xlsx')[0]
wb = load_workbook(path, data_only=True)

print(f'=== "9.Quellen" sheet — first 20 hyperlinks (sources) ===')
ws = wb['9.Quellen']
n = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.hyperlink and n < 20:
            target = cell.hyperlink.target or cell.hyperlink.location
            display = cell.value if cell.value else ''
            print(f'  {cell.coordinate:5s} display={str(display)[:50]!r:55s} link={target[:70] if target else None!r}')
            n += 1
    if n >= 20: break

print(f'\n=== Sheet "1." — first 15 comments (parameter assumptions) ===')
ws = wb['1.']
n = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.comment and n < 15:
            txt = cell.comment.text.replace('\n', ' | ')
            val = str(cell.value)[:25] if cell.value is not None else ''
            print(f'  {cell.coordinate:6s} value={val!r:30s} comment={txt[:120]!r}')
            n += 1
    if n >= 15: break
