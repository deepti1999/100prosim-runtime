"""Value-based mapping audit — resolve which D.xlsx column holds status vs ziel.

For each parameter DB row we have status + ziel (numeric). For each D.xlsx
row we have up to 5 numeric candidates (cols U/V/W/AG/AN).

Method:
 1. For each DB row, find D.xlsx rows where col X is numerically close
    (relative tol 0.5%, abs tol 0.01) to our status OR our ziel.
 2. Combine with label-overlap from audit_import_mapping.
 3. High-confidence = label-match AND value-match on BOTH status+ziel.
 4. Tally column conventions per model: how often does status come from U
    vs W vs AG vs AN? Same for ziel.
 5. Emit per-model CSV with columns:
    our_code, our_label, status_ours, ziel_ours, excel_row, excel_label,
    status_from_col, status_excel_val, ziel_from_col, ziel_excel_val,
    confidence, notes

Output:
 - scripts/audit_out/value_map_<model>.csv  (per-model)
 - scripts/audit_out/column_convention.md   (summary of which col means what)
"""
import json, glob, warnings, csv, os, re
from openpyxl import load_workbook
from collections import Counter
warnings.filterwarnings('ignore')

with open('seed/sqlite_seed.json', encoding='utf-8') as f:
    seed = json.load(f)

D = load_workbook(glob.glob('docs/100prosim_d_*/D.xlsx')[0], data_only=True)
ws = D['1.']

SCEN_COLS = [('U', 21), ('V', 22), ('W', 23), ('AG', 33), ('AN', 40)]

excel_rows = []
for r in range(1, ws.max_row + 1):
    label_e = ws.cell(row=r, column=5).value
    if not isinstance(label_e, str) or len(label_e.strip()) < 3:
        continue
    vals = {}
    for name, c in SCEN_COLS:
        v = ws.cell(row=r, column=c).value
        if isinstance(v, (int, float)):
            vals[name] = float(v)
    excel_rows.append({
        'row': r,
        'label': label_e.strip()[:200],
        'vals': vals,
    })

print(f'D.xlsx sheet "1." — {len(excel_rows)} labelled rows')
print(f'  rows with at least one numeric: {sum(1 for e in excel_rows if e["vals"])}')
print()


def norm(s):
    if not s:
        return ''
    s = s.lower()
    s = re.sub(r'[^\wäöüß]', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


def close(a, b, rel=0.005, abs_tol=0.01):
    if a is None or b is None:
        return False
    if a == 0 and b == 0:
        return True
    if abs(a - b) < abs_tol:
        return True
    denom = max(abs(a), abs(b))
    if denom == 0:
        return False
    return abs(a - b) / denom < rel


def find_value_matches(target, excel_rows):
    """Return list of (excel_row_dict, col_name) where col_name's value ≈ target."""
    if target is None:
        return []
    hits = []
    for er in excel_rows:
        for col_name, v in er['vals'].items():
            if close(target, v):
                hits.append((er, col_name))
    return hits


def find_best_row(status_ours, ziel_ours, label_ours, excel_rows):
    """Find the single best D.xlsx row — must match on value + label overlap.

    Return: (excel_row, status_col, ziel_col, confidence, notes)
    """
    status_hits = find_value_matches(status_ours, excel_rows)
    ziel_hits = find_value_matches(ziel_ours, excel_rows)
    db_words = set(norm(label_ours).split())

    # Index by excel row id
    rows_with_status = {id(er): col for er, col in status_hits}
    rows_with_ziel = {id(er): col for er, col in ziel_hits}

    # Prefer rows that match BOTH status and ziel
    both_rows = set(rows_with_status.keys()) & set(rows_with_ziel.keys())
    if both_rows:
        best = None
        best_score = -1
        for er in excel_rows:
            if id(er) not in both_rows:
                continue
            ex_words = set(norm(er['label']).split())
            overlap = db_words & ex_words
            score = len(overlap) / max(len(db_words) or 1, len(ex_words) or 1) if ex_words else 0
            if score > best_score:
                best_score = score
                best = er
        if best:
            return (best, rows_with_status[id(best)], rows_with_ziel[id(best)],
                    'HIGH' if best_score >= 0.2 else 'MED-VALUE-ONLY',
                    f'both-match; label_overlap={best_score:.2f}')

    # Status only
    if status_hits:
        best = None
        best_score = -1
        for er, col in status_hits:
            ex_words = set(norm(er['label']).split())
            overlap = db_words & ex_words
            score = len(overlap) / max(len(db_words) or 1, len(ex_words) or 1) if ex_words else 0
            if score > best_score:
                best_score = score
                best = (er, col)
        if best and best_score >= 0.2:
            return (best[0], best[1], None, 'MED',
                    f'status-only; label_overlap={best_score:.2f}')

    # Ziel only
    if ziel_hits:
        best = None
        best_score = -1
        for er, col in ziel_hits:
            ex_words = set(norm(er['label']).split())
            overlap = db_words & ex_words
            score = len(overlap) / max(len(db_words) or 1, len(ex_words) or 1) if ex_words else 0
            if score > best_score:
                best_score = score
                best = (er, col)
        if best and best_score >= 0.2:
            return (best[0], None, best[1], 'MED',
                    f'ziel-only; label_overlap={best_score:.2f}')

    return (None, None, None, 'LOW', 'no value+label match')


MODELS = {
    'simulator.landuse': ('code', 'name', 'status_ha', 'target_ha'),
    'simulator.renewabledata': ('code', None, 'status_value', 'target_value'),
    'simulator.verbrauchdata': ('code', 'category', 'status', 'ziel'),
    'simulator.gebaeudewaermedata': ('code', 'category', 'status', 'ziel'),
}

os.makedirs('scripts/audit_out', exist_ok=True)

conv_report = []

for model_name, (code_f, label_f, status_f, ziel_f) in MODELS.items():
    rows = [r for r in seed if r['model'] == model_name]
    short = model_name.split('.')[-1]
    csv_path = f'scripts/audit_out/value_map_{short}.csv'

    status_cols = Counter()
    ziel_cols = Counter()
    conf_counts = Counter()

    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['our_code', 'our_label', 'status_ours', 'ziel_ours',
                    'excel_row', 'excel_label', 'status_col', 'status_excel',
                    'ziel_col', 'ziel_excel', 'confidence', 'notes'])
        for r in rows:
            fv = r['fields']
            code = fv.get(code_f, '')
            if short == 'renewabledata':
                label = (fv.get('subcategory') or '') + ' / ' + (fv.get('category') or '')
            else:
                label = fv.get(label_f or 'category', '') or ''
            status_ours = fv.get(status_f)
            ziel_ours = fv.get(ziel_f)

            er, status_col, ziel_col, confidence, notes = find_best_row(
                status_ours, ziel_ours, label, excel_rows)

            conf_counts[confidence] += 1
            if status_col:
                status_cols[status_col] += 1
            if ziel_col:
                ziel_cols[ziel_col] += 1

            if er:
                w.writerow([code, label[:80], status_ours, ziel_ours,
                            er['row'], er['label'][:80],
                            status_col, er['vals'].get(status_col) if status_col else '',
                            ziel_col, er['vals'].get(ziel_col) if ziel_col else '',
                            confidence, notes])
            else:
                w.writerow([code, label[:80], status_ours, ziel_ours,
                            '', '', '', '', '', '', confidence, notes])

    total = len(rows)
    print(f'=== {short} (n={total}) ===')
    print(f'  confidence: {dict(conf_counts)}')
    print(f'  status column source: {dict(status_cols)}')
    print(f'  ziel column source:   {dict(ziel_cols)}')
    print(f'  -> {csv_path}')
    print()
    conv_report.append({
        'model': short,
        'n': total,
        'confidence': dict(conf_counts),
        'status_cols': dict(status_cols),
        'ziel_cols': dict(ziel_cols),
    })

# Write convention summary
md_path = 'scripts/audit_out/column_convention.md'
with open(md_path, 'w', encoding='utf-8') as f:
    f.write('# D.xlsx column convention (value-match derived)\n\n')
    f.write('For each model, which column of D.xlsx supplied a value matching '
            'our DB `status` / `ziel` fields. Higher count = more reliable convention.\n\n')
    for c in conv_report:
        f.write(f'## {c["model"]} (n={c["n"]})\n\n')
        f.write(f'- confidence: `{c["confidence"]}`\n')
        f.write(f'- status sourced from: `{c["status_cols"]}`\n')
        f.write(f'- ziel sourced from:   `{c["ziel_cols"]}`\n\n')
print(f'Convention summary -> {md_path}')
