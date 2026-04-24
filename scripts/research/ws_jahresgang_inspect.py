"""Extract Jahresgang Strom sheet + any text-box/shape containing '87' from WS.xlsm.

For Q4 — determine whether '87' label on the Gasspeicher flow diagram is a
formula result, typed-in number, or text-box (drawing shape).
"""
import glob
import zipfile
import warnings
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

WS = glob.glob("docs/100prosim_d_*/WS.xlsm")[0]
print(f"Path: {WS}\n")

WS_V = load_workbook(WS, data_only=True, keep_vba=True)
WS_F = load_workbook(WS, data_only=False, keep_vba=True)
print("Sheets:")
for sn in WS_V.sheetnames:
    print(f"  - {sn!r}")
print()

# =========================================================
# Q4a — Look at sheets whose name starts with "2." (Jahresgang Strom)
# =========================================================
candidates = [sn for sn in WS_V.sheetnames if sn.startswith("2.") or "jahresgang" in sn.lower()]
print(f"'2.' sheets: {candidates}")

for sn in candidates:
    ws_v = WS_V[sn]
    ws_f = WS_F[sn]
    print(f"\n=== Sheet: {sn!r} ===")
    # Find all cells with value ~= 87
    hits = []
    for row in ws_v.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, (int, float)) and abs(v - 87) < 0.5:
                f = ws_f[cell.coordinate].value
                is_f = isinstance(f, str) and f.startswith("=")
                # Label in columns 1-6 of the same row
                label = None
                for c in range(1, 7):
                    lv = ws_v.cell(row=cell.row, column=c).value
                    if isinstance(lv, str) and len(lv) > 2:
                        label = lv
                        break
                hits.append((cell.coordinate, v, is_f, f, label))
    if hits:
        print(f"  87-like cells found: {len(hits)}")
        for c, v, isf, f, label in hits:
            print(f"    {c:6s} = {v:.2f} [{'FORMULA' if isf else 'LITERAL'}]  label={label!r}")
            if isf:
                print(f"      formula: {f}")
    else:
        print("  (no 87-like numeric cells)")

# =========================================================
# Q4b — Look inside xl/drawings/ for text-box strings containing '87'
# =========================================================
print("\n" + "=" * 70)
print("Q4b — text-boxes / shapes containing '87' (flow diagram overlays)")
print("=" * 70)
with zipfile.ZipFile(WS) as z:
    drawing_members = [m for m in z.namelist() if "drawing" in m.lower() and m.endswith(".xml")]
    print(f"Drawing xml files: {len(drawing_members)}")
    for m in drawing_members:
        data = z.read(m).decode("utf-8", errors="replace")
        # Look for isolated '87' numeric text
        # XML strings for shapes: <a:t>87</a:t> or similar
        import re
        # Find every <a:t> text node and print those that are exactly '87'
        for match in re.finditer(r'<a:t[^>]*>([^<]+)</a:t>', data):
            text = match.group(1).strip()
            if text == "87" or text == "87,0" or text == "87.0":
                start = max(0, match.start() - 400)
                end = min(len(data), match.end() + 200)
                # Show surrounding context
                context = data[start:end].replace('\n', ' ')[:400]
                print(f"  {m}: {text!r}")
                print(f"    ctx: …{context[:300]}…")
                print()

print("\n" + "=" * 70)
print("Q4c — Also search for '87' in shared strings / sheets via raw xl/")
print("=" * 70)
with zipfile.ZipFile(WS) as z:
    try:
        ss = z.read("xl/sharedStrings.xml").decode("utf-8", errors="replace")
        # Count exact "87" strings
        import re
        count_87 = len(re.findall(r'>87</t>', ss))
        print(f"  sharedStrings.xml: '87' appears as isolated text in {count_87} entries")
    except KeyError:
        print("  (no sharedStrings.xml)")
