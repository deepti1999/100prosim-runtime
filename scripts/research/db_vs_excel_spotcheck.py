"""Q9 — Sample 10 random parameter rows from the DB and cross-check against _S.xlsx
and D.xlsx. Report any discrepancies.

Uses Django ORM; must run via `docker compose exec web python scripts/research/...`.
"""
import os
import sys
import glob
import random
import warnings
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

# Django bootstrap
sys.path.insert(0, "/app")
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "landuse_project.settings")
import django
django.setup()

from simulator.models import LandUse, RenewableData, VerbrauchData, GebaeudewaermeData

S_XLSX = glob.glob("/app/docs/100prosim_d_*/_S.xlsx")[0]
D_XLSX = glob.glob("/app/docs/100prosim_d_*/D.xlsx")[0]
print(f"_S.xlsx: {S_XLSX}")
print(f"D.xlsx : {D_XLSX}")

# Load Excel workbooks (values only)
S = load_workbook(S_XLSX, data_only=True)
D = load_workbook(D_XLSX, data_only=True)

print(f"\n_S sheets: {S.sheetnames}")
print(f"D sheets : {D.sheetnames[:10]} ...")

# _S.xlsx sheets are typically 1. Fl�chen, 2. Erneuerbare, 3. Bedarfsniveau, 4. Verbrauch
# D.xlsx has one sheet per parameter group

# =========================================================
# Pick 10 random DB rows across models (no user scope — use _base_manager)
# =========================================================
random.seed(42)

samples = []

# 3 LandUse rows — LandUse uses status_ha + target_ha
for obj in random.sample(list(LandUse.all_objects.filter(owner__username="testsim")), 3):
    samples.append(("LandUse", obj.code, obj.name, float(getattr(obj, 'status_ha', 0) or 0), float(getattr(obj, 'target_ha', 0) or 0)))

# 3 Renewable
for obj in random.sample(list(RenewableData.all_objects.filter(owner__username="testsim")), 3):
    samples.append(("Renewable", obj.code, obj.name, float(obj.status_value or 0), float(obj.target_value or 0)))

# 2 Verbrauch — VerbrauchData uses status + ziel, no 'name'
for obj in random.sample(list(VerbrauchData.all_objects.filter(owner__username="testsim")), 2):
    samples.append(("Verbrauch", obj.code, obj.category, float(obj.status or 0), float(obj.ziel or 0)))

# 2 Gebaeudewaerme — singleton per code/region (no owner scope)
for obj in random.sample(list(GebaeudewaermeData.all_objects.all()), 2):
    samples.append(("GW", obj.code, obj.category, float(obj.status or 0), float(obj.ziel or 0)))

print(f"\n{len(samples)} samples drawn from DB\n")

# =========================================================
# Cross-check each sample against _S.xlsx and D.xlsx
# =========================================================

def find_in_workbook(wb, code):
    """Scan every sheet for a cell whose value matches the code — return (sheet, cell, row_values)."""
    matches = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() == code:
                    # Grab whole row for context
                    row_vals = [c.value for c in ws[cell.row]]
                    matches.append((sn, cell.coordinate, row_vals))
                    if len(matches) >= 3:
                        return matches
    return matches

print("=" * 90)
print(f"{'Model':10s} {'Code':12s} {'DB_status':>12s} {'DB_target':>12s} | _S match? | D match?")
print("=" * 90)

results = []
for model, code, name, db_status, db_target in samples:
    s_hits = find_in_workbook(S, code)
    d_hits = find_in_workbook(D, code)
    s_str = f"{len(s_hits)} hits" if s_hits else "NO"
    d_str = f"{len(d_hits)} hits" if d_hits else "NO"
    print(f"{model:10s} {code:12s} {db_status:12.2f} {db_target:12.2f} | {s_str:10s} | {d_str}")
    results.append((model, code, name, db_status, db_target, s_hits, d_hits))

# =========================================================
# For each result with at least one _S match, try to pull status + target numeric from the row
# =========================================================
print("\n" + "=" * 90)
print("Detailed numeric comparison for rows with _S.xlsx matches")
print("=" * 90)

for model, code, name, db_status, db_target, s_hits, d_hits in results:
    if not s_hits:
        print(f"\n{model} {code}: no _S match — skipping")
        continue
    sn, addr, row_vals = s_hits[0]
    # Extract numeric values from the row (excluding the code itself)
    nums = [v for v in row_vals if isinstance(v, (int, float))]
    print(f"\n{model} {code} ({name[:40]!r}) on _S!{sn}!{addr}")
    print(f"  DB: status={db_status:.2f}  target={db_target:.2f}")
    print(f"  _S row numerics (first 10): {nums[:10]}")
    # Look for closest match within 5% tolerance
    for n in nums:
        if db_status != 0 and abs(n - db_status) / abs(db_status) < 0.05:
            print(f"    -> status matches ~{n:.2f} (diff {(n-db_status)/db_status*100:.1f}%)")
        if db_target != 0 and abs(n - db_target) / abs(db_target) < 0.05:
            print(f"    -> target matches ~{n:.2f} (diff {(n-db_target)/db_target*100:.1f}%)")

print("\n" + "=" * 90)
print("Detailed numeric comparison for rows with D.xlsx matches")
print("=" * 90)

for model, code, name, db_status, db_target, s_hits, d_hits in results:
    if not d_hits:
        print(f"\n{model} {code}: no D match — skipping")
        continue
    sn, addr, row_vals = d_hits[0]
    nums = [v for v in row_vals if isinstance(v, (int, float))]
    print(f"\n{model} {code} on D!{sn}!{addr}")
    print(f"  DB: status={db_status:.2f}  target={db_target:.2f}")
    print(f"  D row numerics (first 10): {nums[:10]}")
    for n in nums:
        if db_status != 0 and abs(n - db_status) / abs(db_status) < 0.05:
            print(f"    -> status matches ~{n:.2f}")
        if db_target != 0 and abs(n - db_target) / abs(db_target) < 0.05:
            print(f"    -> target matches ~{n:.2f}")
