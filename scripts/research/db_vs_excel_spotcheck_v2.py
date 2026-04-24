"""Q9 v2 — Match DB rows to _S.xlsx by NAME (not code), then compare numeric values."""
import os, sys, glob, warnings
from openpyxl import load_workbook

warnings.filterwarnings("ignore")
sys.path.insert(0, "/app")
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "landuse_project.settings")
import django
django.setup()

from simulator.models import LandUse, RenewableData, VerbrauchData, GebaeudewaermeData

S = load_workbook(glob.glob("/app/docs/100prosim_d_*/_S.xlsx")[0], data_only=True)

# Known-important rows (by human-readable name) we can cross-check
samples = []
for code, model in [
    ("LU_1",   LandUse),         # Siedlung
    ("LU_2",   LandUse),         # Landwirtschaftsfläche
    ("LU_2.1", LandUse),         # Solare Freiflächen
    ("LU_6",   LandUse),         # Windparkfläche
    ("9.1.2",  RenewableData),   # PV
    ("9.1.1",  RenewableData),   # Wind
    ("9.1.3",  RenewableData),   # Laufwasser
    ("9.1.4",  RenewableData),   # Bio
    ("1.1.2",  VerbrauchData),   # some verbrauch
    ("2.10",   GebaeudewaermeData), # Endenergieverbrauch GW
]:
    try:
        if model is GebaeudewaermeData:
            o = model.all_objects.get(code=code)
            name = o.category
            status = float(o.status or 0)
            target = float(o.ziel or 0)
        elif model is VerbrauchData:
            qs = model.all_objects.filter(code=code, owner__username="testsim")
            o = qs.first()
            if not o: continue
            name = o.category
            status = float(o.status or 0)
            target = float(o.ziel or 0)
        elif model is LandUse:
            o = model.all_objects.filter(code=code, owner__username="testsim").first()
            if not o: continue
            name = o.name
            status = float(o.status_ha or 0)
            target = float(o.target_ha or 0)
        else:
            o = model.all_objects.filter(code=code, owner__username="testsim").first()
            if not o: continue
            name = o.name
            status = float(o.status_value or 0)
            target = float(o.target_value or 0)
        samples.append((model.__name__, code, name, status, target))
    except Exception as e:
        print(f"Skip {code}: {e}")

print(f"{'Model':18s} {'Code':10s} {'Name':40s} {'Status':>15s} {'Target':>15s}")
print("=" * 110)
for m, c, n, s, t in samples:
    print(f"{m:18s} {c:10s} {(n or '')[:40]:40s} {s:15.2f} {t:15.2f}")

# For each, scan _S.xlsx for the name string — exact or contains match
def name_match_in_sheet(sheet, target_name):
    """Return list of (row_num, row_values) where target_name substring appears in any cell."""
    hits = []
    ws = S[sheet] if isinstance(sheet, str) else sheet
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and target_name.lower()[:20] in cell.value.lower():
                row_vals = [c.value for c in ws[cell.row]]
                hits.append((cell.row, row_vals))
                break
    return hits

print("\n" + "=" * 110)
print("Name-substring matches in _S.xlsx + numeric comparison")
print("=" * 110)

EXPECTED_SHEETS = {
    "LandUse": "1. Flächen",
    "RenewableData": "2. Erneuerbare",
    "VerbrauchData": "4. Verbrauch",
    "GebaeudewaermeData": "3. Bedarfsniveau",
}

for model, code, name, status, target in samples:
    if not name:
        continue
    sheet = EXPECTED_SHEETS.get(model)
    if not sheet or sheet not in S.sheetnames:
        continue
    hits = name_match_in_sheet(sheet, name)
    if not hits:
        # Try short key
        hits = name_match_in_sheet(sheet, name.split()[0] if name.split() else name[:10])
    if not hits:
        print(f"\n  ❌ {model} {code} {name!r} — NO name match in {sheet}")
        continue
    # Take first hit
    row_num, row_vals = hits[0]
    nums = [v for v in row_vals if isinstance(v, (int, float)) and abs(v) > 0.001]
    print(f"\n  ✓ {model} {code} {name!r} → {sheet} row {row_num}")
    print(f"    DB: status={status:.2f}  target={target:.2f}")
    print(f"    _S numerics: {[round(n,2) for n in nums[:12]]}")
    # Check 5% tolerance
    status_match = any(status and abs(n - status) / max(abs(status), 1) < 0.05 for n in nums)
    target_match = any(target and abs(n - target) / max(abs(target), 1) < 0.05 for n in nums)
    verdict_s = "✓" if status_match else "✗"
    verdict_t = "✓" if target_match else "✗"
    print(f"    Match: status {verdict_s}, target {verdict_t}")
